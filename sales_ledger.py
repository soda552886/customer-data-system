# -*- coding: utf-8 -*-
"""銷售總表：成交／簽約／未報／退戶明細，供週報與請佣摘要自動彙總。"""
from __future__ import annotations

import json
import io
import re
import sqlite3
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


RECORD_TYPES = {
    'deal': '成交（已報）',
    'unreported': '未報',
    'signing': '簽約',
    'purchase': '買進',
    'refund': '退換戶',
}

STATUS_ACTIVE = {'deal', 'unreported', 'signing', 'purchase'}

# 案場請佣預設（可於單筆覆寫；各工地可在銷售總表「本案場請佣設定」改寫）
SITE_COMMISSION_DEFAULTS = {
    'libao_duoyi': {
        'rate': 0.0485,
        'payableRatio': 0.97,
        'retentionRatio': 0.03,
        'scheme': 'simple',
        'label': '鐸藝預設：底價×4.85%，本期可請97%，保留款3%',
    },
    '_default': {
        'rate': 0.0485,
        'payableRatio': 0.97,
        'retentionRatio': 0.03,
        'scheme': 'simple',
        'label': '預設：底價×4.85%，本期可請97%，保留款3%',
    },
}

DEDUCTION_LABEL_DEFAULTS = [
    {'key': 'surcharge', 'label': '附加費'},
    {'key': 'applianceGift', 'label': '家電禮券'},
    {'key': 'pickupVoucher', 'label': '提貨券'},
    {'key': 'decoration', 'label': '裝潢'},
    {'key': 'companyLoanInterest', 'label': '公司貸利息'},
]


def default_sales_settings() -> dict:
    return {
        'rate': 0.0485,
        'payableRatio': 0.97,
        'retentionRatio': 0.03,
        'scheme': 'simple',
        'handoverRetention': 0.005,
        'tiers': [
            {'paidPct': 6, 'claimPct': 2},
            {'paidPct': 8, 'claimPct': 3},
        ],
        'deductionLabels': [dict(x) for x in DEDUCTION_LABEL_DEFAULTS],
        'label': '預設：底價×4.85%，本期可請97%，保留款3%',
    }


def _ratio_from_any(val, default=0.0) -> float:
    """接受 0.0485 或 4.85；小於等於 1 視為比例。"""
    if val in (None, ''):
        return float(default or 0)
    try:
        v = float(val)
    except (TypeError, ValueError):
        return float(default or 0)
    if v > 1:
        return v / 100.0
    return v


def _pct_to_ratio(val, default_ratio=0.0) -> float:
    """表單百分比 → 比例。0.5 代表 0.5%。"""
    if val in (None, ''):
        return float(default_ratio or 0)
    try:
        return float(val) / 100.0
    except (TypeError, ValueError):
        return float(default_ratio or 0)


def _ratio_to_pct(ratio, fallback=0.0) -> float:
    if ratio in (None, ''):
        return float(fallback or 0)
    try:
        v = float(ratio)
    except (TypeError, ValueError):
        return float(fallback or 0)
    if v < 0:
        return float(fallback or 0)
    return round(v * 100.0, 4) if v <= 1 else round(v, 4)


def _normalize_tiers(raw) -> list[dict]:
    out = []
    if isinstance(raw, str):
        rows = []
        for line in raw.replace('；', '\n').splitlines():
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.replace('，', ',').replace('/', ',').split(',') if p.strip()]
            if len(parts) >= 2:
                rows.append({'paidPct': parts[0], 'claimPct': parts[1]})
        raw = rows
    if not isinstance(raw, (list, tuple)):
        return [{'paidPct': 6, 'claimPct': 2}, {'paidPct': 8, 'claimPct': 3}]
    for item in raw:
        if not isinstance(item, dict):
            continue
        try:
            paid = float(item.get('paidPct', item.get('paid')))
            claim = float(item.get('claimPct', item.get('claim')))
        except (TypeError, ValueError):
            continue
        out.append({'paidPct': paid, 'claimPct': claim})
    out.sort(key=lambda x: x['paidPct'])
    return out or [{'paidPct': 6, 'claimPct': 2}, {'paidPct': 8, 'claimPct': 3}]


def _normalize_deduction_labels(raw) -> list[dict]:
    by_key = {d['key']: d['label'] for d in DEDUCTION_LABEL_DEFAULTS}
    if isinstance(raw, dict):
        for key, label in raw.items():
            if key in by_key:
                by_key[key] = str(label or '').strip()
    elif isinstance(raw, (list, tuple)):
        seen = set()
        for item in raw:
            if not isinstance(item, dict):
                continue
            key = str(item.get('key') or '').strip()
            if key in by_key:
                by_key[key] = str(item.get('label') or '').strip()
                seen.add(key)
        # 列表有送來的項目才覆寫；未出現的鍵維持預設
        if seen:
            pass
    return [{'key': k, 'label': by_key[k]} for k in (
        'surcharge', 'applianceGift', 'pickupVoucher', 'decoration', 'companyLoanInterest',
    )]


def format_commission_label(settings: dict) -> str:
    rate = _ratio_to_pct(settings.get('rate'), 4.85)
    if (settings.get('scheme') or 'simple') == 'payment_tiers':
        ho = _ratio_to_pct(settings.get('handoverRetention'), 0.5)
        bits = [
            f'繳{t["paidPct"]:g}%可請{t["claimPct"]:g}%'
            for t in (settings.get('tiers') or [])
        ]
        extra = '；'.join(bits) if bits else '依繳款成數解鎖'
        return f'佣金總額{rate:g}%、交屋保留{ho:g}%；{extra}（各期不再另扣保留款）'
    pay = _ratio_to_pct(settings.get('payableRatio'), 97)
    ret = _ratio_to_pct(settings.get('retentionRatio'), 3)
    return f'底價×{rate:g}%，本期可請{pay:g}%，保留款{ret:g}%'


def parse_sales_settings(raw) -> dict:
    base = default_sales_settings()
    data = raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw or '{}')
        except (TypeError, json.JSONDecodeError):
            data = {}
    if not isinstance(data, dict):
        data = {}
    scheme = str(data.get('scheme') or base['scheme']).strip() or 'simple'
    if scheme not in ('simple', 'payment_tiers'):
        scheme = 'simple'
    out = {
        **base,
        'scheme': scheme,
        'rate': _ratio_from_any(data.get('rate'), base['rate']),
        'payableRatio': _ratio_from_any(data.get('payableRatio'), base['payableRatio']),
        'retentionRatio': _ratio_from_any(data.get('retentionRatio'), base['retentionRatio']),
        'handoverRetention': _ratio_from_any(data.get('handoverRetention'), base['handoverRetention']),
        'tiers': _normalize_tiers(data.get('tiers', base['tiers'])),
        'deductionLabels': _normalize_deduction_labels(data.get('deductionLabels', base['deductionLabels'])),
    }
    # 過大的交屋保留（例如誤存 0.5＝50%）改視為百分比 0.5%
    if out['handoverRetention'] > 0.05:
        out['handoverRetention'] = out['handoverRetention'] / 100.0 if out['handoverRetention'] <= 5 else 0.005
    out['label'] = str(data.get('label') or '').strip() or format_commission_label(out)
    return out


def normalize_sales_settings_from_api(body: dict) -> dict:
    base = default_sales_settings()
    data = body if isinstance(body, dict) else {}
    scheme = str(data.get('scheme') or base['scheme']).strip() or 'simple'
    if scheme not in ('simple', 'payment_tiers'):
        scheme = 'simple'
    out = {
        **base,
        'scheme': scheme,
        'rate': _pct_to_ratio(data.get('ratePct'), base['rate']) if data.get('ratePct') not in (None, '') else _ratio_from_any(data.get('rate'), base['rate']),
        'payableRatio': _pct_to_ratio(data.get('payablePct'), base['payableRatio']) if data.get('payablePct') not in (None, '') else _ratio_from_any(data.get('payableRatio'), base['payableRatio']),
        'retentionRatio': _pct_to_ratio(data.get('retentionPct'), base['retentionRatio']) if data.get('retentionPct') not in (None, '') else _ratio_from_any(data.get('retentionRatio'), base['retentionRatio']),
        'handoverRetention': _pct_to_ratio(data.get('handoverPct'), base['handoverRetention']) if data.get('handoverPct') not in (None, '') else _ratio_from_any(data.get('handoverRetention'), base['handoverRetention']),
        'tiers': _normalize_tiers(data.get('tiers') or data.get('tiersText') or base['tiers']),
        'deductionLabels': _normalize_deduction_labels(data.get('deductionLabels')),
    }
    out['label'] = format_commission_label(out)
    return out


def sales_settings_public(settings: dict) -> dict:
    s = parse_sales_settings(settings)
    return {
        **s,
        'ratePct': _ratio_to_pct(s['rate'], 4.85),
        'payablePct': _ratio_to_pct(s['payableRatio'], 97),
        'retentionPct': _ratio_to_pct(s['retentionRatio'], 3),
        'handoverPct': _ratio_to_pct(s['handoverRetention'], 0.5),
        'labels': commission_matrix_labels(s),
    }


def commission_matrix_labels(settings: dict) -> dict:
    s = parse_sales_settings(settings)
    rate = _ratio_to_pct(s['rate'], 4.85)
    if s.get('scheme') == 'payment_tiers':
        ho = _ratio_to_pct(s['handoverRetention'], 0.5)
        return {
            'claimable': f'{rate:g}%總佣',
            'retention': f'交屋保留 {ho:g}%',
            'payable': '已解鎖可請',
        }
    pay = _ratio_to_pct(s['payableRatio'], 97)
    ret = _ratio_to_pct(s['retentionRatio'], 3)
    return {
        'claimable': f'{rate:g}%',
        'retention': f'{ret:g}%保留',
        'payable': f'{pay:g}%可請',
    }


def commission_defaults_for_site(site_id: str, conn: Optional[sqlite3.Connection] = None) -> dict:
    raw = None
    if conn and site_id and site_id != '_default':
        try:
            row = conn.execute('SELECT sales_settings FROM sites WHERE id = ?', (site_id,)).fetchone()
        except sqlite3.OperationalError:
            row = None
        if row:
            try:
                raw = row['sales_settings']
            except (KeyError, IndexError, TypeError):
                raw = None
    if raw:
        return sales_settings_public(raw)
    hard = SITE_COMMISSION_DEFAULTS.get(site_id) or SITE_COMMISSION_DEFAULTS['_default']
    return sales_settings_public(hard)


def _unlocked_claim_pct(paid_pct: float, tiers: list[dict]) -> float:
    unlocked = 0.0
    for item in sorted(tiers or [], key=lambda x: float(x.get('paidPct') or 0)):
        try:
            if paid_pct + 1e-9 >= float(item.get('paidPct') or 0):
                unlocked = float(item.get('claimPct') or 0)
        except (TypeError, ValueError):
            continue
    return unlocked


def _round4(val) -> float:
    return round(float(val or 0), 4)


def _as_ratio(val: float) -> float:
    """接受 0.0485 或 4.85（百分比）寫法。"""
    return _ratio_from_any(val, 0)


def compute_commission(
    *,
    site_id: str,
    base_total: float,
    actual_total: float,
    body: dict,
    settings: Optional[dict] = None,
) -> dict:
    """依底價／成交價自動算請佣。一般案場用本期可請／保留款；繳款成數方案依已繳%解鎖。"""
    defaults = parse_sales_settings(settings or commission_defaults_for_site(site_id))
    mode = str(body.get('commissionBaseMode') or 'base').strip().lower()
    if mode not in ('base', 'deal'):
        mode = 'base'
    rate = _as_ratio(_num(body.get('commissionRate'), defaults['rate']))
    payable_ratio = _as_ratio(_num(body.get('commissionPayableRatio'), defaults['payableRatio']))
    retention_ratio = _as_ratio(_num(body.get('commissionRetentionRatio'), defaults['retentionRatio']))
    deduction = max(_num(body.get('commissionDeduction')), 0)
    scheme = defaults.get('scheme') or 'simple'

    extra = body.get('extra') if isinstance(body.get('extra'), dict) else {}
    if extra.get('customerPaidPct') in (None, '') and body.get('customerPaidPct') not in (None, ''):
        extra = dict(extra)
        extra['customerPaidPct'] = _num(body.get('customerPaidPct'))

    suggested = base_total if mode == 'base' else actual_total
    # 有明確覆寫時用覆寫值（對應 Excel 紅字改拉成交價／特殊金額）
    if body.get('commissionSalesAmount') not in (None, ''):
        sales_amount = _num(body.get('commissionSalesAmount'), suggested)
    else:
        sales_amount = suggested

    period = str(body.get('commissionPeriod') or '').strip()
    claim_date = _parse_date(body.get('commissionClaimDate'))
    is_claimed = bool(period or claim_date)

    if scheme == 'payment_tiers':
        paid_pct = _num(body.get('customerPaidPct'), _num(extra.get('customerPaidPct')))
        unlocked_pct = _unlocked_claim_pct(paid_pct, defaults.get('tiers') or [])
        handover_ratio = float(defaults.get('handoverRetention') or 0.005)
        claimable = max(sales_amount * rate - deduction, 0)
        payable = max(sales_amount * (unlocked_pct / 100.0), 0)
        retention = max(sales_amount * handover_ratio, 0)
        claimed = payable if is_claimed else 0.0
        unclaimed = max(claimable - claimed - retention, 0)
        payable_ratio = (unlocked_pct / 100.0 / rate) if rate else 0
        retention_ratio = (handover_ratio / rate) if rate else 0
    else:
        claimable = max(sales_amount * rate - deduction, 0)
        payable = claimable * payable_ratio
        retention = claimable * retention_ratio
        claimed = payable if is_claimed else 0.0
        unclaimed = max(claimable - claimed, 0)

    # 請佣總表匯入：若檔案已有可請／已請／未請實數，優先採用（期別仍決定已請狀態）
    if extra.get('importClaimable') not in (None, ''):
        claimable = max(_num(extra.get('importClaimable')), 0)
    if extra.get('importPayable') not in (None, ''):
        payable = max(_num(extra.get('importPayable')), 0)
    elif scheme != 'payment_tiers':
        payable = claimable * payable_ratio
    if extra.get('importRetention') not in (None, ''):
        retention = max(_num(extra.get('importRetention')), 0)
    elif scheme != 'payment_tiers':
        # 保留款一律＝可請佣 × 保留比例，避免匯入的 97% 四捨五入讓 3% 對不上
        retention = claimable * retention_ratio
    if extra.get('importClaimed') not in (None, ''):
        claimed = max(_num(extra.get('importClaimed')), 0)
        is_claimed = is_claimed or claimed > 0
    else:
        claimed = payable if is_claimed else 0.0
    if extra.get('importUnclaimed') not in (None, ''):
        unclaimed = max(_num(extra.get('importUnclaimed')), 0)
    elif scheme == 'payment_tiers':
        unclaimed = max(claimable - claimed - retention, 0)
    else:
        unclaimed = max(claimable - claimed, 0)

    return {
        'commission_base_mode': mode,
        'commission_sales_amount': _round4(sales_amount),
        'commission_rate': rate,
        'commission_payable_ratio': payable_ratio,
        'commission_retention_ratio': retention_ratio,
        'commission_deduction': _round4(deduction),
        'commission_claimable': _round4(claimable),
        'commission_payable': _round4(payable),
        'commission_retention': _round4(retention),
        'commission_period': period,
        'commission_claim_date': claim_date,
        'commission_claimed': _round4(claimed),
        'commission_unclaimed': _round4(unclaimed),
        'commission_status': '已請' if is_claimed else '未請',
    }


def init_sales_tables(conn: sqlite3.Connection):
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS sales_deals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id TEXT NOT NULL,
            record_type TEXT NOT NULL DEFAULT 'deal',
            order_no TEXT NOT NULL DEFAULT '',
            unit_no TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            product_type TEXT NOT NULL DEFAULT '',
            area_ping REAL NOT NULL DEFAULT 0,
            parking_count REAL NOT NULL DEFAULT 0,
            parking_nos TEXT NOT NULL DEFAULT '',
            parking_no1 TEXT NOT NULL DEFAULT '',
            parking_no2 TEXT NOT NULL DEFAULT '',
            list_price REAL NOT NULL DEFAULT 0,
            base_price REAL NOT NULL DEFAULT 0,
            total_price REAL NOT NULL DEFAULT 0,
            house_sale_price REAL NOT NULL DEFAULT 0,
            parking_sale_price REAL NOT NULL DEFAULT 0,
            actual_house_price REAL NOT NULL DEFAULT 0,
            actual_total_price REAL NOT NULL DEFAULT 0,
            surcharge REAL NOT NULL DEFAULT 0,
            appliance_gift REAL NOT NULL DEFAULT 0,
            pickup_voucher REAL NOT NULL DEFAULT 0,
            decoration REAL NOT NULL DEFAULT 0,
            company_loan_interest REAL NOT NULL DEFAULT 0,
            house_base_price REAL NOT NULL DEFAULT 0,
            parking_base_price REAL NOT NULL DEFAULT 0,
            excess_price REAL NOT NULL DEFAULT 0,
            units REAL NOT NULL DEFAULT 1,
            deposit_date TEXT,
            supplement_date TEXT,
            sign_date TEXT,
            report_date TEXT,
            owner_sale_report_date TEXT,
            owner_sign_report_date TEXT,
            salesperson1 TEXT NOT NULL DEFAULT '',
            salesperson2 TEXT NOT NULL DEFAULT '',
            is_co_managed INTEGER NOT NULL DEFAULT 0,
            commission_claimable REAL NOT NULL DEFAULT 0,
            commission_claimed REAL NOT NULL DEFAULT 0,
            commission_booked REAL NOT NULL DEFAULT 0,
            next_month_claimable REAL NOT NULL DEFAULT 0,
            next_month_units REAL NOT NULL DEFAULT 0,
            next_month_parking REAL NOT NULL DEFAULT 0,
            customer_id INTEGER,
            memo TEXT NOT NULL DEFAULT '',
            extra TEXT NOT NULL DEFAULT '{}',
            created_by INTEGER,
            updated_by INTEGER,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_sales_deals_site ON sales_deals(site_id);
        CREATE INDEX IF NOT EXISTS idx_sales_deals_type ON sales_deals(site_id, record_type);
        CREATE INDEX IF NOT EXISTS idx_sales_deals_report ON sales_deals(site_id, report_date);
        CREATE INDEX IF NOT EXISTS idx_sales_deals_sign ON sales_deals(site_id, sign_date);
    ''')
    # SQLite CREATE TABLE IF NOT EXISTS 不會替既有表補欄位，逐欄升級。
    migrations = {
        'order_no': "TEXT NOT NULL DEFAULT ''",
        'parking_no1': "TEXT NOT NULL DEFAULT ''",
        'parking_no2': "TEXT NOT NULL DEFAULT ''",
        'house_sale_price': 'REAL NOT NULL DEFAULT 0',
        'parking_sale_price': 'REAL NOT NULL DEFAULT 0',
        'actual_house_price': 'REAL NOT NULL DEFAULT 0',
        'actual_total_price': 'REAL NOT NULL DEFAULT 0',
        'surcharge': 'REAL NOT NULL DEFAULT 0',
        'appliance_gift': 'REAL NOT NULL DEFAULT 0',
        'pickup_voucher': 'REAL NOT NULL DEFAULT 0',
        'decoration': 'REAL NOT NULL DEFAULT 0',
        'company_loan_interest': 'REAL NOT NULL DEFAULT 0',
        'house_base_price': 'REAL NOT NULL DEFAULT 0',
        'parking_base_price': 'REAL NOT NULL DEFAULT 0',
        'excess_price': 'REAL NOT NULL DEFAULT 0',
        'supplement_date': 'TEXT',
        'owner_sale_report_date': 'TEXT',
        'owner_sign_report_date': 'TEXT',
        'commission_base_mode': "TEXT NOT NULL DEFAULT 'base'",
        'commission_sales_amount': 'REAL NOT NULL DEFAULT 0',
        'commission_rate': 'REAL NOT NULL DEFAULT 0.0485',
        'commission_payable_ratio': 'REAL NOT NULL DEFAULT 0.97',
        'commission_retention_ratio': 'REAL NOT NULL DEFAULT 0.03',
        'commission_deduction': 'REAL NOT NULL DEFAULT 0',
        'commission_payable': 'REAL NOT NULL DEFAULT 0',
        'commission_retention': 'REAL NOT NULL DEFAULT 0',
        'commission_period': "TEXT NOT NULL DEFAULT ''",
        'commission_claim_date': 'TEXT',
        'commission_unclaimed': 'REAL NOT NULL DEFAULT 0',
    }
    existing = {
        row['name'] for row in conn.execute('PRAGMA table_info(sales_deals)').fetchall()
    }
    for name, definition in migrations.items():
        if name not in existing:
            conn.execute(f'ALTER TABLE sales_deals ADD COLUMN {name} {definition}')

    conn.executescript('''
        CREATE TABLE IF NOT EXISTS commission_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id TEXT NOT NULL,
            period_name TEXT NOT NULL,
            claim_month TEXT NOT NULL DEFAULT '',
            amount_payable REAL,
            half1_amount REAL NOT NULL DEFAULT 0,
            deposit_date1 TEXT,
            half2_amount REAL NOT NULL DEFAULT 0,
            deposit_date2 TEXT,
            deduction_amount REAL NOT NULL DEFAULT 0,
            deduction_memo TEXT NOT NULL DEFAULT '',
            memo TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            UNIQUE(site_id, period_name)
        );
        CREATE INDEX IF NOT EXISTS idx_commission_batches_site
            ON commission_batches(site_id);
    ''')


def _comm_bucket():
    return {
        'units': 0.0,
        'parking': 0.0,
        'claimable': 0.0,   # 4.85%
        'retention': 0.0,   # 3%
        'payable': 0.0,     # 97%
    }


def _round_bucket(b: dict, *, derive_from_claimable: bool = True, settings: Optional[dict] = None) -> dict:
    cfg = parse_sales_settings(settings) if settings else default_sales_settings()
    payable_ratio = float(cfg.get('payableRatio') or 0.97)
    retention_ratio = float(cfg.get('retentionRatio') or 0.03)
    claimable = round(b['claimable'], 4)
    if (cfg.get('scheme') == 'payment_tiers') or not derive_from_claimable:
        return {
            'units': round(b['units'], 2),
            'parking': round(b['parking'], 2),
            'claimable': claimable,
            'retention': round(b['retention'], 4),
            'payable': round(b['payable'], 4),
        }
    if derive_from_claimable and claimable > 0:
        return {
            'units': round(b['units'], 2),
            'parking': round(b['parking'], 2),
            'claimable': claimable,
            'retention': round(claimable * retention_ratio, 4),
            'payable': round(claimable * payable_ratio, 4),
        }
    payable = round(b['payable'], 4)
    if derive_from_claimable and payable > 0 and claimable <= 0 and payable_ratio:
        claimable = round(payable / payable_ratio, 4)
    return {
        'units': round(b['units'], 2),
        'parking': round(b['parking'], 2),
        'claimable': claimable,
        'retention': round(claimable * retention_ratio, 4) if claimable else round(b['retention'], 4),
        'payable': round(claimable * payable_ratio, 4) if claimable else payable,
    }


def _add_to_bucket(bucket: dict, units, parking, claimable, retention, payable):
    bucket['units'] += units
    bucket['parking'] += parking
    bucket['claimable'] += claimable
    bucket['retention'] += retention
    bucket['payable'] += payable


def _deposit_due(value) -> bool:
    """預計入帳日已到期（含當日）。"""
    parsed = _parse_date(value)
    if not parsed:
        return False
    try:
        return datetime.strptime(parsed[:10], '%Y-%m-%d').date() <= date.today()
    except ValueError:
        return False


def _batch_status(deposit1, deposit2) -> str:
    due1 = _deposit_due(deposit1)
    due2 = _deposit_due(deposit2)
    if due2:
        return 'full'
    if due1:
        return 'partial'
    return 'none'


def sync_commission_batches(conn: sqlite3.Connection, site_id: str):
    """依銷售單的請佣期別自動建立／對齊批次列。"""
    periods = conn.execute(
        '''
        SELECT DISTINCT TRIM(commission_period) AS period_name
        FROM sales_deals
        WHERE site_id = ?
          AND record_type IN ('deal', 'signing', 'purchase', 'unreported')
          AND TRIM(COALESCE(commission_period, '')) != ''
        ''',
        (site_id,),
    ).fetchall()
    for row in periods:
        name = (row['period_name'] or '').strip()
        if not name:
            continue
        exists = conn.execute(
            'SELECT id FROM commission_batches WHERE site_id = ? AND period_name = ?',
            (site_id, name),
        ).fetchone()
        if not exists:
            conn.execute(
                '''
                INSERT INTO commission_batches (site_id, period_name)
                VALUES (?, ?)
                ''',
                (site_id, name),
            )


def _period_deal_totals(conn: sqlite3.Connection, site_id: str, period_name: str) -> dict:
    settings = commission_defaults_for_site(site_id, conn=conn)
    payable_ratio = float(settings.get('payableRatio') or 0.97)
    retention_ratio = float(settings.get('retentionRatio') or 0.03)
    scheme = settings.get('scheme') or 'simple'
    rows = conn.execute(
        '''
        SELECT units, parking_count, commission_payable, commission_claimable,
               commission_retention, commission_booked, commission_claimed,
               commission_period, commission_claim_date,
               unit_no, customer_name, order_no, id
        FROM sales_deals
        WHERE site_id = ?
          AND record_type IN ('deal', 'signing', 'purchase', 'unreported')
          AND TRIM(COALESCE(commission_period, '')) = ?
        ''',
        (site_id, period_name),
    ).fetchall()
    payable = 0.0
    claimable = 0.0
    retention = 0.0
    booked = 0.0
    units = 0.0
    parking = 0.0
    deals = []
    for r in rows:
        p = _num(r['commission_payable'])
        c = _num(r['commission_claimable'])
        if p <= 0 and c > 0 and scheme != 'payment_tiers':
            p = c * payable_ratio
        ret = _num(r['commission_retention'])
        if ret <= 0 and c > 0 and scheme != 'payment_tiers':
            ret = c * retention_ratio
        payable += p
        claimable += c
        retention += ret
        booked += _num(r['commission_booked'])
        units += _num(r['units'], 1)
        parking += _num(r['parking_count'])
        is_claimed = bool(
            _num(r['commission_claimed']) > 0
            or (r['commission_period'] or '').strip()
            or r['commission_claim_date']
        )
        deals.append({
            'id': r['id'],
            'unitNo': r['unit_no'] or '',
            'customerName': r['customer_name'] or '',
            'orderNo': r['order_no'] or '',
            'units': _num(r['units'], 1),
            'parking': _num(r['parking_count']),
            'payable': _round4(p),
            'claimable': _round4(c),
            'status': '已請' if is_claimed else '未請',
        })
    return {
        'units': round(units, 2),
        'parking': round(parking, 2),
        'payable': _round4(payable),
        'claimable': _round4(claimable),
        'retention': _round4(retention),
        'booked': _round4(booked),
        'dealCount': len(deals),
        'deals': deals,
    }


def list_commission_batches(conn: sqlite3.Connection, site_id: str) -> list[dict]:
    sync_commission_batches(conn, site_id)
    conn.commit()
    rows = conn.execute(
        '''
        SELECT * FROM commission_batches
        WHERE site_id = ?
        ORDER BY claim_month ASC, period_name ASC, id ASC
        ''',
        (site_id,),
    ).fetchall()
    out = []
    for row in rows:
        totals = _period_deal_totals(conn, site_id, row['period_name'])
        auto_payable = totals['payable']
        amount = row['amount_payable']
        if amount is None:
            amount = auto_payable
        else:
            amount = _num(amount)
        half1 = _num(row['half1_amount'])
        half2 = _num(row['half2_amount'])
        if half1 <= 0 and half2 <= 0 and amount > 0:
            half1 = _round4(amount / 2)
            half2 = _round4(amount - half1)
        claim_month = str(row['claim_month'] or '').strip()
        if not claim_month:
            inferred = _infer_period_claim_month(conn, site_id, row['period_name'])
            if inferred:
                claim_month = inferred
                conn.execute(
                    'UPDATE commission_batches SET claim_month = ? WHERE id = ?',
                    (inferred, row['id']),
                )
        status = _batch_status(row['deposit_date1'], row['deposit_date2'])
        booked_total = 0.0
        if _deposit_due(row['deposit_date1']):
            booked_total += half1
        if _deposit_due(row['deposit_date2']):
            booked_total += half2
        out.append({
            'id': row['id'],
            'siteId': row['site_id'],
            'periodName': row['period_name'],
            'claimMonth': claim_month,
            'amountPayable': _round4(amount),
            'autoPayable': auto_payable,
            'half1Amount': _round4(half1),
            'depositDate1': row['deposit_date1'],
            'half2Amount': _round4(half2),
            'depositDate2': row['deposit_date2'],
            'deductionAmount': _round4(row['deduction_amount']),
            'deductionMemo': row['deduction_memo'] or '',
            'memo': row['memo'] or '',
            'status': status,
            'bookedTotal': _round4(booked_total),
            'claimable': totals['claimable'],
            'retention': totals['retention'],
            'units': totals['units'],
            'parking': totals['parking'],
            'dealCount': totals['dealCount'],
            'deals': totals['deals'],
            'updatedAt': row['updated_at'],
        })
    conn.commit()
    return out


def upsert_commission_batch(conn: sqlite3.Connection, site_id: str, body: dict) -> int:
    period_name = str(body.get('periodName') or body.get('period_name') or '').strip()
    if not period_name:
        raise ValueError('請填期別名稱')
    claim_month = str(body.get('claimMonth') or body.get('claim_month') or '').strip()
    amount_raw = body.get('amountPayable', body.get('amount_payable'))
    amount_payable = None if amount_raw in (None, '') else _num(amount_raw)
    half1 = _num(body.get('half1Amount', body.get('half1_amount')))
    half2 = _num(body.get('half2Amount', body.get('half2_amount')))
    deposit1 = _parse_date(body.get('depositDate1', body.get('deposit_date1')))
    deposit2 = _parse_date(body.get('depositDate2', body.get('deposit_date2')))
    deduction_amount = _num(body.get('deductionAmount', body.get('deduction_amount')))
    deduction_memo = str(body.get('deductionMemo') or body.get('deduction_memo') or '').strip()
    memo = str(body.get('memo') or '').strip()
    batch_id = body.get('id')

    if batch_id:
        conn.execute(
            '''
            UPDATE commission_batches SET
              period_name=?, claim_month=?, amount_payable=?,
              half1_amount=?, deposit_date1=?, half2_amount=?, deposit_date2=?,
              deduction_amount=?, deduction_memo=?, memo=?,
              updated_at=datetime('now', 'localtime')
            WHERE id=? AND site_id=?
            ''',
            (
                period_name, claim_month, amount_payable,
                half1, deposit1, half2, deposit2,
                deduction_amount, deduction_memo, memo,
                int(batch_id), site_id,
            ),
        )
        return int(batch_id)

    existing = conn.execute(
        'SELECT id FROM commission_batches WHERE site_id=? AND period_name=?',
        (site_id, period_name),
    ).fetchone()
    if existing:
        conn.execute(
            '''
            UPDATE commission_batches SET
              claim_month=?, amount_payable=?,
              half1_amount=?, deposit_date1=?, half2_amount=?, deposit_date2=?,
              deduction_amount=?, deduction_memo=?, memo=?,
              updated_at=datetime('now', 'localtime')
            WHERE id=?
            ''',
            (
                claim_month, amount_payable,
                half1, deposit1, half2, deposit2,
                deduction_amount, deduction_memo, memo,
                existing['id'],
            ),
        )
        return int(existing['id'])

    cur = conn.execute(
        '''
        INSERT INTO commission_batches (
          site_id, period_name, claim_month, amount_payable,
          half1_amount, deposit_date1, half2_amount, deposit_date2,
          deduction_amount, deduction_memo, memo
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            site_id, period_name, claim_month, amount_payable,
            half1, deposit1, half2, deposit2,
            deduction_amount, deduction_memo, memo,
        ),
    )
    return int(cur.lastrowid)


def delete_commission_batch(conn: sqlite3.Connection, site_id: str, batch_id: int) -> bool:
    cur = conn.execute(
        'DELETE FROM commission_batches WHERE id=? AND site_id=?',
        (batch_id, site_id),
    )
    return cur.rowcount > 0


def build_commission_overview_excel(site_name: str, matrix: dict, batches: list[dict]) -> bytes:
    """匯出請佣摘要＋期別服務費兩張表。"""
    wb = Workbook()
    font_name = '微軟正黑體'
    header_fill = PatternFill('solid', fgColor='1A4D7C')
    section_fill = PatternFill('solid', fgColor='D6EAF8')
    full_fill = PatternFill('solid', fgColor='D5F5E3')
    partial_fill = PatternFill('solid', fgColor='FCF3CF')

    def style_header(ws, row):
        for cell in ws[row]:
            cell.font = Font(name=font_name, bold=True, color='FFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # —— 請佣摘要 ——
    ws = wb.active
    ws.title = '請佣摘要'
    ws.append([f'{site_name}　請佣總覽（可請／已請／未請）'])
    ws['A1'].font = Font(name=font_name, bold=True, size=16, color='1A4D7C')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.append(['區塊', '戶／車', '4.85%可請(萬)', '3%保留款(萬)', '97%本期可請(萬)'])
    style_header(ws, 2)
    for key, label in (
        ('claimable', '可請總金額'),
        ('claimed', '已請款金額'),
        ('unclaimed', '未請款總金額'),
        ('forecast', '預計本月可請'),
    ):
        b = matrix.get(key) or {}
        ws.append([
            label,
            f"{b.get('units', 0)}戶／{b.get('parking', 0)}車",
            b.get('claimable', 0),
            b.get('retention', 0),
            b.get('payable', 0),
        ])
    ws.append([])
    ws.append(['已入帳合計(萬)', (matrix.get('totals') or {}).get('bookedAmount', 0)])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # —— 期別服務費 ——
    ws = wb.create_sheet('期別服務費')
    ws.append([f'{site_name}　期別服務費（拆半入帳）'])
    ws['A1'].font = Font(name=font_name, bold=True, size=16, color='1A4D7C')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.append([
        '請款期別', '請款月份', '戶／車', '服務費97%(萬)',
        '拆半50%(萬)', '入帳日1', '拆半50%(萬)', '入帳日2',
        '墊款／折讓', '備註',
    ])
    style_header(ws, 2)
    for b in batches:
        ws.append([
            b.get('periodName'),
            b.get('claimMonth'),
            f"{b.get('units', 0)}戶／{b.get('parking', 0)}車",
            b.get('amountPayable'),
            b.get('half1Amount'),
            b.get('depositDate1') or '',
            b.get('half2Amount'),
            b.get('depositDate2') or '',
            b.get('deductionMemo') or (
                f"墊／折 {_round4(b.get('deductionAmount') or 0)} 萬"
                if b.get('deductionAmount') else ''
            ),
            b.get('memo') or '',
        ])
        fill = None
        if b.get('status') == 'full':
            fill = full_fill
        elif b.get('status') == 'partial':
            fill = partial_fill
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.font = Font(name=font_name, size=11)
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['I'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_deal_date(value) -> Optional[str]:
    """對外日期解析（民國年 113/12/12 → 2024-12-12）。"""
    return _parse_date(value)


def _parse_date(value) -> Optional[str]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s or s.lower() in ('none', 'nat', 'nan'):
        return None
    s = s.replace('年', '-').replace('月', '-').replace('日', '')
    s = s.split()[0].replace('.', '-').replace('/', '-')
    parts = [p for p in s.split('-') if p]
    if len(parts) >= 3:
        try:
            year = int(parts[0])
            month = int(parts[1])
            day = int(str(parts[2])[:2])
        except ValueError:
            return None
        if 1 <= year <= 200:
            year += 1911
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date().isoformat()
    except ValueError:
        return None


def _norm_roc_month(value) -> str:
    s = str(value or '').strip().replace('.', '/').replace('-', '/')
    if not s:
        return ''
    parts = [p for p in s.split('/') if p]
    if len(parts) < 2:
        return ''
    try:
        year = int(''.join(c for c in parts[0] if c.isdigit()) or '0')
        month = int(''.join(c for c in parts[1] if c.isdigit()) or '0')
    except ValueError:
        return ''
    if year > 1911:
        year -= 1911
    if year <= 0 or not (1 <= month <= 12):
        return ''
    return f'{year}/{month:02d}'


def _current_roc_month(today=None) -> str:
    d = today or date.today()
    return f'{d.year - 1911}/{d.month:02d}'


def _infer_period_claim_month(conn: sqlite3.Connection, site_id: str, period_name: str) -> str:
    rows = conn.execute(
        '''
        SELECT extra, memo FROM sales_deals
        WHERE site_id = ?
          AND TRIM(COALESCE(commission_period, '')) = ?
        ''',
        (site_id, period_name),
    ).fetchall()
    counts = {}
    for row in rows:
        extra = {}
        try:
            extra = json.loads(row['extra'] or '{}')
        except (TypeError, json.JSONDecodeError):
            extra = {}
        raw = str((extra or {}).get('claimMonth') or '').strip()
        if not raw:
            memo = str(row['memo'] or '')
            marker = '請佣月份 '
            if marker in memo:
                raw = memo.split(marker, 1)[-1].split('；', 1)[0].strip()
        key = _norm_roc_month(raw)
        if key:
            counts[key] = counts.get(key, 0) + 1
    if not counts:
        return ''
    return max(counts, key=counts.get)


def _num(val, default=0.0) -> float:
    try:
        if val is None or val == '':
            return float(default)
        return float(val)
    except (TypeError, ValueError):
        return float(default)


def _truthy(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    return str(val or '').strip().lower() in {'1', 'true', 'yes', 'y', '是', '有'}


def _deal_deductions(data) -> float:
    snake = ('surcharge', 'appliance_gift', 'pickup_voucher', 'decoration', 'company_loan_interest')
    camel = ('surcharge', 'applianceGift', 'pickupVoucher', 'decoration', 'companyLoanInterest')
    if isinstance(data, dict):
        if any(k in data for k in ('appliance_gift', 'pickup_voucher', 'company_loan_interest')):
            return sum(_num(data.get(k)) for k in snake)
        if any(k in data for k in ('applianceGift', 'pickupVoucher', 'companyLoanInterest')):
            return sum(_num(data.get(k)) for k in camel)
        return sum(_num(data.get(k)) for k in snake)
    return sum(_num(data[k]) for k in snake if k in data.keys())


def _calc_excess_price(contract_total, base_total, deductions, imported=None) -> float:
    if imported not in (None, ''):
        val = _num(imported)
        if val != 0 or str(imported).strip() in ('0', '0.0', '-0'):
            return _round4(val)
    return _round4(_num(contract_total) - _num(base_total) - _num(deductions))


def _excess_from_body(body: dict, contract_total: float, base_total: float, deductions: float) -> float:
    extra = body.get('extra')
    if isinstance(extra, str):
        try:
            extra = json.loads(extra or '{}')
        except (TypeError, json.JSONDecodeError):
            extra = {}
    if not isinstance(extra, dict):
        extra = {}
    imported = body.get('excessPrice')
    if imported in (None, '') and extra.get('excessPrice') not in (None, ''):
        imported = extra.get('excessPrice')
    return _calc_excess_price(contract_total, base_total, deductions, imported)


def row_to_deal(row) -> dict:
    extra = {}
    try:
        extra = json.loads(row['extra'] or '{}')
    except (TypeError, json.JSONDecodeError):
        extra = {}
    house_base = row['house_base_price'] or extra.get('houseBasePrice') or 0
    parking_base = row['parking_base_price'] or extra.get('parkingBasePrice') or 0
    base_total = row['base_price'] or extra.get('baseTotal') or extra.get('basePrice') or 0
    house_base_display = house_base
    if house_base == 0 and parking_base == 0 and base_total:
        house_base_display = base_total
    parking_no3 = extra.get('parkingNo3') or ''
    if not parking_no3 and (row['parking_nos'] or ''):
        parts = [x.strip() for x in re.split(r'[、,，/／]', row['parking_nos'] or '') if x.strip()]
        if len(parts) > 2:
            parking_no3 = parts[2]
    return {
        'id': row['id'],
        'siteId': row['site_id'],
        'recordType': row['record_type'],
        'recordTypeLabel': RECORD_TYPES.get(row['record_type'], row['record_type']),
        'orderNo': row['order_no'] or '',
        'unitNo': row['unit_no'] or '',
        'customerName': row['customer_name'] or '',
        'phone': row['phone'] or '',
        'productType': row['product_type'] or '',
        'areaPing': row['area_ping'] or 0,
        'parkingCount': row['parking_count'] or 0,
        'parkingNos': row['parking_nos'] or '',
        'parkingNo1': row['parking_no1'] or '',
        'parkingNo2': row['parking_no2'] or '',
        'parkingNo3': parking_no3,
        'builderCompany': extra.get('builderCompany') or '',
        'community': extra.get('community') or '',
        'listPrice': row['list_price'] or 0,
        'basePrice': row['base_price'] or 0,
        'totalPrice': row['total_price'] or 0,
        'contractTotal': row['total_price'] or 0,
        'houseSalePrice': row['house_sale_price'] or 0,
        'parkingSalePrice': row['parking_sale_price'] or 0,
        'actualHousePrice': row['actual_house_price'] or 0,
        'actualTotalPrice': row['actual_total_price'] or 0,
        'surcharge': row['surcharge'] or extra.get('surcharge') or 0,
        'applianceGift': row['appliance_gift'] or extra.get('applianceGift') or 0,
        'pickupVoucher': row['pickup_voucher'] or extra.get('pickupVoucher') or 0,
        'decoration': row['decoration'] or extra.get('decoration') or 0,
        'companyLoanInterest': row['company_loan_interest'] or extra.get('companyLoanInterest') or 0,
        'houseBasePrice': house_base_display,
        'parkingBasePrice': parking_base,
        'baseTotal': row['base_price'] or 0,
        'excessPrice': _calc_excess_price(
            row['total_price'],
            row['base_price'],
            _deal_deductions(row),
        ),
        'units': row['units'] if row['units'] is not None else 1,
        'depositDate': _parse_date(row['deposit_date']),
        'supplementDate': _parse_date(row['supplement_date']),
        'signDate': _parse_date(row['sign_date']),
        'reportDate': _parse_date(row['report_date']),
        'ownerSaleReportDate': _parse_date(row['owner_sale_report_date'] or row['report_date']),
        'ownerSignReportDate': _parse_date(row['owner_sign_report_date']),
        'salesperson1': row['salesperson1'] or '',
        'salesperson2': row['salesperson2'] or '',
        'isCoManaged': bool(row['is_co_managed']),
        'commissionBaseMode': (row['commission_base_mode'] if 'commission_base_mode' in row.keys() else None) or 'base',
        'commissionSalesAmount': row['commission_sales_amount'] if 'commission_sales_amount' in row.keys() else 0,
        'commissionRate': row['commission_rate'] if 'commission_rate' in row.keys() else 0.0485,
        'commissionPayableRatio': row['commission_payable_ratio'] if 'commission_payable_ratio' in row.keys() else 0.97,
        'commissionRetentionRatio': row['commission_retention_ratio'] if 'commission_retention_ratio' in row.keys() else 0.03,
        'commissionDeduction': row['commission_deduction'] if 'commission_deduction' in row.keys() else 0,
        'commissionClaimable': row['commission_claimable'] or 0,
        'commissionPayable': row['commission_payable'] if 'commission_payable' in row.keys() else 0,
        'commissionRetention': row['commission_retention'] if 'commission_retention' in row.keys() else 0,
        'commissionPeriod': (row['commission_period'] if 'commission_period' in row.keys() else '') or '',
        'commissionClaimDate': _parse_date(row['commission_claim_date']) if 'commission_claim_date' in row.keys() else None,
        'commissionClaimed': row['commission_claimed'] or 0,
        'commissionUnclaimed': row['commission_unclaimed'] if 'commission_unclaimed' in row.keys() else 0,
        'commissionStatus': '已請' if (
            (row['commission_period'] if 'commission_period' in row.keys() else '')
            or (row['commission_claim_date'] if 'commission_claim_date' in row.keys() else None)
        ) else '未請',
        'commissionBooked': row['commission_booked'] or 0,
        'nextMonthClaimable': row['next_month_claimable'] or 0,
        'nextMonthUnits': row['next_month_units'] or 0,
        'nextMonthParking': row['next_month_parking'] or 0,
        'customerId': row['customer_id'],
        'memo': row['memo'] or '',
        'extra': extra,
        'createdAt': row['created_at'],
        'updatedAt': row['updated_at'],
    }


def list_sales_deals(
    conn: sqlite3.Connection,
    site_id: str,
    *,
    record_type: Optional[str] = None,
    q: str = '',
    limit: int = 500,
    offset: int = 0,
) -> tuple[list, int]:
    clauses = ['site_id = ?']
    params: list = [site_id]
    if record_type:
        clauses.append('record_type = ?')
        params.append(record_type)
    if q:
        clauses.append(
            '(order_no LIKE ? OR unit_no LIKE ? OR customer_name LIKE ? OR phone LIKE ? '
            'OR salesperson1 LIKE ? OR parking_nos LIKE ? OR parking_no1 LIKE ? OR parking_no2 LIKE ?)'
        )
        like = f'%{q}%'
        params.extend([like] * 8)
    where = ' AND '.join(clauses)
    total = conn.execute(f'SELECT COUNT(*) AS c FROM sales_deals WHERE {where}', params).fetchone()['c']
    rows = conn.execute(
        f'''
        SELECT * FROM sales_deals
        WHERE {where}
        ORDER BY
          COALESCE(owner_sale_report_date, report_date, sign_date, deposit_date, '') ASC,
          id ASC
        LIMIT ? OFFSET ?
        ''',
        [*params, limit, offset],
    ).fetchall()
    return [row_to_deal(r) for r in rows], int(total)


def get_sales_deal(conn: sqlite3.Connection, deal_id: int, site_id: Optional[str] = None):
    if site_id:
        row = conn.execute(
            'SELECT * FROM sales_deals WHERE id = ? AND site_id = ?',
            (deal_id, site_id),
        ).fetchone()
    else:
        row = conn.execute('SELECT * FROM sales_deals WHERE id = ?', (deal_id,)).fetchone()
    return row_to_deal(row) if row else None


def normalize_deal_payload(body: dict, site_id: str = '', settings: Optional[dict] = None) -> dict:
    record_type = str(body.get('recordType') or 'deal').strip()
    if record_type not in RECORD_TYPES:
        record_type = 'deal'
    s1 = str(body.get('salesperson1') or '').strip()
    s2 = str(body.get('salesperson2') or '').strip()
    co = _truthy(body.get('isCoManaged')) or bool(s1 and s2)
    house_sale = _num(body.get('houseSalePrice'))
    parking_sale = _num(body.get('parkingSalePrice'))
    explicit_total = _num(body.get('totalPrice'))
    surcharge = _num(body.get('surcharge'))
    appliance_gift = _num(body.get('applianceGift'))
    pickup_voucher = _num(body.get('pickupVoucher'))
    decoration = _num(body.get('decoration'))
    company_loan_interest = _num(body.get('companyLoanInterest'))
    house_base = _num(body.get('houseBasePrice'))
    parking_base = _num(body.get('parkingBasePrice'))
    deductions = (
        surcharge + appliance_gift + pickup_voucher
        + decoration + company_loan_interest
    )
    contract_total = house_sale + parking_sale
    parking_no1 = str(body.get('parkingNo1') or '').strip()
    parking_no2 = str(body.get('parkingNo2') or '').strip()
    extra_in = body.get('extra') if isinstance(body.get('extra'), dict) else {}
    parking_no3 = str(body.get('parkingNo3') or extra_in.get('parkingNo3') or '').strip()
    parking_nos = '、'.join(x for x in (parking_no1, parking_no2, parking_no3) if x)
    parking_count = sum(1 for x in (parking_no1, parking_no2, parking_no3) if x)
    if explicit_total > 0:
        if contract_total <= 0:
            contract_total = explicit_total
        elif abs(contract_total - explicit_total) > 0.01:
            contract_total = explicit_total
            if parking_count > 0 and parking_sale <= 0 and house_sale > 0:
                diff = explicit_total - house_sale
                if diff > 0.01 and abs(diff - deductions) > 0.01:
                    parking_sale = diff
    # 請佣總表「房售價」多為未含附加；若 房+車+附加 ≈ 合約總價，實際成交＝房售＋車售
    net_of_extras = (
        deductions > 0.01
        and contract_total > 0
        and abs(house_sale + parking_sale + deductions - contract_total) <= 0.05
    )
    if net_of_extras:
        actual_house = house_sale
        actual_total = house_sale + parking_sale
        if explicit_total <= 0:
            contract_total = house_sale + parking_sale + deductions
    else:
        actual_house = house_sale - deductions
        actual_total = actual_house + parking_sale
    base_total = house_base + parking_base
    excess = _excess_from_body(body, contract_total, base_total, deductions)
    # 相容舊資料／API：尚未拆分房售、車售時保留既有總價。
    if contract_total == 0 and explicit_total:
        contract_total = explicit_total
        actual_house = contract_total - deductions - parking_sale
        actual_total = actual_house + parking_sale
        excess = _excess_from_body(body, contract_total, base_total, deductions)
    if base_total == 0 and _num(body.get('basePrice')):
        base_total = _num(body.get('basePrice'))
        excess = _excess_from_body(body, contract_total, base_total, deductions)
    if not parking_nos:
        parking_nos = str(body.get('parkingNos') or '').strip()
        parking_count = _num(body.get('parkingCount'))

    site = site_id or str(body.get('siteId') or '').strip()
    extra = body.get('extra') if isinstance(body.get('extra'), dict) else {}
    extra = dict(extra)
    extra['parkingNo3'] = parking_no3
    extra['builderCompany'] = str(body.get('builderCompany') or extra.get('builderCompany') or '').strip()
    extra['community'] = str(body.get('community') or extra.get('community') or '').strip()
    if body.get('customerPaidPct') not in (None, ''):
        extra['customerPaidPct'] = _num(body.get('customerPaidPct'))
    body_for_comm = dict(body)
    body_for_comm['extra'] = extra
    comm = compute_commission(
        site_id=site,
        base_total=base_total,
        actual_total=actual_total,
        body=body_for_comm,
        settings=settings,
    )

    return {
        'record_type': record_type,
        'order_no': str(body.get('orderNo') or '').strip(),
        'unit_no': str(body.get('unitNo') or '').strip(),
        'customer_name': str(body.get('customerName') or '').strip(),
        'phone': str(body.get('phone') or '').strip(),
        'product_type': str(body.get('productType') or '').strip(),
        'area_ping': _num(body.get('areaPing')),
        'parking_count': parking_count,
        'parking_nos': parking_nos,
        'parking_no1': parking_no1,
        'parking_no2': parking_no2,
        'list_price': _num(body.get('listPrice')),
        'base_price': base_total,
        'total_price': contract_total,
        'house_sale_price': house_sale,
        'parking_sale_price': parking_sale,
        'actual_house_price': actual_house,
        'actual_total_price': actual_total,
        'surcharge': surcharge,
        'appliance_gift': appliance_gift,
        'pickup_voucher': pickup_voucher,
        'decoration': decoration,
        'company_loan_interest': company_loan_interest,
        'house_base_price': house_base,
        'parking_base_price': parking_base,
        'excess_price': excess,
        'units': _num(body.get('units'), 1),
        'deposit_date': _parse_date(body.get('depositDate')),
        'supplement_date': _parse_date(body.get('supplementDate')),
        'sign_date': _parse_date(body.get('signDate')),
        'report_date': _parse_date(body.get('ownerSaleReportDate') or body.get('reportDate')),
        'owner_sale_report_date': _parse_date(body.get('ownerSaleReportDate') or body.get('reportDate')),
        'owner_sign_report_date': _parse_date(body.get('ownerSignReportDate')),
        'salesperson1': s1,
        'salesperson2': s2,
        'is_co_managed': 1 if co else 0,
        'commission_base_mode': comm['commission_base_mode'],
        'commission_sales_amount': comm['commission_sales_amount'],
        'commission_rate': comm['commission_rate'],
        'commission_payable_ratio': comm['commission_payable_ratio'],
        'commission_retention_ratio': comm['commission_retention_ratio'],
        'commission_deduction': comm['commission_deduction'],
        'commission_claimable': comm['commission_claimable'],
        'commission_payable': comm['commission_payable'],
        'commission_retention': comm['commission_retention'],
        'commission_period': comm['commission_period'],
        'commission_claim_date': comm['commission_claim_date'],
        'commission_claimed': comm['commission_claimed'],
        'commission_unclaimed': comm['commission_unclaimed'],
        'commission_booked': _num(body.get('commissionBooked')),
        'next_month_claimable': _num(body.get('nextMonthClaimable')),
        'next_month_units': _num(body.get('nextMonthUnits')),
        'next_month_parking': _num(body.get('nextMonthParking')),
        'customer_id': int(body['customerId']) if body.get('customerId') not in (None, '') else None,
        'memo': str(body.get('memo') or '').strip(),
        'extra': json.dumps(extra, ensure_ascii=False),
    }


def create_sales_deal(conn: sqlite3.Connection, site_id: str, body: dict, user_id=None) -> int:
    data = normalize_deal_payload(
        body, site_id=site_id, settings=commission_defaults_for_site(site_id, conn=conn),
    )
    cur = conn.execute(
        '''
        INSERT INTO sales_deals (
          site_id, record_type, order_no, unit_no, customer_name, phone, product_type,
          area_ping, parking_count, parking_nos, parking_no1, parking_no2,
          list_price, base_price, total_price,
          house_sale_price, parking_sale_price, actual_house_price, actual_total_price,
          surcharge, appliance_gift, pickup_voucher, decoration, company_loan_interest,
          house_base_price, parking_base_price, excess_price, units,
          deposit_date, supplement_date, sign_date, report_date,
          owner_sale_report_date, owner_sign_report_date,
          salesperson1, salesperson2, is_co_managed,
          commission_base_mode, commission_sales_amount, commission_rate,
          commission_payable_ratio, commission_retention_ratio, commission_deduction,
          commission_claimable, commission_payable, commission_retention,
          commission_period, commission_claim_date, commission_claimed, commission_unclaimed,
          commission_booked,
          next_month_claimable, next_month_units, next_month_parking,
          customer_id, memo, extra, created_by, updated_by
        ) VALUES (
          ?,?,?,?,?,?,?,
          ?,?,?,?,?,
          ?,?,?,
          ?,?,?,?,
          ?,?,?,?,?,
          ?,?,?,?,
          ?,?,?,?,?,?,
          ?,?,?,
          ?,?,?,?,?,?,
          ?,?,?,
          ?,?,?,?,
          ?,
          ?,?,?,
          ?,?,?,?,?
        )
        ''',
        (
            site_id, data['record_type'], data['order_no'], data['unit_no'], data['customer_name'],
            data['phone'], data['product_type'],
            data['area_ping'], data['parking_count'], data['parking_nos'],
            data['parking_no1'], data['parking_no2'],
            data['list_price'], data['base_price'], data['total_price'],
            data['house_sale_price'], data['parking_sale_price'], data['actual_house_price'],
            data['actual_total_price'], data['surcharge'], data['appliance_gift'],
            data['pickup_voucher'], data['decoration'], data['company_loan_interest'],
            data['house_base_price'], data['parking_base_price'], data['excess_price'], data['units'],
            data['deposit_date'], data['supplement_date'], data['sign_date'], data['report_date'],
            data['owner_sale_report_date'], data['owner_sign_report_date'],
            data['salesperson1'], data['salesperson2'],
            data['is_co_managed'],
            data['commission_base_mode'], data['commission_sales_amount'], data['commission_rate'],
            data['commission_payable_ratio'], data['commission_retention_ratio'], data['commission_deduction'],
            data['commission_claimable'], data['commission_payable'], data['commission_retention'],
            data['commission_period'], data['commission_claim_date'], data['commission_claimed'],
            data['commission_unclaimed'], data['commission_booked'],
            data['next_month_claimable'], data['next_month_units'], data['next_month_parking'],
            data['customer_id'], data['memo'], data['extra'], user_id, user_id,
        ),
    )
    return cur.lastrowid


def update_sales_deal(conn: sqlite3.Connection, deal_id: int, site_id: str, body: dict, user_id=None) -> bool:
    existing = conn.execute(
        'SELECT id FROM sales_deals WHERE id = ? AND site_id = ?',
        (deal_id, site_id),
    ).fetchone()
    if not existing:
        return False
    existing_deal = get_sales_deal(conn, deal_id, site_id) or {}
    old_extra = existing_deal.get('extra') if isinstance(existing_deal.get('extra'), dict) else {}
    new_extra = body.get('extra') if isinstance(body.get('extra'), dict) else {}
    payload = dict(body)
    payload['extra'] = {**old_extra, **new_extra}
    data = normalize_deal_payload(
        payload, site_id=site_id, settings=commission_defaults_for_site(site_id, conn=conn),
    )
    conn.execute(
        '''
        UPDATE sales_deals SET
          record_type=?, order_no=?, unit_no=?, customer_name=?, phone=?, product_type=?,
          area_ping=?, parking_count=?, parking_nos=?, parking_no1=?, parking_no2=?,
          list_price=?, base_price=?, total_price=?,
          house_sale_price=?, parking_sale_price=?, actual_house_price=?, actual_total_price=?,
          surcharge=?, appliance_gift=?, pickup_voucher=?, decoration=?, company_loan_interest=?,
          house_base_price=?, parking_base_price=?, excess_price=?, units=?,
          deposit_date=?, supplement_date=?, sign_date=?, report_date=?,
          owner_sale_report_date=?, owner_sign_report_date=?,
          salesperson1=?, salesperson2=?, is_co_managed=?,
          commission_base_mode=?, commission_sales_amount=?, commission_rate=?,
          commission_payable_ratio=?, commission_retention_ratio=?, commission_deduction=?,
          commission_claimable=?, commission_payable=?, commission_retention=?,
          commission_period=?, commission_claim_date=?, commission_claimed=?, commission_unclaimed=?,
          commission_booked=?,
          next_month_claimable=?, next_month_units=?, next_month_parking=?,
          customer_id=?, memo=?, extra=?, updated_by=?,
          updated_at=datetime('now', 'localtime')
        WHERE id=? AND site_id=?
        ''',
        (
            data['record_type'], data['order_no'], data['unit_no'], data['customer_name'],
            data['phone'], data['product_type'],
            data['area_ping'], data['parking_count'], data['parking_nos'],
            data['parking_no1'], data['parking_no2'],
            data['list_price'], data['base_price'], data['total_price'],
            data['house_sale_price'], data['parking_sale_price'], data['actual_house_price'],
            data['actual_total_price'], data['surcharge'], data['appliance_gift'],
            data['pickup_voucher'], data['decoration'], data['company_loan_interest'],
            data['house_base_price'], data['parking_base_price'], data['excess_price'], data['units'],
            data['deposit_date'], data['supplement_date'], data['sign_date'], data['report_date'],
            data['owner_sale_report_date'], data['owner_sign_report_date'],
            data['salesperson1'], data['salesperson2'],
            data['is_co_managed'],
            data['commission_base_mode'], data['commission_sales_amount'], data['commission_rate'],
            data['commission_payable_ratio'], data['commission_retention_ratio'], data['commission_deduction'],
            data['commission_claimable'], data['commission_payable'], data['commission_retention'],
            data['commission_period'], data['commission_claim_date'], data['commission_claimed'],
            data['commission_unclaimed'], data['commission_booked'],
            data['next_month_claimable'], data['next_month_units'], data['next_month_parking'],
            data['customer_id'], data['memo'], data['extra'], user_id,
            deal_id, site_id,
        ),
    )
    return True


def delete_sales_deal(conn: sqlite3.Connection, deal_id: int, site_id: str) -> bool:
    cur = conn.execute(
        'DELETE FROM sales_deals WHERE id = ? AND site_id = ?',
        (deal_id, site_id),
    )
    return cur.rowcount > 0


def delete_all_sales_deals(conn: sqlite3.Connection, site_id: str) -> dict:
    """清空指定案場的銷售明細與期別服務費批次。"""
    deals = conn.execute(
        'DELETE FROM sales_deals WHERE site_id = ?',
        (site_id,),
    ).rowcount
    batches = conn.execute(
        'DELETE FROM commission_batches WHERE site_id = ?',
        (site_id,),
    ).rowcount
    return {'deals': int(deals), 'batches': int(batches)}


def _in_range(date_s: Optional[str], start, end) -> bool:
    parsed = _parse_date(date_s)
    if not parsed:
        return False
    try:
        d = datetime.strptime(parsed, '%Y-%m-%d').date()
    except ValueError:
        return False
    return start <= d <= end


def _on_or_before(date_s: Optional[str], end, *, include_blank=True) -> bool:
    parsed = _parse_date(date_s)
    if not parsed:
        return include_blank
    try:
        d = datetime.strptime(parsed, '%Y-%m-%d').date()
    except ValueError:
        return include_blank
    return d <= end


def _block():
    return {'units': 0.0, 'parking': 0.0, 'amount': 0.0}


def _product_group(product_type: str) -> str:
    text = str(product_type or '')
    if '店面' in text:
        return 'storefront'
    if '店鋪' in text or '店舖' in text:
        return 'shop'
    if '事務' in text or '辦公' in text:
        return 'office'
    if text.strip():
        return 'residential'
    return ''


def aggregate_for_weekly(conn: sqlite3.Connection, site_id: str, start, end) -> dict:
    """彙總銷售總表 → 週報成交／簽約／買進／未報／請佣／去化建議值。"""
    settings = commission_defaults_for_site(site_id, conn=conn)
    payable_ratio = float(settings.get('payableRatio') or 0.97)
    retention_ratio = float(settings.get('retentionRatio') or 0.03)
    scheme = settings.get('scheme') or 'simple'
    rows = conn.execute(
        'SELECT * FROM sales_deals WHERE site_id = ?',
        (site_id,),
    ).fetchall()

    deals = _block()
    signings = _block()
    purchases = _block()
    deals_cum = _block()
    signings_cum = _block()
    purchases_cum = _block()
    unreported = _block()
    refunds = _block()

    sellable_amount = 0.0
    claimable_amount = 0.0
    payable_amount = 0.0
    retention_amount = 0.0
    claimed_amount = 0.0
    unclaimed_amount = 0.0
    booked_amount = 0.0
    claimable_units = 0.0
    claimable_parking = 0.0
    claimed_units = 0.0
    claimed_parking = 0.0
    next_month_amount = 0.0
    next_month_units = 0.0
    next_month_parking = 0.0

    matrix_all = _comm_bucket()
    matrix_claimed = _comm_bucket()
    matrix_unclaimed = _comm_bucket()
    matrix_forecast = _comm_bucket()

    sold_units = 0.0
    sold_parking = 0.0
    sold_amount = 0.0
    sold_base = 0.0
    residential_sold = 0.0
    office_sold = 0.0
    shop_sold = 0.0
    storefront_sold = 0.0

    week_deal_rows = []

    for row in rows:
        rt = row['record_type'] or 'deal'
        units = _num(row['units'], 1)
        parking = _num(row['parking_count'])
        contract_amount = _num(row['total_price']) or _num(row['list_price'])
        actual_amount = _num(row['actual_total_price']) if 'actual_total_price' in row.keys() else 0
        if actual_amount <= 0:
            actual_house = _num(row['actual_house_price']) if 'actual_house_price' in row.keys() else 0
            parking_sale = _num(row['parking_sale_price']) if 'parking_sale_price' in row.keys() else 0
            actual_amount = (
                actual_house + parking_sale
                if (actual_house or parking_sale)
                else contract_amount
            )
        # 成交／簽約／買進金額＝實際房價＋車售；未報金額＝合約總價
        amount = contract_amount if rt == 'unreported' else actual_amount
        base = _num(row['base_price'])
        report_d = row['owner_sale_report_date'] or row['report_date']
        sign_d = row['sign_date']
        owner_sign_d = row['owner_sign_report_date']
        deposit_d = row['deposit_date']

        # 累計請佣／銷售（不含退換戶，或退換戶沖銷）
        if rt == 'refund':
            sellable_amount -= actual_amount
            sold_units -= units
            sold_parking -= parking
            sold_amount -= actual_amount
            sold_base -= base
        elif rt in ('deal', 'signing', 'unreported', 'purchase'):
            # 已報／簽約／買進計入累積銷售；未報也常算入銷售金額但週報另列
            if rt != 'unreported':
                sellable_amount += actual_amount
                sold_units += units
                sold_parking += parking
                sold_amount += actual_amount
                sold_base += base
                pt_group = _product_group(row['product_type'])
                if pt_group == 'office':
                    office_sold += units
                elif pt_group == 'shop':
                    shop_sold += units
                elif pt_group == 'storefront':
                    storefront_sold += units
                elif pt_group == 'residential':
                    residential_sold += units

            claimable = _num(row['commission_claimable'])
            payable = _num(row['commission_payable']) if 'commission_payable' in row.keys() else 0
            retention = _num(row['commission_retention']) if 'commission_retention' in row.keys() else 0
            claimed = _num(row['commission_claimed'])
            period = str(row['commission_period'] if 'commission_period' in row.keys() else '')
            claim_date = row['commission_claim_date'] if 'commission_claim_date' in row.keys() else None
            is_claimed = bool(claimed > 0 or period.strip() or claim_date)
            unclaimed = (
                _num(row['commission_unclaimed'])
                if 'commission_unclaimed' in row.keys()
                else max(claimable - claimed, 0)
            )
            if payable <= 0 and claimable > 0 and scheme != 'payment_tiers':
                payable = claimable * payable_ratio
            if claimable > 0 and scheme != 'payment_tiers':
                retention = claimable * retention_ratio
            payable_amount += payable
            retention_amount += retention
            claimable_amount += claimable
            claimed_amount += claimed
            booked_amount += _num(row['commission_booked'])
            if claimable > 0:
                claimable_units += units
                claimable_parking += parking
                _add_to_bucket(matrix_all, units, parking, claimable, retention, payable)
            if is_claimed:
                claimed_units += units
                claimed_parking += parking
                _add_to_bucket(matrix_claimed, units, parking, claimable, retention, payable)
            else:
                unclaimed_amount += unclaimed if unclaimed else claimable
                _add_to_bucket(matrix_unclaimed, units, parking, claimable, retention, payable)

        # 本週區塊
        def add(block, u, p, a):
            block['units'] += u
            block['parking'] += p
            block['amount'] += a

        sale_d = report_d
        sign_report_d = owner_sign_d or sign_d

        if rt in ('deal', 'signing') and _in_range(sale_d, start, end):
            add(deals, units, parking, amount)
            week_deal_rows.append(row_to_deal(row))
        if rt == 'signing' and _in_range(sign_report_d, start, end):
            add(signings, units, parking, amount)
        elif rt == 'purchase' and _in_range(sale_d or report_d or sign_d or deposit_d, start, end):
            add(purchases, units, parking, amount)
        elif rt == 'unreported':
            add(unreported, units, parking, amount)
        elif rt == 'refund' and _in_range(report_d or sign_d, start, end):
            add(refunds, units, parking, amount)

        if rt in ('deal', 'signing') and _on_or_before(sale_d, end):
            add(deals_cum, units, parking, amount)
        if rt == 'signing' and _on_or_before(sign_report_d, end):
            add(signings_cum, units, parking, amount)
        elif rt == 'purchase' and _on_or_before(sale_d or report_d or sign_d or deposit_d, end):
            add(purchases_cum, units, parking, amount)

    def round_block(b):
        return {
            'units': round(b['units'], 2),
            'parking': round(b['parking'], 2),
            'amount': round(b['amount'], 2),
        }

    unclaimed_units = max(claimable_units - claimed_units, 0)
    unclaimed_parking = max(claimable_parking - claimed_parking, 0)
    # 若未請金額未從「未請列」累加到（相容舊資料），用矩陣 payable 補
    if unclaimed_amount <= 0 and matrix_unclaimed['payable'] > 0:
        unclaimed_amount = matrix_unclaimed['payable']

    # 已入帳／預計本月：依期別「預計入帳日」與「請款月份」彙總，不必逐戶勾選
    batches = list_commission_batches(conn, site_id)
    booked_from_batches = 0.0
    has_deposit_dates = False
    this_month = _current_roc_month()
    for b in batches:
        if b.get('depositDate1') or b.get('depositDate2'):
            has_deposit_dates = True
        booked_from_batches += _num(b.get('bookedTotal'))
        if _norm_roc_month(b.get('claimMonth')) != this_month:
            continue
        units_b = _num(b.get('units'))
        parking_b = _num(b.get('parking'))
        payable_b = _num(b.get('amountPayable'))
        claimable_b = _num(b.get('claimable'))
        if claimable_b <= 0 and payable_b > 0 and scheme != 'payment_tiers' and payable_ratio:
            claimable_b = payable_b / payable_ratio
        retention_b = _num(b.get('retention'))
        if scheme != 'payment_tiers':
            retention_b = claimable_b * retention_ratio
            if payable_b <= 0:
                payable_b = claimable_b * payable_ratio
        next_month_units += units_b
        next_month_parking += parking_b
        next_month_amount += payable_b
        _add_to_bucket(matrix_forecast, units_b, parking_b, claimable_b, retention_b, payable_b)
    if has_deposit_dates:
        booked_amount = booked_from_batches

    commission_matrix = {
        'claimable': _round_bucket(matrix_all, settings=settings),
        'claimed': _round_bucket(matrix_claimed, settings=settings),
        'unclaimed': _round_bucket(matrix_unclaimed, settings=settings),
        'forecast': _round_bucket(matrix_forecast, settings=settings),
        'labels': commission_matrix_labels(settings),
        'totals': {
            'bookedAmount': round(booked_amount, 4),
            'sellableAmount': round(sellable_amount, 4),
        },
    }

    return {
        'deals': round_block(deals),
        'dealsCum': round_block(deals_cum),
        'signings': round_block(signings),
        'signingsCum': round_block(signings_cum),
        'unsignedCum': round_block({
            'units': max(deals_cum['units'] - signings_cum['units'], 0),
            'parking': max(deals_cum['parking'] - signings_cum['parking'], 0),
            'amount': max(deals_cum['amount'] - signings_cum['amount'], 0),
        }),
        'purchases': round_block(purchases),
        'purchasesCum': round_block(purchases_cum),
        'unreported': round_block(unreported),
        'refunds': round_block(refunds),
        'commission': {
            'sellableUnits': round(max(sold_units, 0), 2),
            'sellableParking': round(max(sold_parking, 0), 2),
            'sellableAmount': round(sellable_amount, 4),
            'claimableAmount': round(claimable_amount, 4),
            'payableAmount': round(payable_amount, 4),
            'retentionAmount': round(retention_amount, 4),
            'claimedAmount': round(claimed_amount, 4),
            'unclaimedAmount': round(unclaimed_amount, 4),
            'bookedAmount': round(booked_amount, 4),
            'claimableUnits': round(claimable_units, 2),
            'claimableParking': round(claimable_parking, 2),
            'claimedUnits': round(claimed_units, 2),
            'claimedParking': round(claimed_parking, 2),
            'unclaimedUnits': round(unclaimed_units, 2),
            'unclaimedParking': round(unclaimed_parking, 2),
            'nextMonthUnits': round(next_month_units, 2),
            'nextMonthParking': round(next_month_parking, 2),
            'nextMonthAmount': round(next_month_amount, 4),
        },
        'commissionMatrix': commission_matrix,
        'inventory': {
            'soldUnits': round(max(sold_units, 0), 2),
            'soldParking': round(max(sold_parking, 0), 2),
            'soldAmount': round(max(sold_amount, 0), 2),
            'soldBasePrice': round(max(sold_base, 0), 2),
            'residentialSold': round(max(residential_sold, 0), 2),
            'officeSold': round(max(office_sold, 0), 2),
            'shopSold': round(max(shop_sold, 0), 2),
            'storefrontSold': round(max(storefront_sold, 0), 2),
        },
        'weekDealCount': len(week_deal_rows),
        'weekDeals': week_deal_rows[:50],
        'totalRecords': len(rows),
    }


def sales_export_headers() -> list[str]:
    return [
        '類型', '訂單編號', '建設公司', '社區', '戶別', '客戶', '產品類型', '坪數',
        '車位1', '車位2', '車位3', '房售價(萬)', '車位售價(萬)', '合約總價(萬)', '實際成交總價(萬)',
        '附加費(萬)', '家電禮券(萬)', '提貨券(萬)', '裝潢(萬)', '公司貸利息(萬)',
        '房底(萬)', '車底(萬)', '底總(萬)', '超價(萬)', '下訂日', '補足日', '簽約日',
        '請佣計價方式', '請佣銷售金額(萬)', '可請佣(萬)', '本期可請97%(萬)',
        '保留款3%(萬)', '已請(萬)', '未請(萬)', '請佣狀態', '請佣期別', '請佣日期',
        '已入帳金額(萬)', '預計本月可請戶數', '預計本月可請車位',
        '預計本月可請金額(萬)', '業主報售日', '業主報簽日',
        '銷售人員1', '銷售人員2', '備註',
    ]


def sales_export_row_values(row: dict) -> list:
    return [
        row.get('recordTypeLabel') or row.get('recordType'), row.get('orderNo'),
        row.get('builderCompany'), row.get('community'),
        row.get('unitNo'), row.get('customerName'), row.get('productType'),
        row.get('areaPing'), row.get('parkingNo1'), row.get('parkingNo2'), row.get('parkingNo3'),
        row.get('houseSalePrice'), row.get('parkingSalePrice'),
        row.get('contractTotal'), row.get('actualTotalPrice'),
        row.get('surcharge'), row.get('applianceGift'), row.get('pickupVoucher'),
        row.get('decoration'), row.get('companyLoanInterest'),
        row.get('houseBasePrice'), row.get('parkingBasePrice'),
        row.get('baseTotal'), row.get('excessPrice'),
        row.get('depositDate'), row.get('supplementDate'), row.get('signDate'),
        '成交價' if row.get('commissionBaseMode') == 'deal' else '底價',
        row.get('commissionSalesAmount'), row.get('commissionClaimable'),
        row.get('commissionPayable'), row.get('commissionRetention'),
        row.get('commissionClaimed'), row.get('commissionUnclaimed'),
        row.get('commissionStatus'), row.get('commissionPeriod'),
        row.get('commissionClaimDate'), row.get('commissionBooked'),
        row.get('nextMonthUnits'), row.get('nextMonthParking'), row.get('nextMonthClaimable'),
        row.get('ownerSaleReportDate'), row.get('ownerSignReportDate'),
        row.get('salesperson1'), row.get('salesperson2'), row.get('memo'),
    ]


def build_sales_excel(site_name: str, rows: list[dict]) -> bytes:
    """匯出目前篩選到的銷售明細，並於最後一列加總金額。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '銷售總表'
    headers = sales_export_headers()
    ws.append([f'{site_name} 銷售總表'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(name='微軟正黑體', bold=True, size=16, color='1A4D7C')
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='1A4D7C')
    for cell in ws[2]:
        cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    amount_keys = [
        'contractTotal', 'actualTotalPrice',
        'houseBasePrice', 'parkingBasePrice', 'baseTotal', 'excessPrice',
        'commissionSalesAmount', 'commissionClaimable', 'commissionPayable',
        'commissionRetention', 'commissionClaimed', 'commissionUnclaimed',
        'commissionBooked', 'nextMonthUnits', 'nextMonthParking', 'nextMonthClaimable',
    ]
    totals = {key: 0.0 for key in amount_keys}
    for row in rows:
        for key in amount_keys:
            totals[key] += _num(row.get(key))
        ws.append(sales_export_row_values(row))

    total_values = [''] * len(headers)
    total_values[0] = '合計'
    total_values[5] = f'{len(rows)} 筆'
    total_columns = {
        'contractTotal': 13, 'actualTotalPrice': 14,
        'surcharge': 15, 'applianceGift': 16, 'pickupVoucher': 17,
        'decoration': 18, 'companyLoanInterest': 19,
        'houseBasePrice': 20, 'parkingBasePrice': 21, 'baseTotal': 22, 'excessPrice': 23,
        'commissionSalesAmount': 28, 'commissionClaimable': 29, 'commissionPayable': 30,
        'commissionRetention': 31, 'commissionClaimed': 32, 'commissionUnclaimed': 33,
        'commissionBooked': 37, 'nextMonthUnits': 38, 'nextMonthParking': 39,
        'nextMonthClaimable': 40,
    }
    for key, column_idx in total_columns.items():
        total_values[column_idx] = round(totals[key], 4)
    ws.append(total_values)
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(name='微軟正黑體', bold=True, size=11)
        cell.fill = PatternFill('solid', fgColor='E8F1F8')

    thin = Border(
        left=Side(style='thin', color='D9E2EC'),
        right=Side(style='thin', color='D9E2EC'),
        top=Side(style='thin', color='D9E2EC'),
        bottom=Side(style='thin', color='D9E2EC'),
    )
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            if cell.value is not None or cell.font:
                f = cell.font
                cell.font = Font(
                    name='微軟正黑體',
                    size=f.size or 11,
                    bold=bool(f.bold),
                    color=f.color,
                )
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    # 標題列不加資料區邊框過密：上面已統一；維持欄寬
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions[get_column_letter(len(headers))].width = 26
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{max(ws.max_row - 1, 2)}'

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
