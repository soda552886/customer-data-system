import csv
import io
import json
import os
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from flask import Flask, jsonify, request, send_from_directory, Response, session, redirect
from openpyxl import load_workbook

from auth import (
    PROTECTED_PAGES, ROLES, get_allowed_site_ids, get_current_user,
    get_user_by_id, get_user_by_username, hash_password, init_auth_tables, is_public_api,
    migrate_retired_roles, save_user_sites, seed_initial_admin, user_can_access_site,
    user_can_bulk_delete_customers, user_can_delete_customer, user_has_permission,
    verify_password,
)
from audit import init_audit_table, log_operation, row_to_log_dict
from config.sites import SITES as DEFAULT_SITES
from weekly_reports import (
    build_auto_stats, build_weekly_excel, commission_summary, default_week_number,
    empty_manual_payload, enrich_dims_with_phones, init_weekly_tables, inventory_summary,
    list_weekly_reports, load_weekly_report, merge_manual, monday_of,
    normalize_phone_calls_detail, phone_total_from_manual, roc_year, upsert_weekly_report,
    week_bounds, safe_week_number,
)
from sales_ledger import (
    RECORD_TYPES as SALES_RECORD_TYPES, aggregate_for_weekly, build_commission_overview_excel,
    build_sales_excel, commission_defaults_for_site, create_sales_deal, delete_all_sales_deals,
    delete_commission_batch, delete_sales_deal, get_sales_deal, init_sales_tables,
    list_commission_batches, list_sales_deals, update_sales_deal, upsert_commission_batch,
)
from field_options import (
    apply_site_field_options, apply_site_hidden_fields, build_site_field_config,
    build_site_field_visibility, build_site_report_export_config,
    export_column_keys_for_site, init_field_options_table, load_site_hidden_fields,
    load_site_option_overrides, load_site_report_export_config,
    normalize_hidden_fields_payload, normalize_report_export_payload, normalize_save_payload,
    save_site_hidden_fields, save_site_option_overrides, save_site_report_export_config,
    SALES_STAFF_FIELD_KEY, resolve_field_options,
)

BASE_DIR = Path(__file__).parent

_report_cols_path = BASE_DIR / 'config' / 'report_columns.json'
with open(_report_cols_path, encoding='utf-8') as _rc:
    REPORT_COLUMNS = json.load(_rc)['columns']

_fields_path = BASE_DIR / 'config' / 'fields_data.json'
with open(_fields_path, encoding='utf-8') as _f:
    _field_data = json.load(_f)
FIELD_SECTIONS = _field_data['sections']
SALES_STAFF = _field_data['salesStaff']

app = Flask(__name__, static_folder='public', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-in-production-deyijia-2026')

DATA_DIR = Path(os.environ.get('DATA_DIR', str(BASE_DIR / 'data')))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / 'customers.db'


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS sites (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            group_type TEXT NOT NULL DEFAULT 'residential',
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        );
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id TEXT NOT NULL,
            site_name TEXT NOT NULL,
            visit_type TEXT NOT NULL,
            is_deal INTEGER NOT NULL DEFAULT 0,
            visit_date TEXT,
            first_visit_date TEXT,
            return_visit_date TEXT,
            data TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_customers_site ON customers(site_id);
        CREATE INDEX IF NOT EXISTS idx_customers_visit_type ON customers(visit_type);
        CREATE INDEX IF NOT EXISTS idx_customers_visit_date ON customers(visit_date);
    ''')
    init_auth_tables(conn)
    init_field_options_table(conn)
    init_audit_table(conn)
    init_weekly_tables(conn)
    init_sales_tables(conn)
    migrate_retired_roles(conn)
    conn.commit()

    count = conn.execute('SELECT COUNT(*) FROM sites').fetchone()[0]
    if count == 0:
        for s in DEFAULT_SITES:
            conn.execute(
                'INSERT OR IGNORE INTO sites (id, name, group_type) VALUES (?, ?, ?)',
                (s['id'], s['name'], s.get('group', 'residential')),
            )
        conn.commit()
    seed_initial_admin(conn)
    conn.close()


def load_sites():
    conn = get_db()
    rows = conn.execute('''
        SELECT s.id, s.name, s.group_type, s.created_at,
               (SELECT COUNT(*) FROM customers c WHERE c.site_id = s.id) AS customer_count
        FROM sites s ORDER BY s.name
    ''').fetchall()
    conn.close()
    return [{
        'id': r['id'], 'name': r['name'], 'group': r['group_type'],
        'created_at': r['created_at'], 'customer_count': r['customer_count'],
    } for r in rows]


def make_site_id(name):
    base = re.sub(r'[\s/\\]+', '_', name.strip())
    base = re.sub(r'[^\w\u4e00-\u9fff-]', '', base, flags=re.UNICODE)
    if not base:
        base = f'site_{int(datetime.now().timestamp())}'
    base = base[:60]
    conn = get_db()
    candidate = base
    n = 1
    while conn.execute('SELECT 1 FROM sites WHERE id = ?', (candidate,)).fetchone():
        candidate = f'{base}_{n}'
        n += 1
    conn.close()
    return candidate


def get_site_by_id(site_id):
    if not site_id:
        return None
    conn = get_db()
    row = conn.execute(
        'SELECT id, name, group_type FROM sites WHERE id = ?', (site_id,)
    ).fetchone()
    conn.close()
    if not row:
        return None
    return {'id': row['id'], 'name': row['name'], 'group': row['group_type']}


def get_active_sales_staff(conn, site_id: str) -> list:
    """在職銷售名單（案場欄位選項有勾選者）；未自訂則用系統預設。"""
    overrides = load_site_option_overrides(conn, site_id)
    defaults = list(SALES_STAFF.get(site_id) or [])
    if SALES_STAFF_FIELD_KEY in overrides:
        _, enabled, _ = resolve_field_options(defaults, overrides, SALES_STAFF_FIELD_KEY)
        return list(enabled)
    return defaults


def get_site_maps():
    sites = load_sites()
    return {s['name']: s for s in sites}, {s['id']: s for s in sites}


def normalize_phone(phone):
    return re.sub(r'\D', '', str(phone or ''))


def format_phone_for_storage(phone):
    digits = normalize_phone(phone)
    if len(digits) == 9 and digits.startswith('9'):
        return '0' + digits
    if digits:
        return digits
    return str(phone or '').strip()


def normalize_date_value(val):
    if val is None or str(val).strip() == '':
        return None
    s = str(val).strip().split()[0]

    def as_roc_or_ad(year: int) -> int:
        # 民國年常見範圍（避免把錯誤數字加成 2135 這類年份）
        if 1 <= year <= 200:
            return year + 1911
        return year

    # YYYY-MM-DD / YYYY/M/D（若年 < 1911 視為民國，如 0114）
    m = re.match(r'^(\d{4})[./-](\d{1,2})[./-](\d{1,2})$', s)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = as_roc_or_ad(year) if year < 1911 else year
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f'{year}-{month:02d}-{day:02d}'

    # 民國 Y/M/D（2～3 碼年）
    m = re.match(r'^(\d{2,3})[./-](\d{1,2})[./-](\d{1,2})$', s)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = as_roc_or_ad(year)
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f'{year}-{month:02d}-{day:02d}'

    # M/D/YYYY 或 M/D/民國年（如 9/2/0114、9/2/114）
    m = re.match(r'^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$', s)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = as_roc_or_ad(year) if year < 1911 else year
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f'{year}-{month:02d}-{day:02d}'

    return s


def normalize_record_dates(system, data):
    for key in ('visitDate', 'firstVisitDate', 'returnVisitDate', 'prevVisitDate'):
        if data.get(key):
            normalized = normalize_date_value(data[key])
            if normalized:
                data[key] = normalized
    if system.get('visit_date'):
        normalized = normalize_date_value(system['visit_date'])
        if normalized:
            system['visit_date'] = normalized
    if not system.get('visit_date'):
        for key in ('returnVisitDate', 'visitDate', 'firstVisitDate', 'prevVisitDate'):
            if data.get(key):
                system['visit_date'] = data[key]
                break
        if not system.get('visit_date') and system.get('_timestamp'):
            system['visit_date'] = system['_timestamp']


def infer_from_filename(filename):
    name = filename or ''
    site_id = None
    visit_type = None
    for site in load_sites():
        if site['name'] in name:
            site_id = site['id']
            break
    if '回訪' in name:
        visit_type = '回訪'
    elif '新客' in name:
        visit_type = '新客'
    return site_id, visit_type


def row_has_visit_type(row_dict):
    for header, raw_val in row_dict.items():
        if map_header_to_key(header) == '_visit_type' and str(raw_val or '').strip():
            return True
    return False


MULTISELECT_KEYS = {
    'roomType', 'floorNeed', 'areaNeed', 'unitType', 'unitNeed',
    'notPurchasedReason', 'purchasedReason',
}

LABEL_TO_KEY = {
    '日期': '_visit_date', '案場': '_site_name', '客戶類型': '_visit_type',
    '是否成交': '_is_deal', '建檔時間': '_created_at',
    '參觀日期': 'visitDate', '首次參觀日期': 'firstVisitDate',
    '回訪日期': 'returnVisitDate', '前次來訪日期': 'prevVisitDate',
    '回訪次數': 'visitCount', '回籠次數': 'returnCount',
    '客戶姓名': 'customerName', '主要電話': 'phone', '主要聯繫電話': 'phone',
    '電話': 'phone', '聯絡電話': 'phone', '手機': 'phone', '連絡電話': 'phone',
    '次要電話': 'phoneSecondary', '次要聯繫電話': 'phoneSecondary',
    '居住地址': 'address', '街道路名或社區': 'streetCommunity', '區域': 'region',
    '年齡': 'age', '職業': 'occupation', '購屋用途': 'purchasePurpose',
    '購屋動機': 'purchaseMotive', '購屋需求': 'purchaseNeed',
    '總價預算': 'budget', '預算總價': 'budget', '自備款': 'downPayment',
    '媒體1': 'media1', '媒體2': 'media2', '媒體3': 'media3', '媒體': 'media',
    '介紹建案': 'commercialProject', '需求房型': 'roomType', '需求樓層': 'floorNeed',
    '需求坪數': 'areaNeed', '需求戶型': 'unitType', '需求戶別': 'unitNeed',
    '房間需求': 'roomNeed', '車位需求': 'parkingNeed',
    '產品需求-住宅': 'productResidential', '產品需求-事務所': 'productOffice',
    '介紹戶別樓層': 'introUnit', '介紹戶別': 'introUnit',
    '介紹戶別樓層（ex︰A1-2F）': 'introUnit',
    '當日來人': 'visitorCount', '當日來訪人數': 'visitorCount',
    '來人關係': 'visitorRelation', '同行人員關係': 'visitorRelation',
    '未購因素': 'notPurchasedReason', '成交因素': 'purchasedReason',
    '已購/成交因素': 'purchasedReason', '已購因素': 'purchasedReason',
    '洽談內容': 'discussion', '客戶來源': 'customerSource', '客戶誠意度': 'sincerity',
    '銷售人員1': 'salesperson1', '銷售人員2': 'salesperson2',
    '接待人員': 'salesperson1', '接待人員1': 'salesperson1', '接待人員2': 'salesperson2',
    '備註': 'remark', '客戶狀態': 'customerStatus',
    '退戶日期': 'cancelDate', '退戶原因': 'cancelReason',
    '時間戳記': '_timestamp',
    '第一次及前次來訪日期': 'prevVisitDate',
    '回訪次數(不含首次參觀)': 'visitCount',
}

ALL_FIELD_KEYS = set(LABEL_TO_KEY.values()) | MULTISELECT_KEYS | {
    'visitDate', 'firstVisitDate', 'returnVisitDate', 'customerName', 'phone',
    'discussion', 'salesperson1', 'region',
}


def map_header_to_key(header):
    h = str(header).strip().lstrip('\ufeff')
    # Google 表單常見後綴／必填標記
    h = re.sub(r'[\s\*＊]+$', '', h)
    h = re.sub(r'[（(]必填[)）]$', '', h).strip()
    if h in LABEL_TO_KEY:
        return LABEL_TO_KEY[h]
    if h in ALL_FIELD_KEYS:
        return h
    # 模糊比對常見電話／姓名欄
    if '次要' in h and '電話' in h:
        return 'phoneSecondary'
    if '電話' in h or h.lower() in ('phone', 'mobile', 'tel'):
        return 'phone'
    if '姓名' in h or '客戶名' in h:
        return 'customerName'
    return None

TEMPLATE_HEADERS = [
    '案場', '客戶類型', '是否成交', '日期', '客戶姓名', '主要聯繫電話', '區域',
    '年齡', '職業', '總價預算', '媒體1', '媒體2', '洽談內容', '銷售人員1', '備註',
]


def parse_bool_deal(val):
    if val is None or str(val).strip() == '':
        return 0
    s = str(val).strip().lower()
    return 1 if s in ('是', '1', 'true', 'yes', 'y', '成交') else 0


def parse_cell_value(key, val):
    if val is None or str(val).strip() == '':
        return None
    s = str(val).strip()
    if key in MULTISELECT_KEYS:
        parts = re.split(r'[、;；,，]', s)
        return [p.strip() for p in parts if p.strip()]
    return s


def has_existing_customer(site_id, phone, exclude_record_id=None):
    normalized = normalize_phone(phone)
    if not normalized:
        return False
    conn = get_db()
    rows = conn.execute(
        'SELECT id, data FROM customers WHERE site_id = ?',
        (site_id,),
    ).fetchall()
    conn.close()
    for row in rows:
        if exclude_record_id and row['id'] == exclude_record_id:
            continue
        try:
            d = json.loads(row['data'])
        except Exception:
            continue
        if normalized == normalize_phone(d.get('phone', '')):
            return True
    return False


def infer_visit_type(system, data, exclude_record_id=None):
    vt = str(system.get('visit_type') or '').strip()
    # CSV／檔名／表單已明確指定時，直接採用，避免「新客」被誤判成「回訪」
    if vt in ('新客', '回訪'):
        return vt

    has_return_signals = any([
        data.get('returnVisitDate'),
        data.get('firstVisitDate'),
        data.get('prevVisitDate'),
        data.get('visitCount'),
    ])
    if has_return_signals or has_existing_customer(
        system.get('site_id'), data.get('phone'), exclude_record_id,
    ):
        return '回訪'
    return '新客'


def prepare_customer_system(site_id, data, visit_type=None, is_deal=None, exclude_record_id=None):
    site = get_site_by_id(site_id)
    if not site:
        return None, '找不到所選案場'
    if not data.get('customerName'):
        return None, '缺少客戶姓名'
    if not data.get('phone'):
        return None, '缺少主要電話'

    system = {
        'site_id': site_id,
        'site_name': site['name'],
        'visit_type': '新客',
        'is_deal': 0,
        'visit_date': None,
        'first_visit_date': data.get('firstVisitDate'),
        'return_visit_date': data.get('returnVisitDate'),
    }
    if visit_type in ('新客', '回訪'):
        system['visit_type'] = visit_type
    system['visit_type'] = infer_visit_type(system, data, exclude_record_id)

    if is_deal is None:
        system['is_deal'] = infer_deal_status(system, data)
    else:
        system['is_deal'] = infer_deal_status(system, data, explicit=bool(is_deal))

    vt = system['visit_type']
    system['visit_date'] = (
        data.get('returnVisitDate') if vt == '回訪' else data.get('visitDate')
    ) or datetime.now().strftime('%Y-%m-%d')

    apply_customer_status(data)

    return system, None


def apply_customer_status(data):
    if data.get('cancelDate') or data.get('cancelReason'):
        data['customerStatus'] = '退戶'
    elif data.get('customerStatus') == '退戶':
        pass
    elif '退戶' in str(data.get('remark') or ''):
        data['customerStatus'] = '退戶'
    elif not data.get('customerStatus'):
        data['customerStatus'] = '正常'


DEAL_KEYWORDS_STRICT = ('斡旋', '小訂', '足訂', '足定', '已下訂', '簽約')
DEAL_MARKERS = ('成交', '已購')
NOT_DEAL_MARKERS = ('未購', '未成交')


def field_text(data, key):
    val = data.get(key)
    if isinstance(val, list):
        return ' '.join(str(x) for x in val if x)
    return str(val or '').strip()


def infer_deal_status(system, data, explicit=None):
    if explicit is not None:
        return 1 if explicit else 0

    if int(system.get('is_deal') or 0) == 1:
        return 1

    purchased_text = field_text(data, 'purchasedReason')
    not_purchased_text = field_text(data, 'notPurchasedReason')

    if any(marker in not_purchased_text for marker in DEAL_MARKERS):
        return 1

    if purchased_text:
        if any(marker in purchased_text for marker in NOT_DEAL_MARKERS):
            if not any(marker in purchased_text for marker in DEAL_MARKERS):
                return 0
        elif not any(marker in purchased_text for marker in NOT_DEAL_MARKERS):
            return 1

    if not_purchased_text and not any(marker in not_purchased_text for marker in DEAL_MARKERS):
        return 0

    sincerity = field_text(data, 'sincerity')
    if sincerity and not any(marker in sincerity for marker in NOT_DEAL_MARKERS):
        if any(marker in sincerity for marker in DEAL_MARKERS + DEAL_KEYWORDS_STRICT):
            return 1

    for key in ('remark',):
        text = field_text(data, key)
        if not text:
            continue
        if any(marker in text for marker in DEAL_MARKERS):
            return 1
        if any(kw in text for kw in DEAL_KEYWORDS_STRICT):
            return 1

    discussion = field_text(data, 'discussion')
    if discussion:
        if any(kw in discussion for kw in DEAL_KEYWORDS_STRICT):
            return 1

    return 0


def resolve_site(site_name, default_site_id=None):
    by_name, by_id = get_site_maps()
    if site_name:
        name = str(site_name).strip()
        if name in by_name:
            return by_name[name]
        for s in load_sites():
            if s['name'] in name or name in s['name']:
                return s
    if default_site_id and default_site_id in by_id:
        return by_id[default_site_id]
    return None


def row_to_record(row_dict, default_site_id=None, default_visit_type=None):
    system = {
        'site_id': None, 'site_name': None, 'visit_type': '新客',
        'is_deal': 0, 'visit_date': None,
        'first_visit_date': None, 'return_visit_date': None,
    }
    data = {}
    has_visit_type = row_has_visit_type(row_dict)

    for header, raw_val in row_dict.items():
        if raw_val is None or str(raw_val).strip() == '':
            continue
        key = map_header_to_key(header)
        if not key:
            continue

        if key == '_site_name':
            system['site_name'] = str(raw_val).strip()
            continue
        if key == '_visit_type':
            system['visit_type'] = str(raw_val).strip() or '新客'
            continue
        if key == '_is_deal':
            system['is_deal'] = parse_bool_deal(raw_val)
            continue
        if key == '_visit_date':
            system['visit_date'] = str(raw_val).strip()
            continue
        if key == '_timestamp':
            normalized = normalize_date_value(raw_val)
            if normalized:
                system['_timestamp'] = normalized
            continue
        if key == '_created_at':
            continue

        parsed = parse_cell_value(key, raw_val)
        if parsed is not None:
            data[key] = parsed

    if not has_visit_type and default_visit_type in ('新客', '回訪'):
        system['visit_type'] = default_visit_type

    site = resolve_site(system['site_name'], default_site_id)
    if not site:
        return None, '找不到案場（請填寫正確案場名稱、於匯入頁選擇預設案場，或使用含案場名稱的檔名）'

    system['site_id'] = site['id']
    system['site_name'] = site['name']

    if data.get('phone'):
        data['phone'] = format_phone_for_storage(data['phone'])

    if not data.get('customerName'):
        return None, '缺少客戶姓名'
    if not data.get('phone'):
        return None, '缺少主要電話'

    normalize_record_dates(system, data)

    system['visit_type'] = infer_visit_type(system, data)
    system['is_deal'] = infer_deal_status(system, data)
    visit_type = system['visit_type']
    if not system['visit_date']:
        system['visit_date'] = (
            data.get('returnVisitDate') if visit_type == '回訪'
            else data.get('visitDate')
        ) or system.get('_timestamp') or datetime.now().strftime('%Y-%m-%d')

    system['first_visit_date'] = data.get('firstVisitDate')
    system['return_visit_date'] = data.get('returnVisitDate')
    apply_customer_status(data)

    system.pop('_timestamp', None)
    return {'system': system, 'data': data}, None


def update_customer_record(record_id, system, data):
    conn = get_db()
    cur = conn.execute('''
        UPDATE customers
        SET site_id = ?, site_name = ?, visit_type = ?, is_deal = ?, visit_date = ?,
            first_visit_date = ?, return_visit_date = ?, data = ?,
            updated_at = datetime('now', 'localtime')
        WHERE id = ?
    ''', (
        system['site_id'], system['site_name'], system['visit_type'],
        system['is_deal'], system['visit_date'],
        system.get('first_visit_date'), system.get('return_visit_date'),
        json.dumps(data, ensure_ascii=False),
        record_id,
    ))
    conn.commit()
    conn.close()
    return cur.rowcount > 0


def insert_customer_record(system, data):
    conn = get_db()
    cur = conn.execute('''
        INSERT INTO customers (site_id, site_name, visit_type, is_deal, visit_date,
                               first_visit_date, return_visit_date, data)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        system['site_id'], system['site_name'], system['visit_type'],
        system['is_deal'], system['visit_date'],
        system.get('first_visit_date'), system.get('return_visit_date'),
        json.dumps(data, ensure_ascii=False),
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return new_id


def auth_guard(permission=None):
    conn = get_db()
    user = get_current_user(conn)
    if not user:
        conn.close()
        return None, None, jsonify({'error': '請先登入', 'code': 'AUTH_REQUIRED'}), 401
    if permission and not user_has_permission(user, permission):
        conn.close()
        return None, None, jsonify({'error': '權限不足', 'code': 'FORBIDDEN'}), 403
    return conn, user, None


def ensure_site_access(user, site_id):
    if site_id and not user_can_access_site(user, site_id):
        return jsonify({'error': '無權存取此案場', 'code': 'FORBIDDEN'}), 403
    return None


def enrich_customer_record(user, record: dict) -> dict:
    """附加前端用權限旗標：編輯不限時間；刪除依 7 日／最高主管。"""
    record['canEdit'] = user_has_permission(user, 'edit_customers')
    record['canDelete'] = user_can_delete_customer(user, record.get('created_at'))
    return record


def fields_payload_for_site(conn, site_id: str) -> dict:
    overrides = load_site_option_overrides(conn, site_id)
    sections, sales_staff = apply_site_field_options(
        FIELD_SECTIONS, site_id, SALES_STAFF, overrides,
    )
    hidden = load_site_hidden_fields(conn, site_id)
    sections = apply_site_hidden_fields(sections, hidden)
    return {'sections': sections, 'salesStaff': sales_staff}


@app.before_request
def enforce_auth():
    path = request.path
    if path.startswith('/css/') or path.startswith('/js/'):
        return None
    if path in ('/login.html', '/favicon.ico'):
        return None
    if path == '/':
        return None

    if path in PROTECTED_PAGES:
        conn = get_db()
        user = get_current_user(conn)
        conn.close()
        if not user:
            return redirect(f'/login.html?next={path}')
        page_perms = {
            '/search.html': 'view_customers',
            '/sites.html': 'manage_sites',
            '/users.html': 'manage_users',
            '/site-fields.html': 'manage_field_options',
            '/field-options.html': 'manage_field_options',
            '/audit-log.html': 'view_audit_logs',
            '/weekly.html': 'manage_weekly_reports',
            '/sales.html': 'manage_weekly_reports',
        }
        need = page_perms.get(path)
        if need and not user_has_permission(user, need):
            return redirect('/')
        return None

    if path.startswith('/api/'):
        if is_public_api(path, request.method):
            return None
        conn = get_db()
        user = get_current_user(conn)
        conn.close()
        if not user:
            return jsonify({'error': '請先登入', 'code': 'AUTH_REQUIRED'}), 401
    return None


@app.route('/login.html')
def login_page():
    return send_from_directory('public', 'login.html')


@app.route('/users.html')
def users_page():
    return send_from_directory('public', 'users.html')


@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    body = request.get_json() or {}
    username = (body.get('username') or '').strip()
    password = body.get('password') or ''
    if not username or not password:
        return jsonify({'error': '請輸入帳號與密碼'}), 400

    conn = get_db()
    row = conn.execute(
        'SELECT * FROM users WHERE username = ?', (username,),
    ).fetchone()
    if not row or not row['is_active']:
        conn.close()
        return jsonify({'error': '帳號或密碼錯誤'}), 401
    if not verify_password(row['password_hash'], password):
        conn.close()
        return jsonify({'error': '帳號或密碼錯誤'}), 401

    session['user_id'] = row['id']
    user = get_user_by_username(conn, username)
    log_operation(conn, user, 'login', f'{user["displayName"]} 登入系統')
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'user': user})


@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    conn = get_db()
    user = get_current_user(conn)
    if user:
        log_operation(conn, user, 'logout', f'{user["displayName"]} 登出系統')
        conn.commit()
    conn.close()
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/me')
def auth_me():
    conn = get_db()
    user = get_current_user(conn)
    conn.close()
    if not user:
        return jsonify({'authenticated': False})
    return jsonify({'authenticated': True, 'user': user})


@app.route('/api/auth/roles')
def auth_roles():
    conn, user, err = auth_guard('manage_users')
    if err:
        return err
    conn.close()
    return jsonify({
        'roles': [{'id': k, 'label': v} for k, v in ROLES.items()],
    })


@app.route('/api/users')
def list_users():
    conn, user, err = auth_guard('manage_users')
    if err:
        return err
    rows = conn.execute('SELECT * FROM users ORDER BY username').fetchall()
    users = [get_user_by_id(conn, r['id']) for r in rows]
    conn.close()
    return jsonify(users)


@app.route('/api/users', methods=['POST'])
def create_user():
    conn, actor, err = auth_guard('manage_users')
    if err:
        return err
    body = request.get_json() or {}
    username = (body.get('username') or '').strip()
    password = body.get('password') or ''
    display_name = (body.get('displayName') or '').strip()
    role = (body.get('role') or '').strip()
    site_ids = body.get('siteIds') or []

    if not username or not password or not display_name or role not in ROLES:
        conn.close()
        return jsonify({'error': '請填寫完整資料並選擇有效職務'}), 400
    if role == 'field_staff' and not site_ids:
        conn.close()
        return jsonify({'error': '現場人員至少需指派一個案場'}), 400

    try:
        cur = conn.execute(
            'INSERT INTO users (username, password_hash, display_name, role) VALUES (?, ?, ?, ?)',
            (username, hash_password(password), display_name, role),
        )
        user_id = cur.lastrowid
        save_user_sites(conn, user_id, site_ids)
        log_operation(
            conn, actor, 'user_create',
            f'新增人員 {display_name}（{username}）',
            entity_type='user', entity_id=user_id,
            detail={'role': role, 'siteIds': site_ids},
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': '帳號已存在'}), 409
    new_user = get_user_by_username(conn, username)
    conn.close()
    return jsonify({'success': True, 'user': new_user}), 201


@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    conn, actor, err = auth_guard('manage_users')
    if err:
        return err
    row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到使用者'}), 404

    body = request.get_json() or {}
    display_name = (body.get('displayName') or row['display_name']).strip()
    role = (body.get('role') or row['role']).strip()
    is_active = body.get('isActive')
    password = body.get('password') or ''
    site_ids = body.get('siteIds')

    if role not in ROLES:
        conn.close()
        return jsonify({'error': '無效的職務'}), 400

    params = [display_name, role]
    sql = 'UPDATE users SET display_name = ?, role = ?, updated_at = datetime(\'now\', \'localtime\')'
    if is_active is not None:
        sql += ', is_active = ?'
        params.append(1 if is_active else 0)
    if password:
        sql += ', password_hash = ?'
        params.append(hash_password(password))
    sql += ' WHERE id = ?'
    params.append(user_id)
    conn.execute(sql, params)

    if site_ids is not None:
        if role == 'field_staff' and not site_ids:
            conn.close()
            return jsonify({'error': '現場人員至少需指派一個案場'}), 400
        save_user_sites(conn, user_id, site_ids)

    log_operation(
        conn, actor, 'user_update',
        f'更新人員 {display_name}（{row["username"]}）',
        entity_type='user', entity_id=user_id,
        detail={'role': role, 'isActive': is_active},
    )
    conn.commit()
    updated = get_user_by_username(conn, row['username'])
    conn.close()
    return jsonify({'success': True, 'user': updated})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    conn, actor, err = auth_guard('manage_users')
    if err:
        return err
    row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到使用者'}), 404
    if actor['id'] == user_id:
        conn.close()
        return jsonify({'error': '無法刪除自己的帳號'}), 400

    if row['role'] == 'executive' and row['is_active']:
        exec_count = conn.execute(
            "SELECT COUNT(*) FROM users WHERE role = 'executive' AND is_active = 1",
        ).fetchone()[0]
        if exec_count <= 1:
            conn.close()
            return jsonify({'error': '系統至少需保留一位啟用中的最高主管'}), 400

    display_name = row['display_name']
    username = row['username']
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    log_operation(
        conn, actor, 'user_delete',
        f'刪除人員 {display_name}（{username}）',
        entity_type='user', entity_id=user_id,
        detail={'role': row['role']},
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/')
def index():
    return send_from_directory('public', 'index.html')


@app.route('/search.html')
def search_page():
    return send_from_directory('public', 'search.html')


@app.route('/import.html')
def import_page():
    return send_from_directory('public', 'import.html')


@app.route('/weekly.html')
def weekly_page():
    return send_from_directory('public', 'weekly.html')


@app.route('/sales.html')
def sales_page():
    return send_from_directory('public', 'sales.html')


@app.route('/api/import/template')
def import_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow([
        '首學杭州', '新客', '否', '2026-07-01', '王小明', '0912345678', '大安區',
        '31-40歲', '上班族', '2000萬以下', 'FB', '介紹', '客戶對產品有興趣，需回去討論', '簡婉如', '',
    ])
    # ASCII fallback + RFC 5987 UTF-8 filename，避免中文檔名導致下載失敗
    utf8_name = quote('得意佳_客戶資料匯入範本.csv')
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': (
                "attachment; filename=\"deyijia_import_template.csv\"; "
                f"filename*=UTF-8''{utf8_name}"
            ),
            'Cache-Control': 'no-store',
        },
    )


@app.route('/api/customers/import', methods=['POST'])
def import_customers():
    conn = get_db()
    user = get_current_user(conn)
    default_site_id = request.form.get('defaultSiteId', '').strip() or None

    if 'file' not in request.files:
        conn.close()
        return jsonify({'error': '請上傳 CSV 檔案'}), 400

    file = request.files['file']
    if not file.filename:
        conn.close()
        return jsonify({'error': '請選擇檔案'}), 400

    inferred_site_id, inferred_visit_type = infer_from_filename(file.filename)
    if not default_site_id and inferred_site_id:
        default_site_id = inferred_site_id

    if default_site_id and user:
        denied = ensure_site_access(user, default_site_id)
        if denied:
            conn.close()
            return denied

    try:
        raw = file.read()
        for encoding in ('utf-8-sig', 'utf-8', 'big5', 'cp950'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return jsonify({'error': '無法讀取檔案編碼，請使用 UTF-8 或 Big5 編碼的 CSV'}), 400

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return jsonify({'error': 'CSV 檔案沒有欄位標題列'}), 400

        imported = 0
        errors = []

        for i, row in enumerate(reader, start=2):
            if not any(str(v).strip() for v in row.values() if v):
                continue
            record, err = row_to_record(row, default_site_id, inferred_visit_type)
            if err:
                errors.append({'row': i, 'message': err})
                continue
            if user:
                denied = ensure_site_access(user, record['system']['site_id'])
                if denied:
                    errors.append({'row': i, 'message': '無權匯入此案場資料'})
                    continue
            try:
                insert_customer_record(record['system'], record['data'])
                imported += 1
            except Exception as e:
                errors.append({'row': i, 'message': str(e)})

        site_name = None
        if default_site_id:
            site_row = conn.execute(
                'SELECT name FROM sites WHERE id = ?', (default_site_id,),
            ).fetchone()
            site_name = site_row['name'] if site_row else default_site_id
        if user:
            log_operation(
                conn, user, 'customer_import',
                f'匯入客戶資料：成功 {imported} 筆，失敗 {len(errors)} 筆'
                + (f'（{site_name}）' if site_name else ''),
                site_id=default_site_id, site_name=site_name,
                detail={'imported': imported, 'failed': len(errors), 'filename': file.filename},
            )
            conn.commit()

        return jsonify({
            'success': True,
            'imported': imported,
            'failed': len(errors),
            'errors': errors[:50],
            'inferredSiteId': inferred_site_id,
            'inferredVisitType': inferred_visit_type,
        })
    except csv.Error as e:
        return jsonify({'error': f'CSV 格式錯誤：{e}'}), 400
    except Exception as e:
        return jsonify({'error': f'匯入失敗：{e}'}), 500
    finally:
        conn.close()


@app.route('/api/sites')
def api_sites():
    conn = get_db()
    user = get_current_user(conn)
    sites = load_sites()
    # 業務與未登入一樣可看全部案場（填表用）；現場專案才依指派案場過濾
    if user and user['role'] == 'field_staff':
        allowed = get_allowed_site_ids(user)
        if allowed is not None:
            allowed_set = set(allowed)
            sites = [s for s in sites if s['id'] in allowed_set]
    conn.close()
    return jsonify(sites)


@app.route('/api/sites', methods=['POST'])
def create_site():
    conn, user, err = auth_guard('manage_sites')
    if err:
        return err
    body = request.get_json() or {}
    name = (body.get('name') or '').strip()
    group = (body.get('group') or 'residential').strip()
    if not name:
        return jsonify({'error': '請輸入案場名稱'}), 400
    if group not in ('residential', 'commercial'):
        group = 'residential'

    site_id = make_site_id(name)
    try:
        conn.execute(
            'INSERT INTO sites (id, name, group_type) VALUES (?, ?, ?)',
            (site_id, name, group),
        )
        log_operation(
            conn, user, 'site_create',
            f'新增案場 {name}',
            entity_type='site', entity_id=site_id,
            site_id=site_id, site_name=name,
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': '此案場名稱已存在'}), 409
    conn.close()
    site = {'id': site_id, 'name': name, 'group': group}
    return jsonify({'success': True, 'site': site}), 201


@app.route('/api/sites/<site_id>', methods=['DELETE'])
def delete_site(site_id):
    conn, user, err = auth_guard('manage_sites')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404

    count = conn.execute(
        'SELECT COUNT(*) FROM customers WHERE site_id = ?', (site_id,)
    ).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({
            'error': f'此案場已有 {count} 筆客戶資料，無法刪除',
        }), 400

    conn.execute('DELETE FROM sites WHERE id = ?', (site_id,))
    log_operation(
        conn, user, 'site_delete',
        f'刪除案場 {row["name"]}',
        entity_type='site', entity_id=site_id,
        site_id=site_id, site_name=row['name'],
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/sites.html')
def sites_page():
    return send_from_directory('public', 'sites.html')


@app.route('/field-options.html')
def field_options_index_page():
    return send_from_directory('public', 'field-options.html')


@app.route('/site-fields.html')
def site_fields_page():
    return send_from_directory('public', 'site-fields.html')


@app.route('/audit-log.html')
def audit_log_page():
    return send_from_directory('public', 'audit-log.html')


@app.route('/api/weekly/meta')
def weekly_meta():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    today = datetime.now().date()
    start = monday_of(today)
    end = start + timedelta(days=6)
    conn.close()
    return jsonify({
        'defaultWeekStart': start.isoformat(),
        'defaultWeekEnd': end.isoformat(),
        'defaultWeekNumber': default_week_number(start),
        'rocLabel': f'{roc_year(start)}/{start.month}/{start.day}-{roc_year(end)}/{end.month}/{end.day}',
    })


@app.route('/api/weekly/summary')
def weekly_summary():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err

    site_id = (request.args.get('siteId') or '').strip()
    week_start = (request.args.get('weekStart') or '').strip()
    if not site_id or not week_start:
        conn.close()
        return jsonify({'error': '請提供案場與週起始日（週一）'}), 400

    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied

    site = get_site_by_id(site_id)
    if not site:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404

    try:
        start, end = week_bounds(week_start)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

    week_start_s = start.isoformat()
    week_end_s = end.isoformat()
    saved = load_weekly_report(conn, site_id, week_start_s)
    base = empty_manual_payload(start, end)
    manual = merge_manual(base, (saved or {}).get('data') if saved else None)

    phone_sum = phone_total_from_manual(manual)
    active_staff = get_active_sales_staff(conn, site_id)
    auto = build_auto_stats(
        conn, site_id, start, end,
        included_visitor_ids=manual.get('includedVisitorIds'),
        active_staff=active_staff,
        week_phone_total=phone_sum,
    )
    auto = enrich_dims_with_phones(auto, manual.get('phoneCallsDetail'))
    suggested = aggregate_for_weekly(conn, site_id, start, end)
    history = list_weekly_reports(conn, site_id)
    conn.close()

    return jsonify({
        'siteId': site_id,
        'siteName': site['name'],
        'weekStart': week_start_s,
        'weekEnd': week_end_s,
        'weekNumber': manual.get('weekNumber') or default_week_number(start),
        'rocLabel': f'{roc_year(start)}/{start.month}/{start.day}-{roc_year(end)}/{end.month}/{end.day}',
        'saved': bool(saved),
        'updatedAt': (saved or {}).get('updatedAt'),
        'manual': manual,
        'auto': auto,
        'suggested': suggested,
        'activeStaff': active_staff,
        'derived': {
            'inventory': inventory_summary(manual),
            'commission': commission_summary(manual),
        },
        'history': history,
    })


@app.route('/api/weekly/reports', methods=['PUT'])
def save_weekly_report():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err

    body = request.get_json() or {}
    site_id = (body.get('siteId') or '').strip()
    week_start = (body.get('weekStart') or '').strip()
    manual = body.get('manual') or {}
    if not site_id or not week_start:
        conn.close()
        return jsonify({'error': '請提供案場與週起始日'}), 400

    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied

    site = get_site_by_id(site_id)
    if not site:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404

    try:
        start, end = week_bounds(week_start)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

    week_number = manual.get('weekNumber')
    try:
        week_number = int(week_number) if week_number not in (None, '') else default_week_number(start)
    except (TypeError, ValueError):
        week_number = default_week_number(start)
    manual['weekNumber'] = week_number
    manual['phoneCallsDetail'] = normalize_phone_calls_detail(manual.get('phoneCallsDetail'))
    # 若有來電明細，每日來電通數改由明細加總，避免兩套數字不一致。
    if manual['phoneCallsDetail']:
        by_day = {}
        for item in manual['phoneCallsDetail']:
            by_day[item['date']] = by_day.get(item['date'], 0) + float(item['count'] or 0)
        for day in manual.get('days') or []:
            day['phoneCalls'] = round(by_day.get(day.get('date'), 0), 2)

    report_id = upsert_weekly_report(
        conn,
        site_id=site_id,
        site_name=site['name'],
        week_start=start.isoformat(),
        week_end=end.isoformat(),
        week_number=week_number,
        data=manual,
        user_id=user.get('id') if user else None,
    )
    log_operation(
        conn,
        user,
        'save_weekly_report',
        f'儲存週報：{site["name"]} 第{week_number}週（{start.isoformat()}）',
        entity_type='weekly_report',
        entity_id=str(report_id),
        site_id=site_id,
        site_name=site['name'],
        detail={'weekStart': start.isoformat(), 'weekNumber': week_number},
    )
    conn.commit()
    saved = load_weekly_report(conn, site_id, start.isoformat())
    conn.close()
    return jsonify({'success': True, 'report': saved})


@app.route('/api/weekly/export.csv')
def export_weekly_csv():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err

    site_id = (request.args.get('siteId') or '').strip()
    week_start = (request.args.get('weekStart') or '').strip()
    if not site_id or not week_start:
        conn.close()
        return jsonify({'error': '請提供案場與週起始日'}), 400

    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied

    site = get_site_by_id(site_id)
    if not site:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404

    try:
        start, end = week_bounds(week_start)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

    saved = load_weekly_report(conn, site_id, start.isoformat())
    base = empty_manual_payload(start, end)
    manual = merge_manual(base, (saved or {}).get('data') if saved else None)
    phone_sum = phone_total_from_manual(manual)
    auto = build_auto_stats(
        conn, site_id, start, end,
        included_visitor_ids=manual.get('includedVisitorIds'),
        active_staff=get_active_sales_staff(conn, site_id),
        week_phone_total=phone_sum,
    )
    auto = enrich_dims_with_phones(auto, manual.get('phoneCallsDetail'))
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    week_no = manual.get('weekNumber') or default_week_number(start)
    writer.writerow([f'{site["name"]} 第{week_no}週週報'])
    writer.writerow(['週次區間', f'{start.isoformat()} ~ {end.isoformat()}'])
    writer.writerow([])
    writer.writerow(['【自動統計：來人】'])
    writer.writerow(['日期', '星期', '新客', '回訪', '成交', '合計'])
    for day in auto['byDay']:
        writer.writerow([
            day['date'], day['weekday'], day['new'], day['return'], day['deal'], day['total'],
        ])
    t = auto['totals']
    writer.writerow(['本週合計', '', t['new'], t['return'], t['deal'], t['total']])
    writer.writerow(['實際來人', t.get('actualTotal', ''), '納入週報', t.get('reportedTotal', '')])
    writer.writerow([])
    writer.writerow(['【來電明細】'])
    writer.writerow(['日期', '區域', '媒體', '通數'])
    for item in normalize_phone_calls_detail(manual.get('phoneCallsDetail')):
        writer.writerow([item.get('date'), item.get('region'), item.get('media'), item.get('count')])
    writer.writerow(['本週來電合計', phone_sum])
    writer.writerow([])
    writer.writerow(['【人工填寫：來電／天氣】'])
    writer.writerow(['日期', '星期', '天氣', '來電'])
    for day in manual.get('days') or []:
        writer.writerow([day.get('date'), day.get('weekday'), day.get('weather'), day.get('phoneCalls')])
    writer.writerow([])
    writer.writerow(['【成交／簽約／買進】'])
    writer.writerow(['項目', '戶', '車', '金額(萬)'])
    for label, key in [('本週成交', 'deals'), ('本週簽約', 'signings'), ('本週買進', 'purchases'), ('未報', 'unreported')]:
        block = manual.get(key) or {}
        writer.writerow([label, block.get('units', 0), block.get('parking', 0), block.get('amount', 0)])
    writer.writerow([])
    writer.writerow(['成交檢討', manual.get('reviewNotes') or ''])
    writer.writerow(['區域個案分析', manual.get('competitorNotes') or ''])
    writer.writerow(['備註', manual.get('memo') or ''])
    writer.writerow([])
    writer.writerow(['【區域統計（新客＋來電）】', '本週來人', '本週來電', '佔來人%', '佔來電%', '前期', '累計'])
    for row in auto['byRegion']:
        writer.writerow([
            row['name'], row.get('weekVisits'), row.get('weekPhones', 0),
            f"{row.get('weekVisitPct', 0)}%", f"{row.get('weekPhonePct', 0)}%",
            row.get('priorVisits'), row.get('cumVisits'),
        ])
    writer.writerow([])
    writer.writerow(['【媒體統計（新客＋來電）】', '本週來人', '本週來電', '佔來人%', '佔來電%', '前期', '累計'])
    for row in auto['byMedia']:
        writer.writerow([
            row['name'], row.get('weekVisits'), row.get('weekPhones', 0),
            f"{row.get('weekVisitPct', 0)}%", f"{row.get('weekPhonePct', 0)}%",
            row.get('priorVisits'), row.get('cumVisits'),
        ])
    writer.writerow([])
    writer.writerow(['【職業（新客）】', '本週', '前期', '累計', '佔本週%', '佔累計%'])
    for row in auto.get('byOccupation') or []:
        writer.writerow([
            row['name'], row.get('weekVisits'), row.get('priorVisits'), row.get('cumVisits'),
            f"{row.get('weekVisitPct', 0)}%", f"{row.get('cumVisitPct', 0)}%",
        ])
    writer.writerow([])
    writer.writerow(['【年齡（新客）】', '本週', '前期', '累計', '佔本週%', '佔累計%'])
    for row in auto.get('byAge') or []:
        writer.writerow([
            row['name'], row.get('weekVisits'), row.get('priorVisits'), row.get('cumVisits'),
            f"{row.get('weekVisitPct', 0)}%", f"{row.get('cumVisitPct', 0)}%",
        ])
    writer.writerow([])
    writer.writerow(['【成交比】'])
    writer.writerow(['銷售', '累計接待', '累計成交', '成交比', '成交金額', '退戶', '本週接待', '本週成交'])
    for row in auto.get('conversion') or []:
        writer.writerow([
            row['name'], row['visits'], row['deals'], row.get('rate') or '—', row.get('amount'),
            row.get('refunds'), row['weekVisits'], row['weekDeals'],
        ])
    writer.writerow([])
    writer.writerow(['【本週客況明細（納入週報）】'])
    writer.writerow(['日期', '類型', '姓名', '電話', '區域', '媒體', '職業', '年齡', '來源', '誠意度', '銷售', '洽談摘要'])
    for v in auto['visitors']:
        writer.writerow([
            v['date'], v['visitType'], v['customerName'], v['phone'], v['region'],
            v['media'], v.get('occupation'), v.get('age'), v['source'], v['sincerity'],
            v['salesperson1'], v['discussion'],
        ])

    utf8_name = quote(f'{site["name"]}_第{week_no}週週報_{start.isoformat()}.csv')
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': (
                f"attachment; filename=\"weekly_report_{start.isoformat()}.csv\"; "
                f"filename*=UTF-8''{utf8_name}"
            ),
        },
    )


@app.route('/api/weekly/export.xlsx')
def export_weekly_xlsx():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err

    site_id = (request.args.get('siteId') or '').strip()
    week_start = (request.args.get('weekStart') or '').strip()
    if not site_id or not week_start:
        conn.close()
        return jsonify({'error': '請提供案場與週起始日'}), 400

    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied

    site = get_site_by_id(site_id)
    if not site:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404

    try:
        start, end = week_bounds(week_start)
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

    saved = load_weekly_report(conn, site_id, start.isoformat())
    base = empty_manual_payload(start, end)
    manual = merge_manual(base, (saved or {}).get('data') if saved else None)
    phone_sum = phone_total_from_manual(manual)
    auto = build_auto_stats(
        conn, site_id, start, end,
        included_visitor_ids=manual.get('includedVisitorIds'),
        active_staff=get_active_sales_staff(conn, site_id),
        week_phone_total=phone_sum,
    )
    auto = enrich_dims_with_phones(auto, manual.get('phoneCallsDetail'))
    sales_summary = aggregate_for_weekly(conn, site_id, start, end)
    auto['commissionMatrix'] = sales_summary.get('commissionMatrix')
    conn.close()

    week_no = safe_week_number(manual.get('weekNumber'), start)
    try:
        content = build_weekly_excel(site['name'], start, end, week_no, manual, auto)
    except Exception as exc:
        app.logger.exception('weekly export.xlsx failed site=%s week=%s', site_id, week_start)
        return jsonify({'error': f'匯出 Excel 失敗：{exc}'}), 500
    utf8_name = quote(f'{site["name"]}_第{week_no}週週報_{start.isoformat()}.xlsx')
    return Response(
        content,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': (
                f"attachment; filename=\"weekly_report_{start.isoformat()}.xlsx\"; "
                f"filename*=UTF-8''{utf8_name}"
            ),
        },
    )


@app.route('/api/sales/meta')
def sales_meta():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    conn.close()
    site_id = (request.args.get('siteId') or '').strip()
    defaults = commission_defaults_for_site(site_id) if site_id else commission_defaults_for_site('_default')
    return jsonify({
        'recordTypes': [{'id': k, 'label': v} for k, v in SALES_RECORD_TYPES.items()],
        'commissionDefaults': defaults,
    })


@app.route('/api/sales/deals')
def api_list_sales_deals():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    record_type = (request.args.get('recordType') or '').strip() or None
    q = (request.args.get('q') or '').strip()
    try:
        limit = min(max(int(request.args.get('limit', 200)), 1), 1000)
        offset = max(int(request.args.get('offset', 0)), 0)
    except ValueError:
        limit, offset = 200, 0
    rows, total = list_sales_deals(
        conn, site_id, record_type=record_type, q=q, limit=limit, offset=offset,
    )
    staff = get_active_sales_staff(conn, site_id)
    conn.close()
    return jsonify({
        'total': total,
        'records': rows,
        'activeStaff': staff,
        'recordTypes': [{'id': k, 'label': v} for k, v in SALES_RECORD_TYPES.items()],
    })


def _sales_export_rows(conn, site_id):
    record_type = (request.args.get('recordType') or '').strip() or None
    q = (request.args.get('q') or '').strip()
    rows, _ = list_sales_deals(
        conn, site_id, record_type=record_type, q=q, limit=1000, offset=0,
    )
    return rows


def _sales_import_header(value):
    # 去掉空白、括號、斜線，讓「銷售日/(業主)」「合約銷售/總價」能對到別名
    return re.sub(r'[\s\u3000（）()/／]', '', str(value or '')).lower()


# 請佣總表常見重複欄名：底價區／階段區取最後一次（合約區改用前綴欄名）
_SALES_HEADER_KEEP_LAST = {
    _sales_import_header(x) for x in (
        '房底', '車底', '房底價', '車底價',
        '房車總底', '房車總底價', '房單價', '房底單價',
    )
}


SALES_IMPORT_ALIASES = {
    'recordType': ['類型', '狀態', '銷售階段'],
    'orderNo': ['訂單編號', '訂單'],
    'unitNo': ['戶別', '戶號', '戶別樓層'],
    'customerName': ['客戶', '客戶姓名', '姓名'],
    'phone': ['電話', '手機'],
    'productType': ['產品類型', '產品', '用途', '建物型態', '建物型態'],
    'areaPing': ['坪數', '建物坪數', '戶別坪數'],
    'parkingNo1': ['車位1', '車位號碼1', '車位編號1', '編號1', '汽車1'],
    'parkingNo2': ['車位2', '車位號碼2', '車位編號2', '編號2', '汽車2'],
    'parkingNos': ['車位號碼', '車位'],
    'houseSalePrice': [
        '合約房售價', '合約房地', '房售價', '房售價萬', '房屋成交價',
        '房售價未含其他費用', '實際房價',
    ],
    'parkingSalePrice': ['合約車售價', '合約車位', '車售價', '車位售價', '車位售價萬'],
    'totalPrice': [
        '合約銷售總價', '合約總價', '合約總價萬',
        '房車總價未含其他費用', '房車總價', '實際合約金額總價',
    ],
    'actualTotalPrice': [
        '實際成交總價萬', '實際成交總價',
        '房車售價未含附加及其他費用', '房車售價',
    ],
    'houseBasePrice': ['房底', '底價房地', '房底價'],
    'parkingBasePrice': ['車底', '底價車位', '車底價'],
    'basePrice': ['底價總價', '底總萬', '底總', '總底價', '房車總底', '房車總底價'],
    'surcharge': ['附加費', '其它費用', '其他費用'],
    'applianceGift': ['家電禮券', '家電禮'],
    'pickupVoucher': ['提貨券'],
    'decoration': ['裝潢', '裝潢費用'],
    'companyLoanInterest': ['公司貸利息', '公司貸利息', '公司貸利息'],
    'depositDate': [
        '訂金日期', '下訂日', '銷售日期', '銷售日',
        '回報下定日', '成交日期',
    ],
    'supplementDate': ['補足日', '補足日期'],
    'signDate': ['簽約日', '簽約日期', '實際簽約日'],
    'ownerSaleReportDate': ['業主報售日', '報售日', '銷售日業主'],
    'ownerSignReportDate': ['業主報簽日', '報簽日', '簽約日業主', '回報簽約日'],
    'salesperson1': ['銷售人員1', '銷售1', '銷售人員'],
    'salesperson2': ['銷售人員2', '銷售2'],
    'commissionBaseMode': ['請佣計價方式', '請佣方式'],
    'commissionSalesAmount': [
        '請佣銷售金額萬', '請佣銷售金額', '請佣$', '請佣銷售金額$',
    ],
    'commissionClaimable': ['可請佣萬', '可請佣', '可請佣金', '可請4.85%'],
    'commissionPayable': ['本期可請97%萬', '本期可請97%', '可請實際請款', '可請97%'],
    'commissionRetention': ['保留款3%萬', '保留款3%', '可請保留款'],
    'commissionClaimed': [
        '已請萬', '已請佣金額萬', '已請佣金額', '已請實際請款',
        '已請佣金', '已請97%',
    ],
    'commissionUnclaimed': [
        '未請萬', '未請金額萬', '未請金額', '未請實際請款',
        '未請佣金', '未請97%',
    ],
    'commissionStatus': ['請佣狀態', '請佣狀況'],
    'commissionPeriod': ['請佣期別'],
    'commissionClaimMonth': ['請佣月份', '請款月份'],
    'commissionClaimDate': ['請佣日期'],
    'commissionBooked': ['已入帳金額萬', '已入帳金額'],
    'nextMonthUnits': ['預計本月可請戶數', '預計可於本月請款戶數'],
    'nextMonthParking': ['預計本月可請車位', '預計可於本月請款車位'],
    'nextMonthClaimable': ['預計本月可請金額萬', '預計本月可請金額'],
    'memo': ['備註', '狀況備註'],
    'excessPrice': ['執底超價', '超價', '溢價'],
}


def _sales_import_value(row, normalized_headers, key):
    for alias in SALES_IMPORT_ALIASES.get(key, []):
        idx = normalized_headers.get(_sales_import_header(alias))
        if idx is not None and idx < len(row):
            value = row[idx]
            if value not in (None, ''):
                if isinstance(value, datetime):
                    return value.date().isoformat()
                return value
    return None


def _normalize_import_order_no(value) -> str:
    """Excel 訂單編號常為數字或 1120149.0，統一成可比對的字串。"""
    if value is None or value == '':
        return ''
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip().rstrip('0').rstrip('.')
    text = str(value).strip().replace(',', '').replace('，', '')
    if re.fullmatch(r'\d+\.0+', text):
        return text.split('.', 1)[0]
    return text


def _normalize_import_key_text(value) -> str:
    """戶別／客戶比對用：去空白與全形空白。"""
    return re.sub(r'[\s\u3000]+', '', str(value or '').strip())


def _sales_import_num(value):
    if value in (None, ''):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '').replace('，', '')
    text = re.sub(r'[^0-9.\-]', '', text)
    return float(text) if text not in ('', '-', '.', '-.') else 0.0


def _sales_record_type(value, sheet_name=''):
    text = str(value or sheet_name or '').strip()
    if '未報' in text:
        return 'unreported'
    if '退' in text or '換戶' in text:
        return 'refund'
    if '簽約' in text:
        return 'signing'
    if '買進' in text:
        return 'purchase'
    if text in SALES_RECORD_TYPES:
        return text
    return 'deal'


def _sales_import_payload(row, normalized_headers, sheet_name=''):
    get = lambda key: _sales_import_value(row, normalized_headers, key)
    unit_no = str(get('unitNo') or '').strip()
    customer_name = str(get('customerName') or '').strip()
    if not unit_no and not customer_name:
        return None
    if unit_no == '合計' or customer_name.endswith('筆'):
        return None

    payload = {
        key: get(key) for key in SALES_IMPORT_ALIASES
        if key not in {
            'recordType', 'parkingNos', 'actualTotalPrice', 'commissionClaimable',
            'commissionPayable', 'commissionRetention', 'commissionClaimed',
            'commissionUnclaimed', 'commissionStatus', 'commissionClaimMonth',
            'excessPrice',
        }
    }
    payload['recordType'] = _sales_record_type(get('recordType'), sheet_name)
    payload['unitNo'] = unit_no
    payload['customerName'] = customer_name
    payload['orderNo'] = _normalize_import_order_no(get('orderNo') or payload.get('orderNo'))

    parking_nos = str(get('parkingNos') or '').strip()
    if parking_nos and not payload.get('parkingNo1'):
        parts = [x.strip() for x in re.split(r'[、,，/／;；]+', parking_nos) if x.strip()]
        payload['parkingNo1'] = parts[0] if parts else ''
        payload['parkingNo2'] = parts[1] if len(parts) > 1 else ''

    contract = _sales_import_num(get('totalPrice'))
    actual = _sales_import_num(get('actualTotalPrice'))
    house = _sales_import_num(payload.get('houseSalePrice'))
    parking = _sales_import_num(payload.get('parkingSalePrice'))
    if house or parking:
        payload['houseSalePrice'] = house
        payload['parkingSalePrice'] = parking
    if contract:
        payload['totalPrice'] = contract
    surcharge_sum = sum(
        _sales_import_num(payload.get(k))
        for k in ('surcharge', 'applianceGift', 'pickupVoucher', 'decoration', 'companyLoanInterest')
    )
    has_parking = bool(payload.get('parkingNo1') or payload.get('parkingNo2'))
    if has_parking and parking <= 0 and contract > house > 0:
        diff = round(contract - house, 4)
        if diff > 0.01 and abs(diff - surcharge_sum) > 0.01:
            payload['parkingSalePrice'] = diff
    elif not payload.get('houseSalePrice') and contract:
        payload['totalPrice'] = contract
        if actual and contract > actual and surcharge_sum <= 0:
            payload['surcharge'] = contract - actual
    if not payload.get('houseSalePrice') and actual and not payload.get('parkingSalePrice'):
        payload['houseSalePrice'] = actual
    if not payload.get('houseBasePrice') and get('basePrice'):
        payload['basePrice'] = get('basePrice')
    if get('excessPrice') not in (None, ''):
        payload['excessPrice'] = _sales_import_num(get('excessPrice'))

    mode = str(payload.get('commissionBaseMode') or '').strip()
    payload['commissionBaseMode'] = 'deal' if '成交' in mode or mode == 'deal' else 'base'
    sales_amount = _sales_import_num(payload.get('commissionSalesAmount'))
    claimable = _sales_import_num(get('commissionClaimable'))
    payable = _sales_import_num(get('commissionPayable'))
    claimed = _sales_import_num(get('commissionClaimed'))
    unclaimed = _sales_import_num(get('commissionUnclaimed'))
    if sales_amount and claimable:
        payload['commissionDeduction'] = round(max(sales_amount * 0.0485 - claimable, 0), 4)

    # 請佣期別：數字 2 → 第2次服務費（對齊期別批次）
    period = str(get('commissionPeriod') or payload.get('commissionPeriod') or '').strip()
    claim_month = str(get('commissionClaimMonth') or '').strip()
    if period.isdigit():
        period = f'第{int(period)}次服務費'
    if period:
        payload['commissionPeriod'] = period
    # 已請有金額但沒期別 → 標成已請
    if claimed > 0 and not period:
        payload['commissionPeriod'] = claim_month or '已請'
    if claim_month and not payload.get('commissionClaimDate'):
        # 保留月份資訊於備註，避免誤解析成西元日期
        memo = str(payload.get('memo') or '').strip()
        note = f'請佣月份 {claim_month}'
        payload['memo'] = f'{memo}；{note}'.strip('；') if memo else note

    # 若檔案有明確可請／已請／未請金額，帶入 extra 供後續覆寫（compute 後再套）
    if claimable or payable or claimed or unclaimed:
        extra = payload.get('extra') if isinstance(payload.get('extra'), dict) else {}
        if claimable:
            extra['importClaimable'] = claimable
        if payable:
            extra['importPayable'] = payable
        if claimed:
            extra['importClaimed'] = claimed
        if unclaimed:
            extra['importUnclaimed'] = unclaimed
        payload['extra'] = extra

    for key in (
        'areaPing', 'houseSalePrice', 'parkingSalePrice', 'totalPrice',
        'houseBasePrice', 'parkingBasePrice', 'basePrice',
        'surcharge', 'applianceGift', 'pickupVoucher', 'decoration',
        'companyLoanInterest', 'commissionSalesAmount', 'commissionBooked',
        'nextMonthUnits', 'nextMonthParking', 'nextMonthClaimable',
    ):
        if payload.get(key) not in (None, ''):
            payload[key] = _sales_import_num(payload[key])
    for date_key in (
        'depositDate', 'supplementDate', 'signDate',
        'ownerSaleReportDate', 'ownerSignReportDate', 'commissionClaimDate',
    ):
        val = payload.get(date_key)
        if val not in (None, ''):
            if isinstance(val, datetime):
                payload[date_key] = val.date().isoformat()
            else:
                payload[date_key] = str(val).strip()[:10].replace('/', '-')
    return payload


def _merge_header_labels(parent_row, child_row):
    """合併請佣總表雙列表頭：可請+佣金 → 可請佣金；合約／底價區加前綴避免重複欄名。"""
    claim_groups = {'可請', '已請', '未請'}
    ambiguous = {'佣金', '保留款', '實際請款'}
    contract_groups = ('合約銷售', '合約金額')
    base_groups = ('底價', '階段請款', '階段執行')
    contract_child_map = {
        '房售價': '合約房售價',
        '車售價': '合約車售價',
        '房地': '合約房地',
        '車位': '合約車位',
        '總價': '合約總價',
    }
    base_child_map = {
        '總價': '底價總價',
        '房地': '底價房地',
        '車位': '底價車位',
    }
    n = max(len(parent_row or []), len(child_row or []))
    current_group = ''
    merged = []
    for i in range(n):
        p = parent_row[i] if parent_row and i < len(parent_row) else None
        c = child_row[i] if child_row and i < len(child_row) else None
        if p not in (None, ''):
            current_group = re.sub(r'[\s\u3000]+', '', str(p)).strip()
        child_raw = '' if c in (None, '') else str(c).strip()
        child = re.sub(r'[\s\u3000]+', '', child_raw).strip()
        group = current_group
        if group in claim_groups and child in ambiguous:
            label = f'{group}{child}'
        elif group in claim_groups and not child:
            label = group
        elif '銷售金額' in group and (not child or child == '$'):
            label = '請佣銷售金額'
        elif child:
            child_key = _sales_import_header(child_raw)
            if any(token in group for token in contract_groups):
                if '合約銷售' in child_key and '總價' in child_key:
                    label = '合約銷售總價'
                elif child in contract_child_map:
                    label = contract_child_map[child]
                else:
                    label = child_raw
            elif any(token in group for token in base_groups):
                label = base_child_map.get(child, child_raw)
            else:
                label = child_raw
        else:
            label = ''
        merged.append(label or None)
    return merged


def _header_map_from_row(row):
    normalized = {}
    for col, value in enumerate(row or []):
        if value in (None, ''):
            continue
        key = _sales_import_header(value)
        if not key:
            continue
        if key not in normalized or key in _SALES_HEADER_KEEP_LAST:
            normalized[key] = col
    return normalized


def _absorb_sales_subheader(rows, header_idx, normalized):
    """吸收請佣總表比率列／車位編號列（編號1、編號2、$、0.0485…）。"""
    nxt = header_idx + 1
    if nxt >= len(rows):
        return header_idx, normalized
    row = rows[nxt]
    labels = [
        _sales_import_header(v) for v in (row or [])
        if v not in (None, '')
    ]
    rate_like = sum(
        1 for v in (row or [])
        if isinstance(v, (int, float)) and not isinstance(v, bool) and 0 < float(v) < 1
    )
    is_sub = (
        any(t in ('編號1', '編號2', '$') for t in labels)
        or rate_like >= 3
    )
    if not is_sub:
        return header_idx, normalized
    for col, value in enumerate(row or []):
        if value in (None, ''):
            continue
        label = _sales_import_header(value)
        if label in ('編號1', '車位編號1'):
            normalized[_sales_import_header('車位1')] = col
        elif label in ('編號2', '車位編號2'):
            normalized[_sales_import_header('車位2')] = col
        elif label in ('$', '請佣$'):
            normalized[_sales_import_header('請佣銷售金額')] = col
    return nxt, normalized


def _find_sales_header(rows):
    for idx, row in enumerate(rows[:60]):
        # 優先：雙列表頭（上一列是群組：可請／已請／未請／請佣銷售金額）
        if idx > 0:
            parent = rows[idx - 1]
            parent_text = ' '.join(str(v) for v in (parent or []) if v not in (None, ''))
            if any(token in parent_text for token in ('可請', '已請', '未請', '請佣', '合約銷售', '銷售基本')):
                merged = _merge_header_labels(parent, row)
                normalized = _header_map_from_row(merged)
                keys = set(normalized)
                has_unit = any(_sales_import_header(x) in keys for x in SALES_IMPORT_ALIASES['unitNo'])
                has_name = any(_sales_import_header(x) in keys for x in SALES_IMPORT_ALIASES['customerName'])
                if has_unit and has_name:
                    header_idx, normalized = _absorb_sales_subheader(rows, idx, normalized)
                    return header_idx, normalized

        normalized = _header_map_from_row(row)
        keys = set(normalized)
        has_unit = any(_sales_import_header(x) in keys for x in SALES_IMPORT_ALIASES['unitNo'])
        has_name = any(_sales_import_header(x) in keys for x in SALES_IMPORT_ALIASES['customerName'])
        has_order = _sales_import_header('訂單編號') in keys
        if has_unit and (has_name or has_order):
            header_idx, normalized = _absorb_sales_subheader(rows, idx, normalized)
            return header_idx, normalized
    return None, None


def _is_sales_import_sheet(sheet_name: str, headers: dict) -> bool:
    name = str(sheet_name or '')
    # 請佣總表內的期別申請／摘要頁略過，改吃「累計銷售分析」
    skip_tokens = ('請佣no', '請佣no.', '已請未請', '折讓', '週報告', '請佣總表')
    lowered = name.replace(' ', '').lower()
    if any(tok.replace(' ', '') in lowered for tok in skip_tokens):
        if '累計銷售' not in name:
            return False
    if any(tok in name for tok in (
        '銷售總表', '未報明細', '已報明細', '退換戶明細', '累計銷售分析',
    )):
        return True
    if _sales_import_header('類型') in headers:
        return True
    if (
        _sales_import_header('訂單編號') in headers
        and any(_sales_import_header(x) in headers for x in SALES_IMPORT_ALIASES['unitNo'])
        and any(_sales_import_header(x) in headers for x in SALES_IMPORT_ALIASES['customerName'])
    ):
        return True
    return False


def _sales_import_tables(raw: bytes, filename: str):
    ext = Path(filename).suffix.lower()
    if ext == '.csv':
        text = None
        for encoding in ('utf-8-sig', 'utf-8', 'big5', 'cp950'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError('無法讀取 CSV 編碼')
        rows = list(csv.reader(io.StringIO(text)))
        return [('CSV', rows)]
    if ext not in ('.xlsx', '.xlsm'):
        raise ValueError('僅支援 .xlsx、.xlsm 或 .csv 檔案')
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    return [
        (ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
        for ws in wb.worksheets
    ]


@app.route('/api/sales/export.xlsx')
def api_export_sales_excel():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    denied = ensure_site_access(user, site_id) if site_id else None
    site = get_site_by_id(site_id) if site_id else None
    if not site_id or not site:
        conn.close()
        return jsonify({'error': '請選擇有效案場'}), 400
    if denied:
        conn.close()
        return denied
    rows = _sales_export_rows(conn, site_id)
    conn.close()
    content = build_sales_excel(site['name'], rows)
    utf8_name = quote(f'{site["name"]}_銷售總表.xlsx')
    return Response(
        content,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': (
                f"attachment; filename=\"sales_ledger.xlsx\"; filename*=UTF-8''{utf8_name}"
            ),
        },
    )


@app.route('/api/sales/export.csv')
def api_export_sales_csv():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    denied = ensure_site_access(user, site_id) if site_id else None
    site = get_site_by_id(site_id) if site_id else None
    if not site_id or not site:
        conn.close()
        return jsonify({'error': '請選擇有效案場'}), 400
    if denied:
        conn.close()
        return denied
    rows = _sales_export_rows(conn, site_id)
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        '類型', '訂單編號', '戶別', '客戶', '產品類型', '坪數', '車位1', '車位2',
        '房售價(萬)', '車位售價(萬)', '合約總價(萬)', '實際成交總價(萬)',
        '附加費(萬)', '家電禮券(萬)', '提貨券(萬)', '裝潢(萬)', '公司貸利息(萬)',
        '底總(萬)', '超價(萬)', '下訂日', '補足日', '簽約日', '請佣計價方式',
        '請佣銷售金額(萬)', '可請佣(萬)', '本期可請97%(萬)', '保留款3%(萬)',
        '已請(萬)', '未請(萬)', '狀態', '請佣期別', '請佣日期', '已入帳金額(萬)',
        '預計本月可請戶數', '預計本月可請車位', '預計本月可請金額(萬)',
        '業主報售日', '業主報簽日', '銷售人員1', '銷售人員2', '備註',
    ])
    total_fields = [
        'houseSalePrice', 'parkingSalePrice', 'contractTotal', 'actualTotalPrice',
        'surcharge', 'applianceGift', 'pickupVoucher', 'decoration', 'companyLoanInterest',
        'baseTotal', 'excessPrice', 'commissionSalesAmount',
        'commissionClaimable', 'commissionPayable', 'commissionRetention',
        'commissionClaimed', 'commissionUnclaimed', 'commissionBooked',
        'nextMonthUnits', 'nextMonthParking', 'nextMonthClaimable',
    ]
    totals = {key: 0.0 for key in total_fields}
    for row in rows:
        for key in total_fields:
            totals[key] += float(row.get(key) or 0)
        writer.writerow([
            row.get('recordTypeLabel') or row.get('recordType'), row.get('orderNo'),
            row.get('unitNo'), row.get('customerName'), row.get('productType'),
            row.get('areaPing'), row.get('parkingNo1'), row.get('parkingNo2'),
            row.get('houseSalePrice'), row.get('parkingSalePrice'),
            row.get('contractTotal'), row.get('actualTotalPrice'),
            row.get('surcharge'), row.get('applianceGift'), row.get('pickupVoucher'),
            row.get('decoration'), row.get('companyLoanInterest'),
            row.get('baseTotal'), row.get('excessPrice'),
            row.get('depositDate'), row.get('supplementDate'), row.get('signDate'),
            '成交價' if row.get('commissionBaseMode') == 'deal' else '底價',
            row.get('commissionSalesAmount'), row.get('commissionClaimable'),
            row.get('commissionPayable'), row.get('commissionRetention'),
            row.get('commissionClaimed'), row.get('commissionUnclaimed'),
            row.get('commissionStatus'), row.get('commissionPeriod'), row.get('commissionClaimDate'),
            row.get('commissionBooked'), row.get('nextMonthUnits'), row.get('nextMonthParking'),
            row.get('nextMonthClaimable'),
            row.get('ownerSaleReportDate') or row.get('reportDate'),
            row.get('ownerSignReportDate'),
            row.get('salesperson1'), row.get('salesperson2'), row.get('memo'),
        ])
    writer.writerow([
        '合計', '', '', f'{len(rows)} 筆', '', '', '', '',
        round(totals['houseSalePrice'], 4), round(totals['parkingSalePrice'], 4),
        round(totals['contractTotal'], 4), round(totals['actualTotalPrice'], 4),
        round(totals['surcharge'], 4), round(totals['applianceGift'], 4),
        round(totals['pickupVoucher'], 4), round(totals['decoration'], 4),
        round(totals['companyLoanInterest'], 4),
        round(totals['baseTotal'], 4), round(totals['excessPrice'], 4), '', '', '', '',
        round(totals['commissionSalesAmount'], 4),
        round(totals['commissionClaimable'], 4), round(totals['commissionPayable'], 4),
        round(totals['commissionRetention'], 4), round(totals['commissionClaimed'], 4),
        round(totals['commissionUnclaimed'], 4),
        '', '', '', round(totals['commissionBooked'], 4),
        round(totals['nextMonthUnits'], 4), round(totals['nextMonthParking'], 4),
        round(totals['nextMonthClaimable'], 4),
        '', '', '', '', '',
    ])
    utf8_name = quote(f'{site["name"]}_銷售總表.csv')
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': (
                f"attachment; filename=\"sales_ledger.csv\"; filename*=UTF-8''{utf8_name}"
            ),
        },
    )


@app.route('/api/sales/import', methods=['POST'])
def api_import_sales():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.form.get('siteId') or '').strip()
    if not site_id or not get_site_by_id(site_id):
        conn.close()
        return jsonify({'error': '請選擇有效案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    file = request.files.get('file')
    if not file or not file.filename:
        conn.close()
        return jsonify({'error': '請選擇要匯入的 Excel 或 CSV 檔案'}), 400

    imported = 0
    skipped = 0
    errors = []
    skipped_rows = []
    recognized_sheets = []
    try:
        tables = _sales_import_tables(file.read(), file.filename)
        sheet_titles = [name for name, _ in tables]
        prefer_cumulative = any('累計銷售分析' in name for name in sheet_titles)
        existing_rows = conn.execute(
            '''
            SELECT id, order_no, unit_no, customer_name
            FROM sales_deals WHERE site_id=?
            ''',
            (site_id,),
        ).fetchall()
        existing_orders = {}
        existing_unit_customer = {}
        for row in existing_rows:
            norm_order = _normalize_import_order_no(row['order_no'])
            if norm_order:
                existing_orders[norm_order] = row
            unit_key = _normalize_import_key_text(row['unit_no'])
            name_key = _normalize_import_key_text(row['customer_name'])
            if unit_key or name_key:
                existing_unit_customer[(unit_key, name_key)] = row
        seen_orders = {}
        seen_unit_customer = {}
        for sheet_name, rows in tables:
            header_idx, headers = _find_sales_header(rows)
            if header_idx is None or not headers:
                continue
            if prefer_cumulative and any(
                token in sheet_name for token in ('已報明細', '未報明細', '退換戶明細')
            ):
                # 同一本請佣總表已有累計銷售分析時，略過已報／未報避免重複
                continue
            if not _is_sales_import_sheet(sheet_name, headers):
                continue
            recognized_sheets.append(sheet_name)
            for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                try:
                    payload = _sales_import_payload(row, headers, sheet_name)
                    if not payload:
                        continue
                    order_no = _normalize_import_order_no(payload.get('orderNo'))
                    unit_no = str(payload.get('unitNo') or '').strip()
                    customer_name = str(payload.get('customerName') or '').strip()
                    unit_key = _normalize_import_key_text(unit_no)
                    name_key = _normalize_import_key_text(customer_name)
                    payload['orderNo'] = order_no
                    row_ref = f'{sheet_name}!{row_no}'
                    duplicate = None
                    skip_reason = ''
                    if order_no:
                        if order_no in seen_orders:
                            duplicate = seen_orders[order_no]
                            skip_reason = '檔案內訂單編號重複'
                        elif order_no in existing_orders:
                            duplicate = existing_orders[order_no]
                            skip_reason = '訂單編號已存在於系統'
                    elif unit_key or name_key:
                        pair_key = (unit_key, name_key)
                        if pair_key in seen_unit_customer:
                            duplicate = seen_unit_customer[pair_key]
                            skip_reason = '檔案內相同戶別＋客戶重複'
                        elif pair_key in existing_unit_customer:
                            duplicate = existing_unit_customer[pair_key]
                            skip_reason = '相同戶別＋客戶已存在於系統'
                    if duplicate:
                        skipped += 1
                        entry = {
                            'row': row_ref,
                            'orderNo': order_no,
                            'unitNo': unit_no,
                            'customerName': customer_name,
                            'reason': skip_reason,
                        }
                        if skip_reason.startswith('檔案內'):
                            entry['matchRow'] = duplicate if isinstance(duplicate, str) else None
                        else:
                            entry['existingId'] = duplicate['id']
                            entry['existingOrderNo'] = duplicate['order_no']
                            entry['existingUnitNo'] = duplicate['unit_no']
                            entry['existingCustomerName'] = duplicate['customer_name']
                        skipped_rows.append(entry)
                        continue
                    new_id = create_sales_deal(
                        conn, site_id, payload,
                        user_id=user.get('id') if user else None,
                    )
                    imported += 1
                    new_row = {
                        'id': new_id,
                        'order_no': order_no,
                        'unit_no': unit_no,
                        'customer_name': customer_name,
                    }
                    if order_no:
                        seen_orders[order_no] = row_ref
                        existing_orders[order_no] = new_row
                    elif unit_key or name_key:
                        pair_key = (unit_key, name_key)
                        seen_unit_customer[pair_key] = row_ref
                        existing_unit_customer[pair_key] = new_row
                except Exception as exc:
                    errors.append({
                        'row': f'{sheet_name}!{row_no}',
                        'message': str(exc),
                    })
        if not recognized_sheets:
            conn.rollback()
            return jsonify({
                'error': '找不到可匯入的銷售明細；請使用系統匯出檔，或請佣總表的「累計銷售分析」（含戶別、姓名／客戶）',
            }), 400
        log_operation(
            conn, user, 'sales_import',
            f'匯入銷售總表：新增 {imported} 筆、略過重複 {skipped} 筆、失敗 {len(errors)} 筆',
            entity_type='sales_deal', site_id=site_id,
            site_name=(get_site_by_id(site_id) or {}).get('name'),
            detail={
                'filename': file.filename, 'imported': imported,
                'skipped': skipped, 'failed': len(errors),
                'sheets': recognized_sheets,
            },
        )
        conn.commit()
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'skippedRows': skipped_rows[:50],
            'failed': len(errors),
            'errors': errors[:50],
            'sheets': recognized_sheets,
        })
    except ValueError as exc:
        conn.rollback()
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        conn.rollback()
        return jsonify({'error': f'匯入失敗：{exc}'}), 500
    finally:
        conn.close()


@app.route('/api/sales/deals', methods=['POST'])
def api_create_sales_deal():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    body = request.get_json() or {}
    site_id = (body.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    if not str(body.get('unitNo') or '').strip() and not str(body.get('customerName') or '').strip():
        conn.close()
        return jsonify({'error': '請至少填寫戶號或客戶姓名'}), 400
    deal_id = create_sales_deal(conn, site_id, body, user_id=user.get('id') if user else None)
    log_operation(
        conn, user, 'create_sales_deal',
        f'新增銷售明細：{body.get("unitNo") or ""} {body.get("customerName") or ""}',
        entity_type='sales_deal', entity_id=str(deal_id),
        site_id=site_id, site_name=(get_site_by_id(site_id) or {}).get('name'),
    )
    conn.commit()
    deal = get_sales_deal(conn, deal_id, site_id)
    conn.close()
    return jsonify({'success': True, 'record': deal})


@app.route('/api/sales/deals/<int:deal_id>', methods=['PUT'])
def api_update_sales_deal(deal_id):
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    body = request.get_json() or {}
    site_id = (body.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    ok = update_sales_deal(conn, deal_id, site_id, body, user_id=user.get('id') if user else None)
    if not ok:
        conn.close()
        return jsonify({'error': '找不到此筆資料'}), 404
    log_operation(
        conn, user, 'update_sales_deal',
        f'更新銷售明細 #{deal_id}',
        entity_type='sales_deal', entity_id=str(deal_id),
        site_id=site_id, site_name=(get_site_by_id(site_id) or {}).get('name'),
    )
    conn.commit()
    deal = get_sales_deal(conn, deal_id, site_id)
    conn.close()
    return jsonify({'success': True, 'record': deal})


@app.route('/api/sales/deals/<int:deal_id>', methods=['DELETE'])
def api_delete_sales_deal(deal_id):
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    ok = delete_sales_deal(conn, deal_id, site_id)
    if not ok:
        conn.close()
        return jsonify({'error': '找不到此筆資料'}), 404
    log_operation(
        conn, user, 'delete_sales_deal',
        f'刪除銷售明細 #{deal_id}',
        entity_type='sales_deal', entity_id=str(deal_id),
        site_id=site_id, site_name=(get_site_by_id(site_id) or {}).get('name'),
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/sales/deals/all', methods=['DELETE'])
def api_delete_all_sales_deals():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    if not user_can_bulk_delete_customers(user):
        conn.close()
        return jsonify({'error': '僅最高主管可清空銷售總表'}), 403
    body = request.get_json(silent=True) or {}
    site_id = (body.get('siteId') or request.args.get('siteId') or '').strip()
    if not site_id or not get_site_by_id(site_id):
        conn.close()
        return jsonify({'error': '請選擇有效案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    if body.get('confirm') != 'DELETE ALL':
        conn.close()
        return jsonify({'error': '請輸入正確確認碼 DELETE ALL'}), 400
    site = get_site_by_id(site_id)
    counts = delete_all_sales_deals(conn, site_id)
    log_operation(
        conn, user, 'sales_clear_site',
        f'清空「{site["name"]}」銷售總表：刪除明細 {counts["deals"]} 筆、期別 {counts["batches"]} 筆',
        entity_type='sales_deal',
        site_id=site_id, site_name=site['name'],
        detail=counts,
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, **counts})


@app.route('/api/sales/summary')
def api_sales_summary():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    week_start = (request.args.get('weekStart') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    if week_start:
        try:
            start, end = week_bounds(week_start)
        except ValueError as e:
            conn.close()
            return jsonify({'error': str(e)}), 400
    else:
        today = datetime.now().date()
        start, end = week_bounds(monday_of(today).isoformat())
    summary = aggregate_for_weekly(conn, site_id, start, end)
    conn.close()
    return jsonify({
        'siteId': site_id,
        'weekStart': start.isoformat(),
        'weekEnd': end.isoformat(),
        'summary': summary,
    })


@app.route('/api/sales/commission/batches')
def api_list_commission_batches():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    batches = list_commission_batches(conn, site_id)
    conn.close()
    return jsonify({'siteId': site_id, 'batches': batches})


@app.route('/api/sales/commission/batches', methods=['POST'])
def api_upsert_commission_batch():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    body = request.get_json(silent=True) or {}
    site_id = (body.get('siteId') or request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    try:
        batch_id = upsert_commission_batch(conn, site_id, body)
        conn.commit()
    except ValueError as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    batches = list_commission_batches(conn, site_id)
    batch = next((b for b in batches if b['id'] == batch_id), None)
    log_operation(
        conn, user, 'upsert_commission_batch',
        f'更新請佣期別：{body.get("periodName") or batch_id}',
        entity_type='commission_batch', entity_id=str(batch_id),
        site_id=site_id,
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'batch': batch, 'batches': batches})


@app.route('/api/sales/commission/batches/<int:batch_id>', methods=['DELETE'])
def api_delete_commission_batch(batch_id):
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    ok = delete_commission_batch(conn, site_id, batch_id)
    if not ok:
        conn.close()
        return jsonify({'error': '找不到此期別'}), 404
    log_operation(
        conn, user, 'delete_commission_batch',
        f'刪除請佣期別 #{batch_id}',
        entity_type='commission_batch', entity_id=str(batch_id),
        site_id=site_id,
    )
    conn.commit()
    batches = list_commission_batches(conn, site_id)
    conn.close()
    return jsonify({'success': True, 'batches': batches})


@app.route('/api/sales/commission/export.xlsx')
def api_export_commission_overview():
    conn, user, err = auth_guard('manage_weekly_reports')
    if err:
        return err
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        conn.close()
        return jsonify({'error': '請提供案場'}), 400
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    site = conn.execute('SELECT name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not site:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    today = datetime.now().date()
    start, end = week_bounds(monday_of(today).isoformat())
    summary = aggregate_for_weekly(conn, site_id, start, end)
    batches = list_commission_batches(conn, site_id)
    content = build_commission_overview_excel(
        site['name'],
        summary.get('commissionMatrix') or {},
        batches,
    )
    conn.close()
    from urllib.parse import quote
    utf8_name = quote(f'{site["name"]}_請佣總覽.xlsx')
    return Response(
        content,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': (
                f"attachment; filename=\"commission_overview.xlsx\"; filename*=UTF-8''{utf8_name}"
            ),
        },
    )


@app.route('/api/fields')
def api_fields():
    site_id = (request.args.get('siteId') or '').strip()
    if not site_id:
        return jsonify({'sections': FIELD_SECTIONS, 'salesStaff': SALES_STAFF})
    conn = get_db()
    row = conn.execute('SELECT id FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    payload = fields_payload_for_site(conn, site_id)
    conn.close()
    return jsonify(payload)


@app.route('/api/sites/<site_id>/field-options')
def get_site_field_options(site_id):
    conn, user, err = auth_guard('manage_field_options')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    overrides = load_site_option_overrides(conn, site_id)
    config = build_site_field_config(FIELD_SECTIONS, site_id, SALES_STAFF, overrides)
    hidden = load_site_hidden_fields(conn, site_id)
    config['fieldVisibility'] = build_site_field_visibility(FIELD_SECTIONS, site_id, hidden)
    export_saved = load_site_report_export_config(conn, site_id, REPORT_COLUMNS)
    export_config = build_site_report_export_config(REPORT_COLUMNS, export_saved)
    config['siteName'] = row['name']
    config['reportExport'] = export_config
    conn.close()
    return jsonify(config)


@app.route('/api/sites/<site_id>/field-options', methods=['PUT'])
def update_site_field_options(site_id):
    conn, user, err = auth_guard('manage_field_options')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    body = request.get_json() or {}

    if 'options' in body:
        normalized = normalize_save_payload(
            FIELD_SECTIONS, site_id, SALES_STAFF, body.get('options') or {},
        )
        save_site_option_overrides(conn, site_id, normalized)
        log_operation(
            conn, user, 'field_options_update',
            f'更新「{row["name"]}」欄位選項',
            site_id=site_id, site_name=row['name'],
            detail={'customizedFields': list(normalized.keys())},
        )

    if 'hiddenFields' in body:
        hidden = normalize_hidden_fields_payload(
            FIELD_SECTIONS, site_id, body.get('hiddenFields'),
        )
        save_site_hidden_fields(conn, site_id, hidden)
        log_operation(
            conn, user, 'field_options_update',
            f'更新「{row["name"]}」欄位顯示設定（隱藏 {len(hidden)} 項）',
            site_id=site_id, site_name=row['name'],
            detail={'hiddenFields': hidden},
        )

    # Backward compatible: old clients send options map at top level
    if 'options' not in body and 'hiddenFields' not in body:
        normalized = normalize_save_payload(
            FIELD_SECTIONS, site_id, SALES_STAFF, body,
        )
        save_site_option_overrides(conn, site_id, normalized)
        log_operation(
            conn, user, 'field_options_update',
            f'更新「{row["name"]}」欄位選項',
            site_id=site_id, site_name=row['name'],
            detail={'customizedFields': list(normalized.keys())},
        )

    conn.commit()
    overrides = load_site_option_overrides(conn, site_id)
    config = build_site_field_config(FIELD_SECTIONS, site_id, SALES_STAFF, overrides)
    hidden = load_site_hidden_fields(conn, site_id)
    config['fieldVisibility'] = build_site_field_visibility(FIELD_SECTIONS, site_id, hidden)
    export_saved = load_site_report_export_config(conn, site_id, REPORT_COLUMNS)
    config['reportExport'] = build_site_report_export_config(REPORT_COLUMNS, export_saved)
    config['siteName'] = row['name']
    conn.close()
    return jsonify({'success': True, **config})


@app.route('/api/report-columns')
def api_report_columns():
    return jsonify({'columns': REPORT_COLUMNS})


@app.route('/api/sites/<site_id>/field-order')
def get_site_field_order(site_id):
    conn, user, err = auth_guard('manage_field_options')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    saved = load_site_report_export_config(conn, site_id, REPORT_COLUMNS)
    config = build_site_report_export_config(REPORT_COLUMNS, saved)
    config['siteId'] = site_id
    config['siteName'] = row['name']
    conn.close()
    return jsonify(config)


@app.route('/api/sites/<site_id>/export-column-order')
def get_site_export_column_order(site_id):
    """供查看資料頁匯出報表時取得欄位順序（不需管理權限，需可查看該案場）。"""
    conn, user, err = auth_guard('view_customers')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    saved = load_site_report_export_config(conn, site_id, REPORT_COLUMNS)
    export_config = build_site_report_export_config(REPORT_COLUMNS, saved)
    conn.close()
    return jsonify({
        'siteId': site_id,
        'isCustomized': export_config['isCustomized'],
        'columnKeys': export_column_keys_for_site(REPORT_COLUMNS, saved),
    })


@app.route('/api/sites/<site_id>/field-order', methods=['PUT'])
def update_site_field_order(site_id):
    conn, user, err = auth_guard('manage_field_options')
    if err:
        return err
    row = conn.execute('SELECT id, name FROM sites WHERE id = ?', (site_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到此案場'}), 404
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied
    body = request.get_json() or {}
    normalized = normalize_report_export_payload(REPORT_COLUMNS, body)
    if normalized:
        save_site_report_export_config(conn, site_id, normalized)
    else:
        conn.execute('DELETE FROM site_field_order WHERE site_id = ?', (site_id,))
    enabled_count = len(export_column_keys_for_site(REPORT_COLUMNS, normalized))
    log_operation(
        conn, user, 'report_column_order_update',
        f'更新「{row["name"]}」報表匯出設定（{enabled_count} 個欄位）',
        site_id=site_id, site_name=row['name'],
    )
    conn.commit()
    saved = load_site_report_export_config(conn, site_id, REPORT_COLUMNS)
    config = build_site_report_export_config(REPORT_COLUMNS, saved)
    config['siteId'] = site_id
    config['siteName'] = row['name']
    conn.close()
    return jsonify({'success': True, **config})


@app.route('/api/audit-logs')
def list_audit_logs():
    conn, user, err = auth_guard('view_audit_logs')
    if err:
        return err
    page = max(1, int(request.args.get('page', 1)))
    limit = min(10000, max(1, int(request.args.get('limit', 50))))
    site_id = (request.args.get('siteId') or '').strip()
    action = (request.args.get('action') or '').strip()

    if site_id:
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied

    sql = 'SELECT * FROM operation_logs WHERE 1=1'
    params = []
    allowed_sites = get_allowed_site_ids(user)

    if site_id:
        sql += ' AND site_id = ?'
        params.append(site_id)
    elif allowed_sites is not None:
        if not allowed_sites:
            conn.close()
            return jsonify({'total': 0, 'page': page, 'limit': limit, 'records': []})
        placeholders = ','.join('?' * len(allowed_sites))
        sql += f' AND (site_id IN ({placeholders}) OR site_id IS NULL)'
        params.extend(allowed_sites)

    if action:
        sql += ' AND action = ?'
        params.append(action)

    count_sql = sql.replace('SELECT *', 'SELECT COUNT(*)', 1)
    total = conn.execute(count_sql, params).fetchone()[0]
    sql += ' ORDER BY id DESC LIMIT ? OFFSET ?'
    params.extend([limit, (page - 1) * limit])
    rows = [row_to_log_dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify({'total': total, 'page': page, 'limit': limit, 'records': rows})


@app.route('/api/customers/lookup')
def lookup_customer():
    phone = request.args.get('phone', '')
    site_id = request.args.get('siteId', '')
    if not phone or not site_id:
        return jsonify({'error': '請提供電話號碼與案場'}), 400

    conn = get_db()
    user = get_current_user(conn)
    if user:
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied

    normalized = normalize_phone(phone)
    if not normalized:
        conn.close()
        return jsonify({'found': False, 'count': 0, 'records': []})

    rows = conn.execute('''
        SELECT * FROM customers
        WHERE site_id = ?
        ORDER BY visit_date ASC, id ASC
    ''', (site_id,)).fetchall()
    conn.close()

    matches = []
    for row in rows:
        data = json.loads(row['data'])
        if normalized not in normalize_phone(data.get('phone', '')):
            continue
        if not data.get('visitDate') and row['visit_date']:
            data['visitDate'] = row['visit_date']
        if not data.get('firstVisitDate'):
            data['firstVisitDate'] = (
                data.get('visitDate') or row['first_visit_date'] or row['visit_date']
            )
        matches.append({
            'id': row['id'],
            'visitType': row['visit_type'],
            'visitDate': row['visit_date'],
            'siteName': row['site_name'],
            'customerName': data.get('customerName') or '',
            'salesperson1': data.get('salesperson1') or '',
            'salesperson2': data.get('salesperson2') or '',
            'data': data,
            'first_visit_date': row['first_visit_date'],
            'visit_date': row['visit_date'],
        })

    if not matches:
        return jsonify({'found': False, 'count': 0, 'records': []})

    # 預設帶入最早一筆「新客」；若無新客則用最早一筆
    primary = next((m for m in matches if m['visitType'] == '新客'), matches[0])
    return jsonify({
        'found': True,
        'count': len(matches),
        'record': primary,
        'records': matches,
    })


@app.route('/api/customers', methods=['POST'])
def create_customer():
    body = request.get_json()
    site_id = body.get('siteId')
    visit_type = body.get('visitType')
    data = body.get('data')
    if not site_id or not visit_type or not data:
        return jsonify({'error': '缺少必要欄位'}), 400

    conn = get_db()
    user = get_current_user(conn)
    # 填表為公開功能：未登入與業務皆可送出；現場專案仍受案場指派限制
    if user and user['role'] == 'field_staff':
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied

    is_deal = body.get('isDeal') if 'isDeal' in body else None
    system, err = prepare_customer_system(site_id, data, visit_type, is_deal)
    if err:
        conn.close()
        return jsonify({'error': err}), 400

    new_id = insert_customer_record(system, data)
    customer_name = str(data.get('customerName', '')).strip() or '（未填姓名）'
    if user:
        log_operation(
            conn, user, 'customer_create',
            f'新增客戶 {customer_name}（{system["site_name"]}）',
            entity_type='customer', entity_id=new_id,
            site_id=site_id, site_name=system['site_name'],
            detail={'visitType': visit_type},
        )
        conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': new_id})


@app.route('/api/customers')
def list_customers():
    conn, user, err = auth_guard('view_customers')
    if err:
        return err

    year = request.args.get('year', '')
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    site_id = request.args.get('siteId', '')
    visit_type = request.args.get('visitType', '')
    is_deal = request.args.get('isDeal', '')
    region = request.args.get('region', '')
    phone = request.args.get('phone', '')
    name = request.args.get('name', '')
    exclude_new = request.args.get('excludeNew', '') in ('1', 'true')
    exclude_return = request.args.get('excludeReturn', '') in ('1', 'true')
    exclude_deal = request.args.get('excludeDeal', '') in ('1', 'true')
    status_filter = request.args.get('customerStatus', '').strip()
    sort_order = (request.args.get('sortOrder', 'desc') or 'desc').lower()
    page = max(1, int(request.args.get('page', 1)))
    limit = min(10000, max(1, int(request.args.get('limit', 50))))

    allowed_sites = get_allowed_site_ids(user)
    if site_id:
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied

    sql = 'SELECT * FROM customers WHERE 1=1'
    params = []

    if year:
        sql += " AND strftime('%Y', visit_date) = ?"
        params.append(year)
    if start_date:
        sql += ' AND visit_date >= ?'
        params.append(start_date)
    if end_date:
        sql += ' AND visit_date <= ?'
        params.append(end_date)
    if site_id:
        sql += ' AND site_id = ?'
        params.append(site_id)
    elif allowed_sites is not None:
        if not allowed_sites:
            conn.close()
            return jsonify({'total': 0, 'page': page, 'limit': limit, 'records': []})
        placeholders = ','.join('?' * len(allowed_sites))
        sql += f' AND site_id IN ({placeholders})'
        params.extend(allowed_sites)
    if visit_type:
        sql += ' AND visit_type = ?'
        params.append(visit_type)
    if is_deal != '':
        sql += ' AND is_deal = ?'
        params.append(1 if is_deal in ('true', '1') else 0)

    all_rows = conn.execute(sql, params).fetchall()

    return_count_map = {}
    for row in all_rows:
        if row['visit_type'] != '回訪':
            continue
        data = json.loads(row['data'])
        key = (row['site_id'], normalize_phone(data.get('phone')))
        if key[1]:
            return_count_map[key] = return_count_map.get(key, 0) + 1

    filtered = []
    for row in all_rows:
        data = json.loads(row['data'])
        if exclude_new and row['visit_type'] == '新客':
            continue
        if exclude_return and row['visit_type'] == '回訪':
            continue
        if exclude_deal and int(row['is_deal']) == 1:
            continue
        if status_filter:
            row_status = str(data.get('customerStatus', '')).strip()
            if status_filter == '正常':
                if row_status == '退戶':
                    continue
            elif status_filter != row_status:
                continue
        if region and region not in str(data.get('region', '')):
            continue
        if phone and normalize_phone(phone) not in normalize_phone(data.get('phone', '')):
            continue
        if name and name not in str(data.get('customerName', '')):
            continue
        record = dict(row)
        phone_key = normalize_phone(data.get('phone'))
        record['return_visit_total'] = return_count_map.get((row['site_id'], phone_key), 0)
        if row['visit_type'] == '回訪' and not data.get('visitCount') and record['return_visit_total'] > 0:
            data['visitCount'] = f"第{record['return_visit_total']}次"
        # 統一日期顯示（修正 9/2/0114 等民國／匯入格式）
        for dk in ('visitDate', 'firstVisitDate', 'returnVisitDate', 'prevVisitDate'):
            if data.get(dk):
                normalized = normalize_date_value(data[dk])
                if normalized:
                    data[dk] = normalized
        record['data'] = data
        visit_norm = normalize_date_value(record.get('visit_date'))
        if visit_norm:
            record['visit_date'] = visit_norm
        filtered.append(record)

    newest_first = sort_order != 'asc'
    filtered.sort(
        key=lambda r: (r.get('visit_date') or '', r.get('id') or 0),
        reverse=newest_first,
    )

    total = len(filtered)
    start = (page - 1) * limit
    page_rows = [enrich_customer_record(user, dict(r)) for r in filtered[start:start + limit]]
    conn.close()

    return jsonify({'total': total, 'page': page, 'limit': limit, 'records': page_rows})


@app.route('/api/customers/<int:record_id>')
def get_customer(record_id):
    conn, user, err = auth_guard('view_customers')
    if err:
        return err
    row = conn.execute('SELECT * FROM customers WHERE id = ?', (record_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到資料'}), 404
    denied = ensure_site_access(user, row['site_id'])
    if denied:
        conn.close()
        return denied
    record = enrich_customer_record(user, dict(row))
    record['data'] = json.loads(record['data'])
    conn.close()
    return jsonify(record)


@app.route('/api/customers/<int:record_id>', methods=['PUT'])
def update_customer(record_id):
    conn, user, err = auth_guard('edit_customers')
    if err:
        return err
    body = request.get_json() or {}
    data = body.get('data')
    if not data:
        conn.close()
        return jsonify({'error': '缺少資料內容'}), 400

    row = conn.execute('SELECT * FROM customers WHERE id = ?', (record_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到資料'}), 404

    site_id = body.get('siteId') or row['site_id']
    denied = ensure_site_access(user, row['site_id'])
    if denied:
        conn.close()
        return denied
    if user['role'] == 'field_staff' and site_id != row['site_id']:
        conn.close()
        return jsonify({'error': '現場人員無法變更資料所屬案場'}), 403
    denied = ensure_site_access(user, site_id)
    if denied:
        conn.close()
        return denied

    visit_type = body.get('visitType') or row['visit_type']
    is_deal = body.get('isDeal') if 'isDeal' in body else bool(row['is_deal'])

    system, prep_err = prepare_customer_system(
        site_id, data, visit_type, is_deal, exclude_record_id=record_id,
    )
    if prep_err:
        conn.close()
        return jsonify({'error': prep_err}), 400

    if not update_customer_record(record_id, system, data):
        conn.close()
        return jsonify({'error': '更新失敗'}), 500
    customer_name = str(data.get('customerName', '')).strip() or '（未填姓名）'
    log_operation(
        conn, user, 'customer_update',
        f'編輯客戶 {customer_name}（{system["site_name"]}）',
        entity_type='customer', entity_id=record_id,
        site_id=site_id, site_name=system['site_name'],
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': record_id})


@app.route('/api/customers/<int:record_id>', methods=['DELETE'])
def delete_customer(record_id):
    conn, user, err = auth_guard('delete_customers')
    if err:
        return err
    row = conn.execute('SELECT * FROM customers WHERE id = ?', (record_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '找不到資料'}), 404
    denied = ensure_site_access(user, row['site_id'])
    if denied:
        conn.close()
        return denied
    if not user_can_delete_customer(user, row['created_at']):
        conn.close()
        return jsonify({
            'error': '此筆資料建檔已超過 7 日，無法刪除。若需刪除請聯繫最高主管。',
            'code': 'DELETE_WINDOW_EXPIRED',
        }), 403
    data = json.loads(row['data'])
    customer_name = str(data.get('customerName', '')).strip() or '（未填姓名）'
    cur = conn.execute('DELETE FROM customers WHERE id = ?', (record_id,))
    log_operation(
        conn, user, 'customer_delete',
        f'刪除客戶 {customer_name}（{row["site_name"]}）',
        entity_type='customer', entity_id=record_id,
        site_id=row['site_id'], site_name=row['site_name'],
    )
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        return jsonify({'error': '找不到資料'}), 404
    return jsonify({'success': True})


@app.route('/api/customers/all', methods=['DELETE'])
def delete_all_customers():
    conn, user, err = auth_guard('delete_customers')
    if err:
        return err
    if not user_can_bulk_delete_customers(user):
        conn.close()
        return jsonify({'error': '僅最高主管可清空案場或全部客戶資料'}), 403
    body = request.get_json() or {}
    if body.get('confirm') != 'DELETE ALL':
        conn.close()
        return jsonify({'error': '請輸入正確確認碼 DELETE ALL'}), 400

    site_id = (body.get('siteId') or '').strip()
    site_name = None
    if site_id:
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied
        site_row = conn.execute('SELECT name FROM sites WHERE id = ?', (site_id,)).fetchone()
        site_name = site_row['name'] if site_row else site_id
        cur = conn.execute('DELETE FROM customers WHERE site_id = ?', (site_id,))
    else:
        cur = conn.execute('DELETE FROM customers')
    deleted = cur.rowcount
    log_operation(
        conn, user, 'customer_clear_site',
        f'清空{site_name or "全部案場"}客戶資料：刪除 {deleted} 筆',
        site_id=site_id or None, site_name=site_name,
        detail={'deleted': deleted},
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'deleted': deleted})


@app.route('/api/stats')
def stats():
    conn, user, err = auth_guard('view_customers')
    if err:
        return err
    year = request.args.get('year', '')
    site_id = request.args.get('siteId', '')
    allowed_sites = get_allowed_site_ids(user)
    if site_id:
        denied = ensure_site_access(user, site_id)
        if denied:
            conn.close()
            return denied

    sql = '''
        SELECT site_name, visit_type, is_deal, COUNT(*) as count
        FROM customers WHERE 1=1
    '''
    params = []
    if year:
        sql += " AND strftime('%Y', visit_date) = ?"
        params.append(year)
    if site_id:
        sql += ' AND site_id = ?'
        params.append(site_id)
    elif allowed_sites is not None:
        if not allowed_sites:
            conn.close()
            return jsonify([])
        placeholders = ','.join('?' * len(allowed_sites))
        sql += f' AND site_id IN ({placeholders})'
        params.extend(allowed_sites)
    sql += ' GROUP BY site_name, visit_type, is_deal ORDER BY site_name'

    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)


init_db()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    print(f'客戶資料系統已啟動: http://localhost:{port}')
    app.run(host='0.0.0.0', port=port, debug=False)
