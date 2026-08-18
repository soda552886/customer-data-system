# -*- coding: utf-8 -*-
"""Weekly report helpers: stats from customers + draft persistence + Excel export."""
from __future__ import annotations

import io
import json
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


WEEKDAY_LABELS = ['一', '二', '三', '四', '五', '六', '日']
HOPE_SINCERITY = {'A', 'A+', 'A-', 'B', 'B+', '有望', '高'}
FORMER_SALES_LABEL = '前期銷售'
COMMISSION_RATE_DEFAULT = 0.0485
# 公司習慣字體；貼到 PPT／Word 時較少再手動改
EXCEL_FONT_NAME = '微軟正黑體'
EXCEL_FONT_SIZE = 11


def _font(**kwargs):
    kwargs.setdefault('name', EXCEL_FONT_NAME)
    kwargs.setdefault('size', EXCEL_FONT_SIZE)
    return Font(**kwargs)


def _conversion_ratio_label(visits, deals) -> str:
    """接待組數 ÷ 成交組數，呈現如 18.2:1（接 18.2 組成交 1 組）。"""
    v = _num(visits)
    d = _num(deals)
    if d <= 0:
        return '—'
    return f'{round(v / d, 1)}:1'


def _apply_sheet_font(ws, name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE):
    """整張工作表統一為微軟正黑體（保留粗體／顏色／原字級）。"""
    for row in ws.iter_rows():
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=name,
                size=f.size or size,
                bold=bool(f.bold),
                italic=bool(f.italic),
                color=f.color,
                strike=bool(f.strike),
                underline=f.underline,
            )


def _finalize_workbook_fonts(wb):
    for ws in wb.worksheets:
        _apply_sheet_font(ws)


def _to_zh_int(n: int) -> str:
    """把正整數轉成中文數字（供週次標題，如 86 → 八十六）。"""
    digits = '零一二三四五六七八九'
    n = int(n or 0)
    if n < 0:
        return str(n)
    if n < 10:
        return digits[n]
    if n < 20:
        return '十' + (digits[n - 10] if n > 10 else '')
    if n < 100:
        tens, ones = divmod(n, 10)
        return digits[tens] + '十' + (digits[ones] if ones else '')
    if n < 1000:
        hundreds, rest = divmod(n, 100)
        if rest == 0:
            return digits[hundreds] + '百'
        if rest < 10:
            return digits[hundreds] + '百零' + digits[rest]
        return digits[hundreds] + '百' + _to_zh_int(rest)
    return str(n)


def _fmt_num(val) -> str:
    try:
        x = float(val or 0)
    except (TypeError, ValueError):
        return str(val or 0)
    if abs(x - round(x)) < 1e-9:
        return f'{int(round(x)):,}'
    return f'{x:,.4f}'.rstrip('0').rstrip('.')


def _fmt_upc(units=0, parking=0, amount=0) -> str:
    return f'{_fmt_num(units)}戶／{_fmt_num(parking)}車／{_fmt_num(amount)}萬'


def _star_counts(rows, value_key='weekVisits') -> str:
    parts = []
    for r in rows or []:
        name = r.get('name')
        if not name or name == '合計':
            continue
        val = _num(r.get(value_key))
        if not val:
            continue
        parts.append(f'{name}*{_fmt_num(val)}')
    return '、'.join(parts) if parts else '—'


def _phone_star_counts(bucket: dict) -> str:
    parts = []
    for name, val in sorted((bucket or {}).items(), key=lambda x: (-x[1], x[0])):
        if not name or not val:
            continue
        parts.append(f'{name}*{_fmt_num(val)}')
    return '、'.join(parts) if parts else '—'


def init_weekly_tables(conn: sqlite3.Connection):
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS weekly_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id TEXT NOT NULL,
            site_name TEXT NOT NULL,
            week_number INTEGER,
            week_start TEXT NOT NULL,
            week_end TEXT NOT NULL,
            data TEXT NOT NULL DEFAULT '{}',
            created_by INTEGER,
            updated_by INTEGER,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            UNIQUE(site_id, week_start)
        );
        CREATE INDEX IF NOT EXISTS idx_weekly_reports_site ON weekly_reports(site_id);
        CREATE INDEX IF NOT EXISTS idx_weekly_reports_start ON weekly_reports(week_start);
    ''')


def parse_ymd(value: str):
    if not value:
        return None
    s = str(value).strip().split()[0].replace('/', '-')
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


def monday_of(d):
    return d - timedelta(days=d.weekday())


def week_bounds(week_start: str):
    start = parse_ymd(week_start)
    if not start:
        raise ValueError('無效的週起始日期')
    start = monday_of(start)
    end = start + timedelta(days=6)
    return start, end


def roc_year(d) -> int:
    return d.year - 1911


def normalize_week1_start(value) -> Optional[str]:
    """第 1 週起始日：對齊該週週一，回傳 YYYY-MM-DD。"""
    parsed = value if hasattr(value, 'toordinal') else parse_ymd(value)
    if not parsed:
        return None
    return monday_of(parsed).isoformat()


def default_week_number(start, origin=None) -> int:
    """有設定第一週起始日時，依間隔算第 N 週；否則用日曆 ISO 週次。"""
    start_d = start if hasattr(start, 'toordinal') else parse_ymd(str(start or ''))
    if not start_d:
        return 1
    origin_d = origin if hasattr(origin, 'toordinal') else parse_ymd(origin)
    if origin_d:
        return (monday_of(start_d) - monday_of(origin_d)).days // 7 + 1
    return int(start_d.isocalendar()[1])


def safe_week_number(week_number, start, origin=None) -> int:
    if week_number in (None, ''):
        return default_week_number(start, origin)
    try:
        return int(float(week_number))
    except (TypeError, ValueError):
        return default_week_number(start, origin)


def _parse_included_visitor_ids(raw) -> Optional[set]:
    """將週報儲存的 includedVisitorIds 轉成 int 集合；格式異常則略過。"""
    if raw is None:
        return None
    if not isinstance(raw, (list, tuple, set)):
        return set()
    out = set()
    for x in raw:
        try:
            if x not in (None, ''):
                out.add(int(x))
        except (TypeError, ValueError):
            continue
    return out


def empty_manual_payload(start, end, week_number=None, origin=None):
    days = []
    for i in range(7):
        d = start + timedelta(days=i)
        days.append({
            'date': d.isoformat(),
            'weekday': WEEKDAY_LABELS[i],
            'weather': '',
            'phoneCalls': 0,
        })
    return {
        'weekNumber': week_number if week_number is not None else default_week_number(start, origin),
        'days': days,
        # null = 尚未選取（預設納入全部本週來人）；[] = 刻意報 0 組；[id…] = 僅報選取組
        'includedVisitorIds': None,
        'deals': {'units': 0, 'parking': 0, 'amount': 0},
        'dealsCum': {'units': 0, 'parking': 0, 'amount': 0},
        'signings': {'units': 0, 'parking': 0, 'amount': 0},
        'signingsCum': {'units': 0, 'parking': 0, 'amount': 0},
        'purchases': {'units': 0, 'parking': 0, 'amount': 0},
        'purchasesCum': {'units': 0, 'parking': 0, 'amount': 0},
        'unreported': {'units': 0, 'parking': 0, 'amount': 0},
        'unsignedCum': {'units': 0, 'parking': 0, 'amount': 0},
        'commission': {
            'sellableUnits': 0,
            'sellableParking': 0,
            'sellableAmount': 0,
            'claimableAmount': 0,
            'payableAmount': 0,
            'retentionAmount': 0,
            'unclaimedAmount': 0,
            'claimedAmount': 0,
            'bookedAmount': 0,
            'claimableUnits': 0,
            'claimableParking': 0,
            'claimedUnits': 0,
            'claimedParking': 0,
            'nextMonthUnits': 0,
            'nextMonthParking': 0,
            'nextMonthAmount': 0,
        },
        'inventory': {
            'totalUnits': 122,
            'soldUnits': 0,
            'totalParking': 99,
            'soldParking': 0,
            'totalAmount': 175190,
            'soldAmount': 0,
            'soldBasePrice': 0,
            'residentialTotal': 80,
            'residentialSold': 0,
            'officeTotal': 42,
            'officeSold': 0,
            'shopTotal': 0,
            'shopSold': 0,
            'storefrontTotal': 0,
            'storefrontSold': 0,
        },
        'phoneCallsDetail': [],
        'reviewNotes': '',
        'competitorNotes': '',
        'memo': '',
    }


def merge_manual(base: dict, saved: Optional[dict]) -> dict:
    if not saved:
        return base
    out = json.loads(json.dumps(base))
    for key, val in saved.items():
        if isinstance(val, dict) and isinstance(out.get(key), dict):
            merged = dict(out[key])
            merged.update(val)
            out[key] = merged
        else:
            out[key] = val
    if isinstance(out.get('days'), list) and isinstance(base.get('days'), list):
        days = []
        saved_days = {d.get('date'): d for d in out['days'] if isinstance(d, dict)}
        for stub in base['days']:
            prev = saved_days.get(stub['date']) or {}
            days.append({**stub, **{k: prev[k] for k in ('weather', 'phoneCalls') if k in prev}})
        out['days'] = days
    return out


def _record_date(row, data: dict):
    raw = row['visit_date'] or data.get('returnVisitDate') or data.get('visitDate') or ''
    return parse_ymd(raw)


def _is_hope(sincerity: str) -> bool:
    s = (sincerity or '').strip().upper()
    if not s:
        return False
    if s in {x.upper() for x in HOPE_SINCERITY}:
        return True
    return '有望' in (sincerity or '')


def _truthy(val) -> bool:
    if isinstance(val, bool):
        return val
    s = str(val or '').strip().lower()
    return s in {'1', 'true', 'yes', 'y', '是', '有', '共同', '共同經營'}


def _num(val, default=0.0) -> float:
    try:
        if val is None or val == '':
            return float(default)
        return float(val)
    except (TypeError, ValueError):
        return float(default)


def _pct(part, whole, digits=1) -> float:
    if not whole:
        return 0.0
    return round(part / whole * 100, digits)


def _media_of(data: dict) -> str:
    media = data.get('media1') or data.get('media') or data.get('media2') or '未填'
    return str(media).strip() or '未填'


def _is_co_managed(data: dict) -> bool:
    if _truthy(data.get('isCoManaged')):
        return True
    s1 = str(data.get('salesperson1') or '').strip()
    s2 = str(data.get('salesperson2') or '').strip()
    return bool(s1 and s2)


def _sales_bucket(name: str, active_staff: Optional[set]) -> str:
    n = (name or '').strip() or '未填'
    if active_staff is not None and n not in active_staff and n != '未填':
        return FORMER_SALES_LABEL
    return n


def _credit_weight(data: dict) -> float:
    return 0.5 if _is_co_managed(data) else 1.0


def _dim_bucket():
    return {
        'priorVisits': 0.0, 'weekVisits': 0.0, 'cumVisits': 0.0,
        'priorDeals': 0.0, 'weekDeals': 0.0, 'cumDeals': 0.0,
        'priorAmount': 0.0, 'weekAmount': 0.0, 'cumAmount': 0.0,
    }


def _finalize_dimension(counter: dict, week_visits: float, week_deals: float,
                        cum_visits: float, cum_deals: float,
                        week_phones: float, cum_phones: float) -> list:
    rows = []
    for name, st in counter.items():
        st['cumVisits'] = st['priorVisits'] + st['weekVisits']
        st['cumDeals'] = st['priorDeals'] + st['weekDeals']
        st['cumAmount'] = st['priorAmount'] + st['weekAmount']
        rows.append({
            'name': name,
            'priorVisits': round(st['priorVisits'], 2),
            'weekVisits': round(st['weekVisits'], 2),
            'cumVisits': round(st['cumVisits'], 2),
            'priorDeals': round(st['priorDeals'], 2),
            'weekDeals': round(st['weekDeals'], 2),
            'cumDeals': round(st['cumDeals'], 2),
            'priorAmount': round(st['priorAmount'], 2),
            'weekAmount': round(st['weekAmount'], 2),
            'cumAmount': round(st['cumAmount'], 2),
            'weekVisitPct': _pct(st['weekVisits'], week_visits),
            'cumVisitPct': _pct(st['cumVisits'], cum_visits),
            'weekDealPct': _pct(st['weekDeals'], week_deals),
            'cumDealPct': _pct(st['cumDeals'], cum_deals),
            'weekPhonePct': _pct(st['weekVisits'], week_visits) if week_phones else 0.0,
            'cumPhonePct': _pct(st['cumVisits'], cum_visits) if cum_phones else 0.0,
            'count': round(st['weekVisits'], 2),
        })
    rows.sort(key=lambda x: (-x['weekVisits'], -x['cumVisits'], x['name']))
    if rows:
        tot_prior_v = sum(r['priorVisits'] for r in rows)
        tot_week_v = sum(r['weekVisits'] for r in rows)
        tot_cum_v = sum(r['cumVisits'] for r in rows)
        tot_prior_d = sum(r['priorDeals'] for r in rows)
        tot_week_d = sum(r['weekDeals'] for r in rows)
        tot_cum_d = sum(r['cumDeals'] for r in rows)
        tot_prior_a = sum(r['priorAmount'] for r in rows)
        tot_week_a = sum(r['weekAmount'] for r in rows)
        tot_cum_a = sum(r['cumAmount'] for r in rows)
        rows.append({
            'name': '合計',
            'priorVisits': round(tot_prior_v, 2),
            'weekVisits': round(tot_week_v, 2),
            'cumVisits': round(tot_cum_v, 2),
            'priorDeals': round(tot_prior_d, 2),
            'weekDeals': round(tot_week_d, 2),
            'cumDeals': round(tot_cum_d, 2),
            'priorAmount': round(tot_prior_a, 2),
            'weekAmount': round(tot_week_a, 2),
            'cumAmount': round(tot_cum_a, 2),
            'weekVisitPct': _pct(tot_week_v, week_visits),
            'cumVisitPct': _pct(tot_cum_v, cum_visits),
            'weekDealPct': _pct(tot_week_d, week_deals),
            'cumDealPct': _pct(tot_cum_d, cum_deals),
            'weekPhonePct': _pct(tot_week_v, week_visits) if week_phones else 0.0,
            'cumPhonePct': _pct(tot_cum_v, cum_visits) if cum_phones else 0.0,
            'count': round(tot_week_v, 2),
        })
    return rows


def build_auto_stats(
    conn: sqlite3.Connection,
    site_id: str,
    start,
    end,
    *,
    included_visitor_ids=None,
    active_staff=None,
    week_phone_total: float = 0,
) -> dict:
    """Build auto stats. included_visitor_ids=None means include all week visitors."""
    rows = conn.execute(
        'SELECT * FROM customers WHERE site_id = ?',
        (site_id,),
    ).fetchall()

    active_set = set(active_staff) if active_staff is not None else None
    include_set = _parse_included_visitor_ids(included_visitor_ids)

    day_keys = [(start + timedelta(days=i)).isoformat() for i in range(7)]
    by_day = {k: {'new': 0.0, 'return': 0.0, 'deal': 0.0, 'total': 0.0} for k in day_keys}

    dim_all = {
        'region': defaultdict(_dim_bucket),
        'media': defaultdict(_dim_bucket),
        'occupation': defaultdict(_dim_bucket),
        'age': defaultdict(_dim_bucket),
        'source': defaultdict(_dim_bucket),
        'purpose': defaultdict(_dim_bucket),
    }
    # 新客專用（區域／媒體／來源／職業／年齡／購屋用途報表）
    dim_new = {
        'region': defaultdict(_dim_bucket),
        'media': defaultdict(_dim_bucket),
        'source': defaultdict(_dim_bucket),
        'occupation': defaultdict(_dim_bucket),
        'age': defaultdict(_dim_bucket),
        'purpose': defaultdict(_dim_bucket),
    }

    visitors_all_week = []
    visitors = []
    return_visits = []
    deals = []
    hope_customers = []

    month_start = start.replace(day=1)
    year_start = start.replace(month=1, day=1)
    period = {
        'week': {'visits': 0.0, 'new': 0.0, 'return': 0.0, 'deals': 0.0, 'newDeals': 0.0, 'amount': 0.0},
        'month': {'visits': 0.0, 'new': 0.0, 'return': 0.0, 'deals': 0.0, 'newDeals': 0.0, 'amount': 0.0},
        'year': {'visits': 0.0, 'new': 0.0, 'return': 0.0, 'deals': 0.0, 'newDeals': 0.0, 'amount': 0.0},
        'all': {'visits': 0.0, 'new': 0.0, 'return': 0.0, 'deals': 0.0, 'newDeals': 0.0, 'amount': 0.0},
        'prior': {'visits': 0.0, 'new': 0.0, 'return': 0.0, 'deals': 0.0, 'newDeals': 0.0, 'amount': 0.0},
    }

    sales_stats = defaultdict(lambda: {
        'visits': 0.0, 'deals': 0.0, 'amount': 0.0, 'refunds': 0.0, 'refundAmount': 0.0,
        'weekVisits': 0.0, 'weekDeals': 0.0, 'weekAmount': 0.0, 'weekRefunds': 0.0,
    })

    def add_dim(store, key, weight, is_deal, amount, in_week, in_prior):
        if not key:
            key = '未填'
        st = store[key]
        if in_week:
            st['weekVisits'] += weight
            if is_deal:
                st['weekDeals'] += weight
                st['weekAmount'] += amount * weight
        elif in_prior:
            st['priorVisits'] += weight
            if is_deal:
                st['priorDeals'] += weight
                st['priorAmount'] += amount * weight

    for row in rows:
        data = json.loads(row['data'] or '{}')
        d = _record_date(row, data)
        vt = row['visit_type'] or ''
        is_new = vt != '回訪'
        is_deal = int(row['is_deal'] or 0) == 1
        is_refund = (
            _truthy(data.get('isRefund'))
            or bool(str(data.get('cancelDate') or '').strip())
            or str(data.get('customerStatus') or '').strip() == '退戶'
        )
        amount = _num(data.get('dealAmount'))
        refund_amount = _num(data.get('refundAmount'))
        weight = _credit_weight(data)
        staff1 = str(data.get('salesperson1') or '').strip() or '未填'
        staff2 = str(data.get('salesperson2') or '').strip()
        co = _is_co_managed(data)
        recipients = [staff1]
        if co and staff2:
            recipients = [staff1, staff2]
        elif co and not staff2:
            recipients = [staff1]

        region = str(data.get('region') or '未填').strip() or '未填'
        media = _media_of(data)
        occupation = str(data.get('occupation') or '未填').strip() or '未填'
        age = str(data.get('age') or '未填').strip() or '未填'
        source = str(data.get('customerSource') or '未填').strip() or '未填'
        purpose = str(
            data.get('purchasePurpose') or data.get('purchaseMotive') or data.get('purchaseNeed') or '未填'
        ).strip() or '未填'
        sincerity = str(data.get('sincerity') or '').strip()

        in_week = bool(d and start <= d <= end)
        in_prior = bool(d and d < start)
        in_month = bool(d and d >= month_start)
        in_year = bool(d and d >= year_start)

        # 週報選取：僅影響「本週」呈現與本週小計；前期累計仍吃全歷史
        week_included = True
        if in_week and include_set is not None:
            week_included = int(row['id']) in include_set

        item = {
            'id': row['id'],
            'date': d.isoformat() if d else '',
            'visitType': vt,
            'isDeal': is_deal,
            'isRefund': is_refund,
            'isCoManaged': co,
            'dealAmount': amount,
            'refundAmount': refund_amount,
            'customerName': data.get('customerName') or '',
            'phone': data.get('phone') or '',
            'region': region,
            'media': media,
            'occupation': occupation,
            'age': age,
            'source': source,
            'purpose': purpose,
            'sincerity': sincerity,
            'salesperson1': staff1,
            'salesperson2': staff2,
            'discussion': data.get('discussion') or '',
            'introUnit': data.get('introUnit') or '',
            'notPurchasedReason': (
                '、'.join(str(x) for x in data.get('notPurchasedReason') if x)
                if isinstance(data.get('notPurchasedReason'), (list, tuple))
                else str(data.get('notPurchasedReason') or '')
            ),
            'included': week_included if in_week else None,
        }

        if in_week:
            visitors_all_week.append(item)

        # —— 銷售成交比（僅新客；含共同經營拆分、前期銷售歸戶）——
        if d and is_new:
            for name in recipients:
                bucket = _sales_bucket(name, active_set)
                w = weight if len(recipients) == 1 else 0.5
                # 若只勾共同經營但沒填銷售2，仍算 0.5 給銷售1（另一半不歸戶）
                if co and len(recipients) == 1:
                    w = 0.5
                sales_stats[bucket]['visits'] += w
                if is_deal and not is_refund:
                    sales_stats[bucket]['deals'] += w
                    sales_stats[bucket]['amount'] += amount * w
                if is_refund:
                    sales_stats[bucket]['refunds'] += w
                    sales_stats[bucket]['refundAmount'] += refund_amount * w
                if in_week and week_included:
                    sales_stats[bucket]['weekVisits'] += w
                    if is_deal and not is_refund:
                        sales_stats[bucket]['weekDeals'] += w
                        sales_stats[bucket]['weekAmount'] += amount * w
                    if is_refund:
                        sales_stats[bucket]['weekRefunds'] += w

        if not d:
            continue

        # period / dims：本週僅計「有納入週報」者
        use_week = in_week and week_included

        def bump_period(key, w=1.0):
            period[key]['visits'] += w
            if is_new:
                period[key]['new'] += w
            else:
                period[key]['return'] += w
            if is_deal and not is_refund:
                period[key]['deals'] += w
                period[key]['amount'] += amount * w
                if is_new:
                    period[key]['newDeals'] += w

        bump_period('all')
        if in_year:
            bump_period('year')
        if in_month:
            bump_period('month')
        if in_prior:
            bump_period('prior')
        if use_week:
            bump_period('week')

        # 維度：前期 + 本週（本週受選取過濾）
        for dim_key, dim_val in (
            ('region', region), ('media', media),
            ('occupation', occupation), ('age', age), ('source', source),
            ('purpose', purpose),
        ):
            add_dim(dim_all[dim_key], dim_val, 1.0, is_deal and not is_refund, amount, use_week, in_prior)
            if is_new and dim_key in dim_new:
                add_dim(dim_new[dim_key], dim_val, 1.0, is_deal and not is_refund, amount, use_week, in_prior)

        if not use_week:
            continue

        key = d.isoformat()
        if is_new:
            by_day[key]['new'] += 1
        else:
            by_day[key]['return'] += 1
        by_day[key]['total'] += 1
        if is_deal and not is_refund:
            by_day[key]['deal'] += 1

        visitors.append(item)
        if not is_new:
            return_visits.append(item)
        if is_deal and not is_refund:
            deals.append(item)
        if _is_hope(sincerity):
            hope_customers.append(item)

    visitors_all_week.sort(key=lambda x: (x['date'], x['id']))
    visitors.sort(key=lambda x: (x['date'], x['id']))
    return_visits.sort(key=lambda x: (x['date'], x['id']))
    deals.sort(key=lambda x: (x['date'], x['id']))
    hope_customers.sort(key=lambda x: (x['date'], x['id']))

    totals = {
        'new': sum(v['new'] for v in by_day.values()),
        'return': sum(v['return'] for v in by_day.values()),
        'deal': sum(v['deal'] for v in by_day.values()),
        'total': sum(v['total'] for v in by_day.values()),
        'actualTotal': len(visitors_all_week),
        'reportedTotal': len(visitors),
    }

    week_visits = period['week']['visits']
    week_deals = period['week']['deals']
    cum_visits = period['prior']['visits'] + period['week']['visits']
    cum_deals = period['prior']['deals'] + period['week']['deals']
    # 新客專用分母（區域／媒體／來源表不含回訪）
    week_new = period['week']['new']
    week_new_deals = period['week']['newDeals']
    cum_new = period['prior']['new'] + period['week']['new']
    cum_new_deals = period['prior']['newDeals'] + period['week']['newDeals']
    week_phones = float(week_phone_total or 0)
    cum_phones = week_phones

    def dims_pack(src, week_v, week_d, cum_v, cum_d):
        return {
            key: _finalize_dimension(
                src[key], week_v, week_d, cum_v, cum_d, week_phones, cum_phones,
            )
            for key in src
        }

    conversion = []
    for name, st in sales_stats.items():
        visits = st['visits']
        deal_n = st['deals']
        conversion.append({
            'name': name,
            'visits': round(visits, 2),
            'deals': round(deal_n, 2),
            'amount': round(st['amount'], 2),
            'refunds': round(st['refunds'], 2),
            'refundAmount': round(st['refundAmount'], 2),
            'rate': _conversion_ratio_label(visits, deal_n),
            'ratio': round(visits / deal_n, 1) if deal_n else None,
            'weekVisits': round(st['weekVisits'], 2),
            'weekDeals': round(st['weekDeals'], 2),
            'weekAmount': round(st['weekAmount'], 2),
            'weekRefunds': round(st['weekRefunds'], 2),
        })
    conversion.sort(key=lambda x: (
        0 if x['name'] == FORMER_SALES_LABEL else 1,
        -x['deals'], -x['visits'], x['name'],
    ))
    # 合計列
    if conversion:
        tot = {
            'name': '合計',
            'visits': round(sum(x['visits'] for x in conversion), 2),
            'deals': round(sum(x['deals'] for x in conversion), 2),
            'amount': round(sum(x['amount'] for x in conversion), 2),
            'refunds': round(sum(x['refunds'] for x in conversion), 2),
            'refundAmount': round(sum(x['refundAmount'] for x in conversion), 2),
            'weekVisits': round(sum(x['weekVisits'] for x in conversion), 2),
            'weekDeals': round(sum(x['weekDeals'] for x in conversion), 2),
            'weekAmount': round(sum(x['weekAmount'] for x in conversion), 2),
            'weekRefunds': round(sum(x['weekRefunds'] for x in conversion), 2),
        }
        tot['rate'] = _conversion_ratio_label(tot['visits'], tot['deals'])
        tot['ratio'] = round(tot['visits'] / tot['deals'], 1) if tot['deals'] else None
        conversion.append(tot)

    all_dims = dims_pack(dim_all, week_visits, week_deals, cum_visits, cum_deals)
    new_dims = dims_pack(dim_new, week_new, week_new_deals, cum_new, cum_new_deals)

    return {
        'byDay': [
            {
                'date': k,
                'weekday': WEEKDAY_LABELS[i],
                'new': by_day[k]['new'],
                'return': by_day[k]['return'],
                'deal': by_day[k]['deal'],
                'total': by_day[k]['total'],
            }
            for i, k in enumerate(day_keys)
        ],
        'totals': totals,
        'period': period,
        # 區域／媒體／來源／職業／年齡皆僅統計新客
        'byRegion': new_dims['region'],
        'byMedia': new_dims['media'],
        'bySource': new_dims['source'],
        'byOccupation': new_dims['occupation'],
        'byAge': new_dims['age'],
        'byPurpose': new_dims['purpose'],
        'dimensions': {
            'all': all_dims,
            'newOnly': new_dims,
        },
        'conversion': conversion,
        'visitors': visitors,
        'visitorsAllWeek': visitors_all_week,
        'returnVisits': return_visits,
        'hopeCustomers': hope_customers,
        'dealsFromCustomers': deals,
    }


def inventory_summary(manual: dict) -> dict:
    inv = manual.get('inventory') or {}
    total_u = _num(inv.get('totalUnits'))
    sold_u = _num(inv.get('soldUnits'))
    total_p = _num(inv.get('totalParking'))
    sold_p = _num(inv.get('soldParking'))
    total_a = _num(inv.get('totalAmount'))
    sold_a = _num(inv.get('soldAmount'))
    sold_base = _num(inv.get('soldBasePrice'))
    res_t = _num(inv.get('residentialTotal'))
    res_s = _num(inv.get('residentialSold'))
    off_t = _num(inv.get('officeTotal'))
    off_s = _num(inv.get('officeSold'))
    shop_t = _num(inv.get('shopTotal'))
    shop_s = _num(inv.get('shopSold'))
    store_t = _num(inv.get('storefrontTotal'))
    store_s = _num(inv.get('storefrontSold'))
    # 未售金額以底價：總底價金額 - 已售底價（totalAmount 視為總底價）
    remain_base = max(total_a - sold_base, 0)
    return {
        'unitRate': round(sold_u / total_u * 100, 2) if total_u else 0,
        'parkingRate': round(sold_p / total_p * 100, 2) if total_p else 0,
        'amountRate': round(sold_a / total_a * 100, 2) if total_a else 0,
        'basePriceRate': round(sold_base / total_a * 100, 2) if total_a else 0,
        'residentialRate': round(res_s / res_t * 100, 2) if res_t else 0,
        'officeRate': round(off_s / off_t * 100, 2) if off_t else 0,
        'shopRate': round(shop_s / shop_t * 100, 2) if shop_t else 0,
        'storefrontRate': round(store_s / store_t * 100, 2) if store_t else 0,
        'remainUnits': max(total_u - sold_u, 0),
        'remainParking': max(total_p - sold_p, 0),
        'remainAmount': max(total_a - sold_a, 0),
        'remainBasePrice': remain_base,
    }


def commission_summary(manual: dict) -> dict:
    c = manual.get('commission') or {}
    claimable_amt = _num(c.get('claimableAmount'))
    claimed_amt = _num(c.get('claimedAmount'))
    claimable_u = _num(c.get('claimableUnits'))
    claimed_u = _num(c.get('claimedUnits'))
    claimable_p = _num(c.get('claimableParking'))
    claimed_p = _num(c.get('claimedParking'))
    unclaimed_amt = _num(c.get('unclaimedAmount'))
    if not unclaimed_amt:
        unclaimed_amt = round(max(claimable_amt - claimed_amt, 0), 4)
    unclaimed_u = _num(c.get('unclaimedUnits'))
    if not unclaimed_u:
        unclaimed_u = max(claimable_u - claimed_u, 0)
    unclaimed_p = _num(c.get('unclaimedParking'))
    if not unclaimed_p:
        unclaimed_p = max(claimable_p - claimed_p, 0)
    return {
        'unclaimedAmount': round(unclaimed_amt, 4),
        'unclaimedUnits': unclaimed_u,
        'unclaimedParking': unclaimed_p,
        'payableAmount': round(_num(c.get('payableAmount')) or claimable_amt * 0.97, 4),
        'retentionAmount': round(_num(c.get('retentionAmount')) or claimable_amt * 0.03, 4),
        'bookedAmount': round(_num(c.get('bookedAmount')), 4),
        'nextMonthUnits': _num(c.get('nextMonthUnits')),
        'nextMonthParking': _num(c.get('nextMonthParking')),
        'nextMonthAmount': round(_num(c.get('nextMonthAmount')), 4),
    }


def load_weekly_report(conn: sqlite3.Connection, site_id: str, week_start: str):
    row = conn.execute(
        'SELECT * FROM weekly_reports WHERE site_id = ? AND week_start = ?',
        (site_id, week_start),
    ).fetchone()
    if not row:
        return None
    return {
        'id': row['id'],
        'siteId': row['site_id'],
        'siteName': row['site_name'],
        'weekNumber': row['week_number'],
        'weekStart': row['week_start'],
        'weekEnd': row['week_end'],
        'data': json.loads(row['data'] or '{}'),
        'updatedAt': row['updated_at'],
    }


def upsert_weekly_report(
    conn: sqlite3.Connection,
    *,
    site_id: str,
    site_name: str,
    week_start: str,
    week_end: str,
    week_number: Optional[int],
    data: dict,
    user_id: Optional[int],
):
    existing = conn.execute(
        'SELECT id FROM weekly_reports WHERE site_id = ? AND week_start = ?',
        (site_id, week_start),
    ).fetchone()
    payload = json.dumps(data, ensure_ascii=False)
    if existing:
        conn.execute(
            '''
            UPDATE weekly_reports
            SET site_name = ?, week_number = ?, week_end = ?, data = ?,
                updated_by = ?, updated_at = datetime('now', 'localtime')
            WHERE id = ?
            ''',
            (site_name, week_number, week_end, payload, user_id, existing['id']),
        )
        return existing['id']
    cur = conn.execute(
        '''
        INSERT INTO weekly_reports
          (site_id, site_name, week_number, week_start, week_end, data, created_by, updated_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (site_id, site_name, week_number, week_start, week_end, payload, user_id, user_id),
    )
    return cur.lastrowid


def list_weekly_reports(conn: sqlite3.Connection, site_id: str, limit: int = 30):
    rows = conn.execute(
        '''
        SELECT id, site_id, site_name, week_number, week_start, week_end, updated_at
        FROM weekly_reports
        WHERE site_id = ?
        ORDER BY week_start DESC
        LIMIT ?
        ''',
        (site_id, limit),
    ).fetchall()
    return [{
        'id': r['id'],
        'siteId': r['site_id'],
        'siteName': r['site_name'],
        'weekNumber': r['week_number'],
        'weekStart': r['week_start'],
        'weekEnd': r['week_end'],
        'updatedAt': r['updated_at'],
    } for r in rows]


def _style_header(ws, row=1):
    header_font = _font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A4D7C')
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def normalize_phone_calls_detail(items) -> list:
    """來電明細：日期／區域／媒體／通數。"""
    out = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        count = max(_num(item.get('count'), 1), 0)
        if count <= 0:
            continue
        region = str(item.get('region') or '').strip() or '未填'
        media = str(item.get('media') or '').strip() or '未填'
        date_s = parse_ymd(item.get('date'))
        out.append({
            'date': date_s.isoformat() if date_s else str(item.get('date') or '').strip()[:10],
            'region': region,
            'media': media,
            'count': round(count, 2),
        })
    return out


def phone_total_from_manual(manual: dict) -> float:
    detail = normalize_phone_calls_detail(manual.get('phoneCallsDetail'))
    if detail:
        return round(sum(_num(x.get('count')) for x in detail), 2)
    return round(sum(_num(d.get('phoneCalls')) for d in (manual.get('days') or [])), 2)


def enrich_dims_with_phones(auto: dict, phone_details) -> dict:
    """把來電明細併入區域／媒體表，與來人同表呈現。"""
    detail = normalize_phone_calls_detail(phone_details)
    by_region = defaultdict(float)
    by_media = defaultdict(float)
    by_day = defaultdict(float)
    total = 0.0
    for item in detail:
        count = _num(item.get('count'))
        by_region[item['region']] += count
        by_media[item['media']] += count
        if item.get('date'):
            by_day[item['date']] += count
        total += count

    def patch(rows, bucket):
        rows = [dict(r) for r in (rows or [])]
        existing = {r.get('name') for r in rows}
        insert_at = len(rows) - 1 if rows and rows[-1].get('name') == '合計' else len(rows)
        for name, phones in sorted(bucket.items(), key=lambda x: (-x[1], x[0])):
            if name in existing:
                continue
            rows.insert(insert_at, {
                'name': name,
                'priorVisits': 0, 'weekVisits': 0, 'cumVisits': 0,
                'priorDeals': 0, 'weekDeals': 0, 'cumDeals': 0,
                'priorAmount': 0, 'weekAmount': 0, 'cumAmount': 0,
                'weekVisitPct': 0, 'cumVisitPct': 0,
                'weekDealPct': 0, 'cumDealPct': 0,
                'count': 0,
            })
            insert_at += 1
            existing.add(name)
        for r in rows:
            if r.get('name') == '合計':
                phones = total
            else:
                phones = bucket.get(r.get('name'), 0)
            r['weekPhones'] = round(phones, 2)
            r['weekPhonePct'] = _pct(phones, total)
        return rows

    auto = dict(auto or {})
    auto['byRegion'] = patch(auto.get('byRegion'), by_region)
    auto['byMedia'] = patch(auto.get('byMedia'), by_media)
    auto['phoneTotal'] = round(total, 2)
    auto['phoneByDay'] = {k: round(v, 2) for k, v in by_day.items()}
    auto['phoneByRegion'] = {k: round(v, 2) for k, v in by_region.items()}
    auto['phoneByMedia'] = {k: round(v, 2) for k, v in by_media.items()}
    return auto


def _append_dim_table(ws, title, rows, *, week_only=True, with_phones=False):
    if title:
        ws.append([title])
    if with_phones:
        ws.append([
            '項目', '本週來人', '佔本週來人%', '本週來電', '佔本週來電%',
            '本週成交', '佔本週成交%', '前期來人', '累計來人',
        ])
    else:
        ws.append([
            '項目', '前期累計(來人)', '本週小計(來人)', '目前累計(來人)',
            '佔本週來人%', '佔累計來人%',
            '前期成交', '本週成交', '累計成交',
            '佔本週成交%', '佔累計成交%',
        ])
    _style_header(ws, ws.max_row)
    total_row = None
    shown = 0
    for r in rows or []:
        if r.get('name') == '合計':
            total_row = r
            continue
        has_week = float(r.get('weekVisits') or 0) or float(r.get('weekPhones') or 0)
        if week_only and not has_week:
            continue
        if with_phones:
            ws.append([
                r.get('name'), r.get('weekVisits'), r.get('weekVisitPct'),
                r.get('weekPhones', 0), r.get('weekPhonePct', 0),
                r.get('weekDeals'), r.get('weekDealPct'),
                r.get('priorVisits'), r.get('cumVisits'),
            ])
            for col in (3, 5, 7):
                ws.cell(ws.max_row, col).number_format = '0.##"%"'
        else:
            ws.append([
                r.get('name'), r.get('priorVisits'), r.get('weekVisits'), r.get('cumVisits'),
                r.get('weekVisitPct'), r.get('cumVisitPct'),
                r.get('priorDeals'), r.get('weekDeals'), r.get('cumDeals'),
                r.get('weekDealPct'), r.get('cumDealPct'),
            ])
            for col in (5, 6, 10, 11):
                ws.cell(ws.max_row, col).number_format = '0.##"%"'
        shown += 1
    if total_row:
        if with_phones:
            ws.append([
                total_row.get('name'), total_row.get('weekVisits'), total_row.get('weekVisitPct'),
                total_row.get('weekPhones', 0), total_row.get('weekPhonePct', 0),
                total_row.get('weekDeals'), total_row.get('weekDealPct'),
                total_row.get('priorVisits'), total_row.get('cumVisits'),
            ])
            for col in (3, 5, 7):
                ws.cell(ws.max_row, col).number_format = '0.##"%"'
        else:
            ws.append([
                total_row.get('name'), total_row.get('priorVisits'), total_row.get('weekVisits'),
                total_row.get('cumVisits'), total_row.get('weekVisitPct'), total_row.get('cumVisitPct'),
                total_row.get('priorDeals'), total_row.get('weekDeals'), total_row.get('cumDeals'),
                total_row.get('weekDealPct'), total_row.get('cumDealPct'),
            ])
            for col in (5, 6, 10, 11):
                ws.cell(ws.max_row, col).number_format = '0.##"%"'
    if week_only and shown == 0 and not total_row:
        ws.append(['（本週無資料）'])
    ws.append([])


def _append_ppt_dim_rows(ws, rows):
    """PPT 摘要用：區域／媒體列，來人與來電同表。"""
    shown = 0
    total_row = None
    for r in rows or []:
        if r.get('name') == '合計':
            total_row = r
            continue
        visits = _num(r.get('weekVisits'))
        phones = _num(r.get('weekPhones'))
        deals = _num(r.get('weekDeals'))
        if not visits and not phones and not deals:
            continue
        ws.append([
            r.get('name'),
            visits,
            phones,
            deals,
            f"{r.get('weekVisitPct', 0)}%",
            f"{r.get('weekPhonePct', 0)}%",
        ])
        shown += 1
    if total_row:
        ws.append([
            total_row.get('name'),
            total_row.get('weekVisits', 0),
            total_row.get('weekPhones', 0),
            total_row.get('weekDeals', 0),
            f"{total_row.get('weekVisitPct', 0)}%",
            f"{total_row.get('weekPhonePct', 0)}%",
        ])
        for col in range(1, 7):
            ws.cell(ws.max_row, col).font = _font(bold=True)
    elif shown == 0:
        ws.append(['（本週無資料）', '', '', '', '', ''])


def _ppt_fill_row(ws, values, fill, end_col=None):
    ws.append(list(values))
    last = end_col or len(values)
    for col in range(1, last + 1):
        cell = ws.cell(ws.max_row, col)
        cell.fill = fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def _ppt_section_title(ws, title, fill, end_col=10):
    ws.append([title])
    cell = ws.cell(ws.max_row, 1)
    cell.font = _font(bold=True, size=12, color='1A4D7C')
    cell.fill = fill
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=end_col)


def _build_ppt_summary_sheet(ws, *, site_name, start, end, week_number, manual, auto):
    """對齊既有週報 PPT 首頁：來人成交｜累計請佣｜客況備註。"""
    title_fill = PatternFill('solid', fgColor='1A4D7C')
    section_fill = PatternFill('solid', fgColor='D6EAF8')
    soft_fill = PatternFill('solid', fgColor='F8FBFE')
    label_fill = PatternFill('solid', fgColor='EEF4FA')
    thin = Border(
        left=Side(style='thin', color='B0BEC5'),
        right=Side(style='thin', color='B0BEC5'),
        top=Side(style='thin', color='B0BEC5'),
        bottom=Side(style='thin', color='B0BEC5'),
    )

    t = auto.get('totals') or {}
    phone_sum = phone_total_from_manual(manual)
    deals = manual.get('deals') or {}
    signings = manual.get('signings') or {}
    purchases = manual.get('purchases') or {}
    deals_cum = manual.get('dealsCum') or {}
    signings_cum = manual.get('signingsCum') or {}
    purchases_cum = manual.get('purchasesCum') or {}
    c = manual.get('commission') or {}
    com = commission_summary(manual)
    matrix = auto.get('commissionMatrix') or {}
    inv_raw = manual.get('inventory') or {}

    visit_total = t.get('reportedTotal', t.get('total', 0))
    return_total = t.get('return', 0)
    cum_deal_u = _num(deals_cum.get('units')) or _num(c.get('sellableUnits')) or _num(inv_raw.get('soldUnits'))
    cum_deal_p = _num(deals_cum.get('parking')) or _num(c.get('sellableParking')) or _num(inv_raw.get('soldParking'))
    cum_deal_a = _num(deals_cum.get('amount')) or _num(c.get('sellableAmount')) or _num(inv_raw.get('soldAmount'))
    cum_sign_u = _num(signings_cum.get('units'))
    cum_sign_p = _num(signings_cum.get('parking'))
    cum_sign_a = _num(signings_cum.get('amount'))
    cum_buy_u = _num(purchases_cum.get('units'))
    cum_buy_p = _num(purchases_cum.get('parking'))
    cum_buy_a = _num(purchases_cum.get('amount'))
    sold_u = _num(c.get('sellableUnits')) or _num(inv_raw.get('soldUnits'))
    sold_p = _num(c.get('sellableParking')) or _num(inv_raw.get('soldParking'))
    sold_a = _num(c.get('sellableAmount')) or _num(inv_raw.get('soldAmount'))
    rate = COMMISSION_RATE_DEFAULT
    claimable_amt = _num(c.get('claimableAmount'))
    claimed_amt = _num(c.get('claimedAmount'))
    booked_amt = _num(c.get('bookedAmount'))
    retention_amt = _num(c.get('retentionAmount')) or com.get('retentionAmount')
    unclaimed_amt = _num(c.get('unclaimedAmount')) or com.get('unclaimedAmount')
    unclaimed_units = _num(c.get('unclaimedUnits')) or com.get('unclaimedUnits')
    unclaimed_parking = _num(c.get('unclaimedParking')) or com.get('unclaimedParking')
    claimable_units = _num(c.get('claimableUnits')) or cum_deal_u
    claimable_parking = _num(c.get('claimableParking')) or cum_deal_p
    claimed_units = _num(c.get('claimedUnits'))
    claimed_parking = _num(c.get('claimedParking'))

    # 優先用銷售總表矩陣（與請佣總覽同一套）
    m_claim = matrix.get('claimable') or {}
    m_claimed = matrix.get('claimed') or {}
    m_unclaimed = matrix.get('unclaimed') or {}
    if m_claim:
        claimable_units = _num(m_claim.get('units'), claimable_units)
        claimable_parking = _num(m_claim.get('parking'), claimable_parking)
        claimable_amt = _num(m_claim.get('claimable'), claimable_amt)
        retention_amt = _num(m_claim.get('retention'), retention_amt) or retention_amt
    if m_claimed:
        claimed_units = _num(m_claimed.get('units'), claimed_units)
        claimed_parking = _num(m_claimed.get('parking'), claimed_parking)
        claimed_amt = _num(m_claimed.get('payable'), claimed_amt) or claimed_amt
        claimed_total_485 = _num(m_claimed.get('claimable')) or round(claimed_amt + retention_amt, 4)
        claimed_retention = _num(m_claimed.get('retention'), retention_amt)
    else:
        claimed_total_485 = round(claimed_amt + retention_amt, 4) if (claimed_amt or retention_amt) else claimable_amt
        claimed_retention = retention_amt
    if m_unclaimed:
        unclaimed_units = _num(m_unclaimed.get('units'), unclaimed_units)
        unclaimed_parking = _num(m_unclaimed.get('parking'), unclaimed_parking)
        # PPT「未請佣金額」對的是 4.85% 可請，不是 97%
        unclaimed_amt = (
            _num(m_unclaimed.get('claimable'))
            or _num(m_unclaimed.get('payable'), unclaimed_amt)
            or unclaimed_amt
        )

    comm_sales_amount = round(claimable_amt / rate, 4) if claimable_amt else cum_deal_a
    unclaimed_sales = round(unclaimed_amt / rate, 4) if unclaimed_amt else 0
    currently_claimed = booked_amt if booked_amt else claimed_amt

    week_title = (
        f'第{_to_zh_int(safe_week_number(week_number, start))}週　'
        f'{start.year}/{start.month}/{start.day}～{end.year}/{end.month}/{end.day}'
    )
    cols = 11
    ws.append([week_title])
    ws['A1'].font = _font(bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.row_dimensions[1].height = 28

    ws.append([site_name, '', '', '', '貼到 PPT 首頁用（戶／車／萬已分欄）'])
    ws['A2'].font = _font(bold=True, size=12, color='1A4D7C')
    ws['E2'].font = _font(size=10, color='607D8B')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=cols)
    ws.append([])

    # —— 一、本週來人成交狀況（左來人｜中本週｜右累計）——
    _ppt_section_title(ws, '一、本週來人成交狀況', section_fill, cols)
    ws.append([
        '來人資訊', '數值', '單位',
        '本週', '戶', '車', '萬',
        '累計', '戶', '車', '萬',
    ])
    _style_header(ws, ws.max_row)
    for col in range(1, cols + 1):
        ws.cell(ws.max_row, col).border = thin

    visit_rows = [
        (
            '來人', visit_total, '組',
            '本週成交', deals.get('units'), deals.get('parking'), deals.get('amount'),
            '累計成交', cum_deal_u, cum_deal_p, cum_deal_a,
        ),
        (
            '來電', phone_sum, '通',
            '本週簽約', signings.get('units'), signings.get('parking'), signings.get('amount'),
            '累計簽約', cum_sign_u, cum_sign_p, cum_sign_a,
        ),
        (
            '回訪', return_total, '組',
            '本週實進', purchases.get('units'), purchases.get('parking'), purchases.get('amount'),
            '累計買進', cum_buy_u, cum_buy_p, cum_buy_a,
        ),
    ]
    for row in visit_rows:
        values = [
            row[0], _fmt_num(row[1]), row[2],
            row[3],
            _fmt_num(row[4]) if row[3] else '',
            _fmt_num(row[5]) if row[3] else '',
            _fmt_num(row[6]) if row[3] else '',
            row[7],
            _fmt_num(row[8]) if row[7] else '',
            _fmt_num(row[9]) if row[7] else '',
            _fmt_num(row[10]) if row[7] else '',
        ]
        _ppt_fill_row(ws, values, soft_fill, cols)
        for col in range(1, cols + 1):
            ws.cell(ws.max_row, col).border = thin
        for col in (1, 4, 8):
            ws.cell(ws.max_row, col).fill = label_fill
            ws.cell(ws.max_row, col).font = _font(bold=True)

    ws.append([])

    # —— 二、累積銷售／請佣（對齊 PPT 五列）——
    _ppt_section_title(ws, '二、累積銷售金額　戶數及可請佣金', section_fill, cols)
    ws.append(['項目', '戶', '車', '金額(萬)', '補充（可直接貼 PPT）', '', '', '', '', '', ''])
    _style_header(ws, ws.max_row)
    ws.merge_cells(start_row=ws.max_row, start_column=5, end_row=ws.max_row, end_column=cols)
    for col in range(1, cols + 1):
        ws.cell(ws.max_row, col).border = thin

    commission_rows = [
        ('累積銷售金額', sold_u, sold_p, sold_a, ''),
        ('請佣銷售金額', claimable_units, claimable_parking, comm_sales_amount, ''),
        ('可請佣戶數車位', claimable_units, claimable_parking, claimable_amt, ''),
        (
            '已請佣金戶數車位',
            claimed_units,
            claimed_parking,
            claimed_total_485,
            (
                f'總4.85% {_fmt_num(claimed_total_485)}　'
                f'月底已請 {_fmt_num(currently_claimed)}　'
                f'尾款未請 {_fmt_num(claimed_retention)}'
            ),
        ),
        (
            '未請佣金戶數車位',
            unclaimed_units,
            unclaimed_parking,
            unclaimed_amt,
            (
                f'未請銷售金額 {_fmt_num(unclaimed_sales)}　'
                f'未請佣金額 {_fmt_num(unclaimed_amt)}'
            ),
        ),
    ]
    for label, units, parking, amount, note in commission_rows:
        _ppt_fill_row(ws, [
            label, _fmt_num(units), _fmt_num(parking), _fmt_num(amount), note,
            '', '', '', '', '', '',
        ], soft_fill, cols)
        ws.merge_cells(start_row=ws.max_row, start_column=5, end_row=ws.max_row, end_column=cols)
        for col in range(1, cols + 1):
            ws.cell(ws.max_row, col).border = thin
        ws.cell(ws.max_row, 1).fill = label_fill
        ws.cell(ws.max_row, 1).font = _font(bold=True)

    ws.append([])

    # —— 三、本週客況（備註文字）——
    hope_n = len(auto.get('hopeCustomers') or [])
    new_n = t.get('new', 0)
    _ppt_section_title(ws, '三、本週客況（備註）', section_fill, cols)
    situation_lines = [
        f'1. 來人 {_fmt_num(visit_total)} 組（新客 {_fmt_num(new_n)}、回訪 {_fmt_num(return_total)}、有望 {_fmt_num(hope_n)}）',
        f'2. 來人區域：{_star_counts(auto.get("byRegion") or [])}',
        f'3. 來人媒體：{_star_counts(auto.get("byMedia") or [])}',
        f'4. 購屋型態：{_star_counts(auto.get("byPurpose") or [])}',
    ]
    if phone_sum:
        situation_lines.insert(
            1,
            f'   來電 {_fmt_num(phone_sum)} 通；來電區域：{_phone_star_counts(auto.get("phoneByRegion") or {})}',
        )
    if manual.get('reviewNotes'):
        situation_lines.append(f'成交檢討：{manual.get("reviewNotes")}')
    if manual.get('memo'):
        situation_lines.append(f'備註：{manual.get("memo")}')
    for line in situation_lines:
        ws.append([line])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=cols)
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(ws.max_row, 1).fill = soft_fill
        ws.row_dimensions[ws.max_row].height = 22

    widths = [18, 10, 8, 12, 10, 10, 12, 12, 10, 10, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_ppt_region_media_sheet(ws, auto):
    """PPT 第二頁素材：區域／媒體（來人＋來電）。"""
    section_fill = PatternFill('solid', fgColor='D6EAF8')
    ws.append(['區域／媒體（貼 PPT 客況或分析頁用）'])
    ws['A1'].font = _font(bold=True, size=14, color='1A4D7C')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.append([])
    ws.append(['區域', '本週來人', '本週來電', '本週成交', '佔來人%', '佔來電%'])
    _style_header(ws, ws.max_row)
    _append_ppt_dim_rows(ws, auto.get('byRegion') or [])
    ws.append([])
    ws.append(['媒體', '本週來人', '本週來電', '本週成交', '佔來人%', '佔來電%'])
    _style_header(ws, ws.max_row)
    _append_ppt_dim_rows(ws, auto.get('byMedia') or [])
    for col, width in enumerate([14, 12, 12, 12, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_weekly_excel(site_name: str, start, end, week_number, manual: dict, auto: dict) -> bytes:
    wb = Workbook()
    inv = inventory_summary(manual)
    com = commission_summary(manual)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    t = auto.get('totals') or {}
    p = auto.get('period') or {}
    phone_sum = phone_total_from_manual(manual)
    deals = manual.get('deals') or {}
    signings = manual.get('signings') or {}
    purchases = manual.get('purchases') or {}
    deals_cum = manual.get('dealsCum') or {}
    signings_cum = manual.get('signingsCum') or {}
    purchases_cum = manual.get('purchasesCum') or {}
    unreported = manual.get('unreported') or {}

    # —— PPT摘要（對齊既有週報投影片首頁）——
    ws = wb.active
    ws.title = 'PPT摘要'
    _build_ppt_summary_sheet(
        ws,
        site_name=site_name,
        start=start,
        end=end,
        week_number=week_number,
        manual=manual,
        auto=auto,
    )

    # —— PPT區域媒體（第二頁素材）——
    ws_dim = wb.create_sheet('PPT區域媒體')
    _build_ppt_region_media_sheet(ws_dim, auto)

    # —— 週報一覽（完整資料）——
    ws = wb.create_sheet('週報一覽')
    ws.append([f'{site_name}　第{week_number}週週報告（完整）'])
    ws['A1'].font = _font(bold=True, size=16, color='1A4D7C')
    ws.append([f'區間：{start.isoformat()} ～ {end.isoformat()}'])
    ws.append([])
    ws.append(['一、本週來人成交狀況'])
    ws.append(['項目', '數值', '項目', '數值'])
    _style_header(ws, ws.max_row)
    overview_pairs = [
        ('本週來人(組)', t.get('total', 0), '實際來人(組)', t.get('actualTotal', t.get('total', 0))),
        ('本週新客／回訪', f"{t.get('new', 0)} / {t.get('return', 0)}", '本週來電(通)', phone_sum),
        ('本週成交(戶/車/萬)', f"{deals.get('units', 0)}/{deals.get('parking', 0)}/{deals.get('amount', 0)}",
         '累計成交(戶/車/萬)', f"{deals_cum.get('units', 0)}/{deals_cum.get('parking', 0)}/{deals_cum.get('amount', 0)}"),
        ('本週簽約(戶/車/萬)', f"{signings.get('units', 0)}/{signings.get('parking', 0)}/{signings.get('amount', 0)}",
         '累計簽約(戶/車/萬)', f"{signings_cum.get('units', 0)}/{signings_cum.get('parking', 0)}/{signings_cum.get('amount', 0)}"),
        ('本週買進(戶/車/萬)', f"{purchases.get('units', 0)}/{purchases.get('parking', 0)}/{purchases.get('amount', 0)}",
         '累計買進(戶/車/萬)', f"{purchases_cum.get('units', 0)}/{purchases_cum.get('parking', 0)}/{purchases_cum.get('amount', 0)}"),
        ('未報(戶/車/萬)', f"{unreported.get('units', 0)}/{unreported.get('parking', 0)}/{unreported.get('amount', 0)}",
         '客資成交筆數', t.get('deal', 0)),
        ('本月來人／成交', f"{(p.get('month') or {}).get('visits', 0)} / {(p.get('month') or {}).get('deals', 0)}",
         '本年來人／成交', f"{(p.get('year') or {}).get('visits', 0)} / {(p.get('year') or {}).get('deals', 0)}"),
        ('去化率(戶/底價%)', f"{inv['unitRate']}% / {inv['basePriceRate']}%",
         '未售底價(萬)', inv['remainBasePrice']),
        ('可請佣-已請(萬)', com['unclaimedAmount'], '已入帳(萬)', com['bookedAmount']),
        ('預計本月可請(戶/車/萬)',
         f"{com['nextMonthUnits']}/{com['nextMonthParking']}/{com['nextMonthAmount']}",
         '剩餘戶／車', f"{inv['remainUnits']} / {inv['remainParking']}"),
    ]
    for a, b, c, d in overview_pairs:
        ws.append([a, b, c, d])
    ws.append([])
    ws.append(['二、成交檢討'])
    ws.append([manual.get('reviewNotes') or ''])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    ws.append(['三、區域個案分析'])
    ws.append([manual.get('competitorNotes') or ''])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    ws.append(['備註', manual.get('memo') or ''])
    ws.append([])
    ws.append(['四、銷售成交比'])
    ws.append(['銷售人員', '累計接待', '累計成交', '成交比', '成交金額', '退戶組數', '退戶金額',
               '本週接待', '本週成交', '本週金額'])
    _style_header(ws, ws.max_row)
    for row in auto.get('conversion') or []:
        ws.append([
            row.get('name'), row.get('visits'), row.get('deals'), row.get('rate'),
            row.get('amount'), row.get('refunds'), row.get('refundAmount'),
            row.get('weekVisits'), row.get('weekDeals'), row.get('weekAmount'),
        ])
    ws.append([])
    ws.append(['五、區域分析（來人／來電並陳）'])
    _append_dim_table(ws, '', auto.get('byRegion') or [], with_phones=True)
    ws.append(['六、媒體分析（來人／來電並陳）'])
    _append_dim_table(ws, '', auto.get('byMedia') or [], with_phones=True)
    ws.append(['七、職業／年齡分析（新客）'])
    _append_dim_table(ws, '職業（新客）', auto.get('byOccupation') or [])
    _append_dim_table(ws, '年齡（新客）', auto.get('byAge') or [])
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 22

    # —— 每日 ——
    ws = wb.create_sheet('每日統計')
    ws.append(['日期', '星期', '新客', '回訪', '合計', '成交', '來電', '天氣'])
    _style_header(ws, ws.max_row)
    days = manual.get('days') or []
    phone_by_day = auto.get('phoneByDay') or {}
    for i, d in enumerate(auto.get('byDay') or []):
        m = days[i] if i < len(days) else {}
        day_phones = phone_by_day.get(d.get('date'), m.get('phoneCalls', 0))
        ws.append([
            d.get('date'), d.get('weekday'), d.get('new'), d.get('return'),
            d.get('total'), d.get('deal'), day_phones, m.get('weather', ''),
        ])

    # —— 來電明細 ——
    ws = wb.create_sheet('來電明細')
    ws.append(['日期', '區域', '媒體', '通數'])
    _style_header(ws)
    phone_detail = normalize_phone_calls_detail(manual.get('phoneCallsDetail'))
    if phone_detail:
        for item in phone_detail:
            ws.append([item.get('date'), item.get('region'), item.get('media'), item.get('count')])
    else:
        ws.append(['（尚無來電明細）'])

    # —— 成交比 ——
    ws = wb.create_sheet('成交比')
    ws.append(['銷售人員', '累計接待', '累計成交', '成交比', '成交金額', '退戶組數', '退戶金額',
               '本週接待', '本週成交', '本週金額', '本週退戶'])
    _style_header(ws)
    for row in auto.get('conversion') or []:
        ws.append([
            row.get('name'), row.get('visits'), row.get('deals'), row.get('rate'),
            row.get('amount'), row.get('refunds'), row.get('refundAmount'),
            row.get('weekVisits'), row.get('weekDeals'), row.get('weekAmount'), row.get('weekRefunds'),
        ])

    # —— 去化／請佣 ——
    ws = wb.create_sheet('去化與請佣')
    inv_m = manual.get('inventory') or {}
    ws.append(['去化項目', '數值'])
    _style_header(ws)
    for label, key in [
        ('總戶數', 'totalUnits'), ('已售戶數', 'soldUnits'),
        ('總車位', 'totalParking'), ('已售車位', 'soldParking'),
        ('總金額／總底價(萬)', 'totalAmount'), ('已售成交價(萬)', 'soldAmount'),
        ('已售底價(萬)', 'soldBasePrice'),
        ('住宅總戶', 'residentialTotal'), ('住宅已售', 'residentialSold'),
        ('事務所總戶', 'officeTotal'), ('事務所已售', 'officeSold'),
        ('店鋪總戶', 'shopTotal'), ('店鋪已售', 'shopSold'),
        ('店面總戶', 'storefrontTotal'), ('店面已售', 'storefrontSold'),
    ]:
        ws.append([label, inv_m.get(key, 0)])
    ws.append(['戶數去化率%', inv['unitRate']])
    ws.cell(ws.max_row, 2).number_format = '0.##"%"'
    ws.append(['底價去化率%', inv['basePriceRate']])
    ws.cell(ws.max_row, 2).number_format = '0.##"%"'
    ws.append(['未售底價(萬)', inv['remainBasePrice']])
    ws.append([])
    ws.append(['請佣項目', '數值'])
    _style_header(ws, ws.max_row)
    c = manual.get('commission') or {}
    matrix = auto.get('commissionMatrix') or {}
    if matrix:
        ws.append(['【可請／已請／未請矩陣】'])
        cards_labels = matrix.get('labels') or {}
        rate_h = cards_labels.get('claimable') or '佣金(萬)'
        ret_h = cards_labels.get('retention') or '保留(萬)'
        pay_h = cards_labels.get('payable') or '可請(萬)'
        ws.append(['區塊', '戶／車', rate_h, ret_h, pay_h])
        _style_header(ws, ws.max_row)
        for key, label in (
            ('claimable', '可請總金額'),
            ('claimed', '已請款金額'),
            ('unclaimed', '未請款總金額'),
            ('forecast', '預計本月可請'),
        ):
            b = matrix.get(key) or {}
            ws.append([
                label,
                f"{b.get('units', 0)}戶／{b.get('parking', 0)}車",
                b.get('claimable', 0),
                b.get('retention', 0),
                b.get('payable', 0),
            ])
        ws.append([])
    for label, key in [
        ('累積銷售戶數', 'sellableUnits'),
        ('累積銷售車位', 'sellableParking'),
        ('累積銷售金額(萬)', 'sellableAmount'),
        ('可請佣金額(萬)', 'claimableAmount'),
        ('已請佣金額(萬)', 'claimedAmount'),
        ('已入帳金額(萬)', 'bookedAmount'),
        ('可請佣戶數', 'claimableUnits'),
        ('已請佣戶數', 'claimedUnits'),
        ('可請佣車位', 'claimableParking'),
        ('已請佣車位', 'claimedParking'),
        ('未請佣戶數', 'unclaimedUnits'),
        ('未請佣車位', 'unclaimedParking'),
        ('預計本月可請戶數', 'nextMonthUnits'),
        ('預計本月可請車位', 'nextMonthParking'),
        ('預計本月可請金額(萬)', 'nextMonthAmount'),
    ]:
        ws.append([label, c.get(key, 0)])
    ws.append(['未請佣金額(萬)', com['unclaimedAmount']])
    if not c.get('unclaimedUnits'):
        ws.append(['未請佣戶數', com['unclaimedUnits']])
        ws.append(['未請佣車位', com['unclaimedParking']])

    # —— 維度詳表 ——
    ws = wb.create_sheet('區域媒體職業年齡')
    for title, rows, with_phones in [
        ('區域（新客＋來電）', auto.get('byRegion'), True),
        ('媒體（新客＋來電）', auto.get('byMedia'), True),
        ('來源（新客）', auto.get('bySource'), False),
        ('購屋型態（新客）', auto.get('byPurpose'), False),
        ('職業（新客）', auto.get('byOccupation'), False),
        ('年齡（新客）', auto.get('byAge'), False),
    ]:
        _append_dim_table(ws, title, rows, week_only=False, with_phones=with_phones)

    # —— 客況（僅納入週報者）——
    ws = wb.create_sheet('本週客況')
    ws.append(['納入', '日期', '類型', '姓名', '區域', '媒體', '職業', '年齡',
               '介紹戶別', '洽談內容', '未購因素', '誠意度', '銷售1', '銷售2', '共同', '成交', '成交金額', '退戶'])
    _style_header(ws)
    for v in auto.get('visitors') or []:
        ws.append([
            '是', v.get('date'), v.get('visitType'), v.get('customerName'),
            v.get('region'), v.get('media'), v.get('occupation'), v.get('age'),
            v.get('introUnit'), v.get('discussion'), v.get('notPurchasedReason'),
            v.get('sincerity'), v.get('salesperson1'), v.get('salesperson2'),
            '是' if v.get('isCoManaged') else '否',
            '是' if v.get('isDeal') else '否', v.get('dealAmount'),
            '是' if v.get('isRefund') else '否',
        ])
    for cell in ws['J'] + ws['K']:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 22

    ws = wb.create_sheet('回訪與有望客')
    ws.append(['【回訪】'])
    ws.append(['日期', '姓名', '電話', '區域', '媒體', '誠意度', '銷售'])
    _style_header(ws, ws.max_row)
    for v in auto.get('returnVisits') or []:
        ws.append([v.get('date'), v.get('customerName'), v.get('phone'), v.get('region'),
                   v.get('media'), v.get('sincerity'), v.get('salesperson1')])
    ws.append([])
    ws.append(['【有望客】'])
    ws.append(['日期', '類型', '姓名', '電話', '區域', '誠意度', '銷售'])
    _style_header(ws, ws.max_row)
    for v in auto.get('hopeCustomers') or []:
        ws.append([v.get('date'), v.get('visitType'), v.get('customerName'), v.get('phone'),
                   v.get('region'), v.get('sincerity'), v.get('salesperson1')])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80), max_col=min(sheet.max_column, 18)):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin

    _finalize_workbook_fonts(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
